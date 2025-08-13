import streamlit as st
import pandas as pd
import re
from typing import List, Dict, Any, Tuple
from io import BytesIO
import openpyxl
import json
import networkx as nx 
from concurrent.futures import ThreadPoolExecutor, as_completed 
from components.AVA_Assistant import *
from components.functions import *
import plotly.graph_objects as go
from pyvis.network import Network
from typing import Dict, List, Any
import streamlit as st
import pandas as pd

class DataLineageAnalyzer(AVAAssistant):
    """Data lineage analyzer with parallel processing."""
    
    def __init__(self):
        super().__init__()
        self.graph = nx.DiGraph()
        self.lineage_cache = {}
        self.node_colors = {
            'source': '#4CAF50',
            'transform': '#2196F3',
            'output': '#FF9800',
            'variable': '#9C27B0',
            'operation': '#00BCD4'
        }
        # Attributes for processing
        self.compiled_code = ""
        self.file_order = []
        self.code_batches = []
        self.functions_macros_map = {}
    
    def _compile_code_in_logical_order(self, code_files: list) -> None:
        """Compile all code in logical execution order."""
        
        def get_sort_key(filename: str) -> tuple:
            """Generate sort key for logical ordering."""
            name = filename.lower()
            base_name = name.rsplit('.', 1)[0]
            
            # Pattern 1: Numbers (1, 2, 3, 10, 11)
            if re.match(r'^\d+$', base_name):
                return (1, int(base_name), '')
            
            # Pattern 2: Number + letter (1a, 1b, 2a)
            match = re.match(r'^(\d+)([a-z]+)$', base_name)
            if match:
                return (2, int(match.group(1)), match.group(2))
            
            # Pattern 3: Single letters (a, b, c)
            if re.match(r'^[a-z]$', base_name):
                return (3, 0, base_name)
            
            # Pattern 4: Words with numbers (script1, file2, step3)
            match = re.search(r'(\d+)', base_name)
            if match:
                prefix = base_name[:match.start()]
                number = int(match.group(1))
                suffix = base_name[match.end():]
                return (4, number, prefix + suffix)
            
            # Pattern 5: Common ordering words
            order_words = ['main', 'init', 'setup', 'process', 'transform', 'analyze', 'final', 'output']
            for i, word in enumerate(order_words):
                if word in base_name:
                    return (5, i, base_name)
            
            # Pattern 6: Everything else (alphabetical)
            return (6, 0, base_name)
        
        # Sort files logically
        sorted_files = sorted(code_files, key=lambda f: get_sort_key(f.name))
        
        self.compiled_code = ""
        self.file_order = []
        
        for i, file in enumerate(sorted_files, 1):
            content = file.read().decode('utf-8')
            
            # Add to compiled code
            self.compiled_code += f"\n\n# ========== FILE {i}: {file.name} ==========\n"
            self.compiled_code += content
            
            self.file_order.append(file.name)
            file.seek(0)  # Reset for later use
        
        # Create batches of 200 lines each
        self._create_code_batches()
    
    def _create_code_batches(self) -> None:
        """Create batches of code for parallel processing."""
        lines = self.compiled_code.split('\n')
        batch_size = 200
        
        self.code_batches = []
        
        for i in range(0, len(lines), batch_size):
            batch_lines = lines[i:i + batch_size]
            batch_content = '\n'.join(batch_lines)
            
            self.code_batches.append({
                'batch_id': len(self.code_batches) + 1,
                'start_line': i + 1,
                'end_line': i + len(batch_lines),
                'content': batch_content,
                'line_count': len(batch_lines)
            })
    
    def _extract_functions_and_macros(self) -> None:
        """Extract functions and macros for context."""
        
        # Patterns for different languages
        patterns = {
            'python_function': r'def\s+(\w+)\s*\([^)]*\):',
            'python_class': r'class\s+(\w+).*:',
            'sas_macro': r'%macro\s+(\w+)',
            'sas_function': r'proc\s+(\w+)',
            'r_function': r'(\w+)\s*<-\s*function\s*\(',
            'sql_function': r'CREATE\s+(?:OR\s+REPLACE\s+)?FUNCTION\s+(\w+)',
            'sql_procedure': r'CREATE\s+(?:OR\s+REPLACE\s+)?PROCEDURE\s+(\w+)'
        }
        
        for pattern_name, pattern in patterns.items():
            matches = re.findall(pattern, self.compiled_code, re.IGNORECASE | re.MULTILINE)
            if matches:
                self.functions_macros_map[pattern_name] = matches
    
    def _simple_ai_variable_extraction(self, code_files: list) -> Dict[str, Any]:
        """AI-powered variable extraction with parallel processing"""
        
        # Step 1: Compile code in logical order
        st.write("📋 **Compiling code in logical order...**")
        self._compile_code_in_logical_order(code_files)
        
        # Step 2: Extract functions/macros for context
        st.write("🔧 **Extracting functions and macros...**")
        self._extract_functions_and_macros()
        
        # Step 3: Parallel AI variable discovery
        st.write("🤖 **AI discovering variables in parallel batches...**")
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Process batches in parallel 
        all_variables = {
            'input_variables': [],
            'created_variables': [],
            'intermediate_variables': [],
            'all_variables': []
        }
        
        def process_single_batch(batch_info):
            """Process a single batch for variable discovery """
            
            batch_id = batch_info['batch_id']
            content = batch_info['content']
            
            # Include function context
            function_context = ""
            if self.functions_macros_map:
                function_context = f"\n\nFUNCTIONS/MACROS AVAILABLE: {json.dumps(self.functions_macros_map)}"
            
            # Ultra-strict prompt for variable discovery
            prompt = f"""ANALYZE CODE BATCH {batch_id} FOR VARIABLE/COLUMN NAMES. RESPOND ONLY WITH JSON.

CODE BATCH:
{content}{function_context}

FIND VARIABLES/COLUMN NAMES FROM THE CODE BATCH BY CATEGORY. 
CRITICAL: ONLY IDENTIFY DATA COLUMNS AS VARIABLES

R: ONLY columns that are part of dataframes/datasets
✓ data$new_column <- data$old_column * 2
✓ mutate(new_var = old_var + 1)
no constant_value <- 0.05

Python: ONLY columns that are part of pandas DataFrames
✓ df['new_column'] = df['old_column'] * 2
✓ df.assign(new_var = df.old_var + 1)
no interest_rate = 0.05

SAS: ONLY variables in DATA steps that are dataset columns
✓ new_var = existing_var * 0.1; (in data steps)
✓ PROC SQL: SELECT new_var = old_var * 2
no %let macro_var = 2023;

IGNORE THESE COMPLETELY:
- Constants/literals (numbers, strings)
- Macro variables (%let, %global)
- File paths and filenames
- Loop counters (i, j, k)
- Configuration variables
- Function parameters
- Temporary calculations not assigned to datasets


STRICTLY RESPOND WITH ONLY THIS JSON:

{{
  "input_variables": ["var1", "var2"],
  "created_variables": ["var3", "var4"], 
  "intermediate_variables": ["var5", "var6"]
}}

CATEGORIES:
- INPUT: Variables/Columns assumed to be present in the data that are already EXISTING
- CREATED: New variables created through calculations/assignments 
- INTERMEDIATE: Variables modified/transformed from EXISTING/INPUT ones
CRITICAL - DO NOT MISS ANY VARIABLES FROM List in batch

JSON RESPONSE:"""
            
            try:
                response = self.send_message(prompt, {
                    "type": "batch_variable_discovery",
                    "temperature": 0.0,
                    "max_tokens": 60000,
                    "batch_id": batch_id
                })
                
                if response:
                    return self._parse_ai_variable_response(response, batch_id)
                else:
                    st.warning(f"⚠️ No AI response for batch {batch_id}")
                    return {'input_variables': [], 'created_variables': [], 'intermediate_variables': []}
                    
            except Exception as e:
                st.warning(f"⚠️ AI error for batch {batch_id}: {str(e)}")
                return {'input_variables': [], 'created_variables': [], 'intermediate_variables': []}
        
        # Execute parallel processing
        status_text.text("🚀 Processing batches in parallel...")
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            # Submit all batch jobs
            future_to_batch = {
                executor.submit(process_single_batch, batch): batch['batch_id'] 
                for batch in self.code_batches
            }
            
            completed_batches = 0
            
            # Collect results as they complete
            for future in as_completed(future_to_batch):
                try:
                    batch_results = future.result()
                    
                    # Merge results
                    for category in ['input_variables', 'created_variables', 'intermediate_variables']:
                        all_variables[category].extend(batch_results.get(category, []))
                    
                    # Update progress
                    completed_batches += 1
                    progress = completed_batches / len(self.code_batches)
                    progress_bar.progress(progress)
                    status_text.text(f"Completed {completed_batches}/{len(self.code_batches)} batches")
                    
                except Exception as e:
                    st.error(f"Error in batch processing: {str(e)}")
        
        # Remove duplicates and create all_variables list
        for category in ['input_variables', 'created_variables', 'intermediate_variables']:
            all_variables[category] = list(set(all_variables[category]))
            all_variables['all_variables'].extend(all_variables[category])
        
        all_variables['all_variables'] = list(set(all_variables['all_variables']))
        
        progress_bar.empty()
        status_text.empty()
        
        return all_variables
    
    def _parse_ai_variable_response(self, response: str, batch_id: int) -> Dict[str, List[str]]:
        """Parse AI response for variables - robust JSON parsing."""
        
        variables = {'input_variables': [], 'created_variables': [], 'intermediate_variables': []}
        
        # Multiple parsing attempts with progressive cleaning
        for attempt in range(3):
            try:
                clean_response = response.strip()
                
                if attempt == 1:
                    # Remove markdown formatting
                    clean_response = re.sub(r'```\w*', '', clean_response)
                    clean_response = clean_response.replace('```', '')
                elif attempt == 2:
                    # Extract JSON block
                    json_match = re.search(r'\{.*\}', clean_response, re.DOTALL)
                    if json_match:
                        clean_response = json_match.group()
                
                # Parse JSON
                data = json.loads(clean_response)
                
                # Extract variables by category
                for category in ['input_variables', 'created_variables', 'intermediate_variables']:
                    if category in data and isinstance(data[category], list):
                        # Clean variable names
                        clean_vars = []
                        for var in data[category]:
                            if isinstance(var, str) and var.strip():
                                clean_vars.append(var.strip())
                        variables[category] = clean_vars
                
                st.write(f"✅ Batch {batch_id}: AI found {sum(len(v) for v in variables.values())} variables")
                return variables
                
            except json.JSONDecodeError:
                if attempt == 2:  # Last attempt
                    st.warning(f"⚠️ AI response not valid JSON for batch {batch_id}")
                continue
        
        return variables
    

    def generate_detailed_variable_lineage(self, code_files: list, selected_variables: List[str], all_variables: Dict[str, Any]) -> Dict[str, Any]:
        """Generate detailed lineage for selected variables using AI analysis"""
        
        if not selected_variables:
            return {}
        
        st.write(f"🗺️ **Generating detailed lineage for {len(selected_variables)} variables...**")
        
        # Group variables into batches 
        batch_size = 1
        variable_batches = []
        
        for i in range(0, len(selected_variables), batch_size):
            batch = selected_variables[i:i + batch_size]
            variable_batches.append({
                'batch_id': len(variable_batches) + 1,
                'variables': batch
            })
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        detailed_lineage = {}
        
        def process_lineage_batch(batch_info):
            """Process a batch of variables for lineage analysis"""
            
            batch_id = batch_info['batch_id']
            variables = batch_info['variables']
            
            # Include function context
            function_context = ""
            if self.functions_macros_map:
                function_context = f"\n\nFUNCTIONS/MACROS AVAILABLE: {json.dumps(self.functions_macros_map)}"
            
            # Create variable analysis prompt
            variables_str = ", ".join(variables)
            
            # Combine all code
            all_code = ""
            file_info = []
            
            for file in code_files:
                content = file.read().decode('utf-8')
                file_type = file.name.split('.')[-1].upper()
                all_code += f"\n\n=== FILE: {file.name} ({file_type}) ===\n{content}"
                file_info.append(file.name)
                file.seek(0)
            
            all_code = ""
            for file in code_files:
                content = file.read().decode('utf-8')
                all_code += f"\n{content}\n"
                file.seek(0)

            detailed_lineage = {}
            
            for variable in selected_variables:
                st.write(f"🔍 Analyzing {variable}...")
                
                # Search for variable assignment lines
                variable_lines = []
                code_lines = all_code.split('\n')
                
                for i, line in enumerate(code_lines, 1):
                    line_clean = line.strip()
                    # Check if variable appears before = sign (meaning it's being assigned)
                    if variable in line_clean and '=' in line_clean:
                        equals_pos = line_clean.find('=')
                        variable_pos = line_clean.find(variable)
                        if variable_pos < equals_pos and variable_pos >= 0:
                            variable_lines.append(f"Line {i}: {line_clean}")
                
                # Fallback search if no assignment lines found
                if not variable_lines:
                    for i, line in enumerate(code_lines, 1):
                        if variable in line.strip():
                            variable_lines.append(f"Line {i}: {line.strip()}")
                
                # Prepare code context for AI
                if variable_lines:
                    specific_code = "\n".join([line.split(": ", 1)[1] for line in variable_lines if ": " in line])
                else:
                    specific_code = all_code
            
            lineage_prompt = f"""ANALYZE LINEAGE FOR ALL OF THESE VARIABLES. RESPOND ONLY WITH JSON.

VARIABLES TO ANALYZE: {variables_str}
    CODE:
    {all_code}

FOR EACH VARIABLE, PROVIDE DETAILED LINEAGE. DO NOT MISS ANY VARIABLE.
RESPOND WITH ONLY THIS JSON:

{{
  "lineage_results": [
    {{
      "variable": "variable_name",
      "category": "input|created|intermediate",
      "source_variables": ["source_var1", "source_var2"], 
      "dataset": ["dataset1"],
      "calculation": "exact_formula_or_logic",
      "description": "business_explanation",
      "business_purpose": "why_this_variable_exists",
      "code_location": "file_and_line_reference",
      "detailed_steps": ["step1", "step2", "step3"]
    }}
  ]
}}

FIND:
- Exact and detailed calculation formulas
- Source variables used
- Dataset: ASSIGNMENT RULES (CRITICAL - FOLLOW EXACTLY):
Rule 1: IF variable "category" is "input" → "dataset" field MUST be INPUT dataset name
- Look for: pd.read_csv, SET dataset, FROM table, IMPORT commands
- Example: "customer_data.csv", "sales_table", "input_file.xlsx"

Rule 2: IF variable "category" is "created" OR "intermediate" → "dataset" field MUST be OUTPUT dataset name  
- Look for: to_csv, DATA output, CREATE TABLE, EXPORT commands
- Example: "results.csv", "output_table", "final_report.xlsx"

Rule 3: IF no clear dataset found → Use "N/A" but NEVER mix input/output categories

Rule 4: There can be a single input/output dataset for a variable

- Business purpose explanation
- Step-by-step transformation logic
- If there are source_variables found for the variable being processed, then try to get all the details for all the 
    source_variables as well along with their exact formulae.
- Do not miss any of the 3 variable from the list

JSON RESPONSE:"""
            
            try:
                response = self.send_message(lineage_prompt, {
                    "type": "batch_lineage_analysis",
                    "temperature": 0.0,
                    "max_tokens": 60000,
                    "batch_id": batch_id
                })
                
                if response:
                    return self._parse_ai_lineage_response(response, batch_id)
                else:
                    st.warning(f"⚠️ No AI response for lineage batch {batch_id}")
                    return {}
                    
            except Exception as e:
                st.warning(f"⚠️ AI error in lineage batch {batch_id}: {str(e)}")
                return {}
        
        # Process lineage batches in parallel
        status_text.text("🚀 Processing lineage batches in parallel...")
        
        with ThreadPoolExecutor(max_workers=1) as executor:
            future_to_batch = {
                executor.submit(process_lineage_batch, batch): batch['batch_id']
                for batch in variable_batches
            }
            
            completed_batches = 0
            
            for future in as_completed(future_to_batch):
                try:
                    batch_results = future.result()
                    detailed_lineage.update(batch_results)
                    
                    completed_batches += 1
                    progress = completed_batches / len(variable_batches)
                    progress_bar.progress(progress)
                    status_text.text(f"Analyzed {completed_batches}/{len(variable_batches)} lineage batches")
                    
                except Exception as e:
                    st.error(f"Error in lineage batch processing: {str(e)}")
        
        progress_bar.empty()
        status_text.empty()
        
        st.success(f"✅ AI generated detailed lineage for {len(detailed_lineage)} variables")
        
        return detailed_lineage
    
    def _parse_ai_lineage_response(self, response: str, batch_id: int) -> Dict[str, Any]:
        """Parse AI lineage response - robust JSON parsing."""
        
        lineage_results = {}
        
        # Multiple parsing attempts
        for attempt in range(3):
            try:
                clean_response = response.strip()
                
                if attempt == 1:
                    # Remove markdown
                    clean_response = re.sub(r'```\w*', '', clean_response)
                    clean_response = clean_response.replace('```', '')
                elif attempt == 2:
                    # Extract JSON
                    json_match = re.search(r'\{.*\}', clean_response, re.DOTALL)
                    if json_match:
                        clean_response = json_match.group()
                
                data = json.loads(clean_response)
                
                if 'lineage_results' in data and isinstance(data['lineage_results'], list):
                    for item in data['lineage_results']:
                        if isinstance(item, dict) and 'variable' in item:
                            variable = item['variable']
                            lineage_results[variable] = {
                                'variable': variable,
                                'category': item.get('category', 'unknown'),
                                'source_variables': item.get('source_variables', []),
                                'dataset': item.get('dataset', []),
                                'calculation': item.get('calculation', ''),
                                'description': item.get('description', ''),
                                'business_purpose': item.get('business_purpose', ''),
                                'code_location': item.get('code_location', ''),
                                'detailed_steps': item.get('detailed_steps', [])
                            }
                    
                    st.write(f"✅ Lineage batch {batch_id}: AI analyzed {len(lineage_results)} variables")
                    return lineage_results
                    
            except json.JSONDecodeError:
                if attempt == 2:
                    st.warning(f"⚠️ AI lineage response not valid JSON for batch {batch_id}")
                continue
        
        return lineage_results
    
    def generate_full_data_lineage(self, all_variables, code_files: list, data_files: list = None) -> Dict[str, Any]:
        """Generate comprehensive data lineage."""
        
        try:
            st.write("🤖 **Starting AI-powered full data lineage analysis...**")
            
            if not all_variables.get('all_variables'):
                return {
                    'error': 'No variables found by AI',
                    'lineage_table': [],
                    'input_datasets': [],
                    'output_datasets': []
                }
            
            # Step 1: Generate lineage for all discovered variables
            all_discovered_vars = all_variables['all_variables']
            
            detailed_lineage = self.generate_detailed_variable_lineage(
                code_files, all_discovered_vars, all_variables
            )
            
            # Step 2: Convert to table format
            lineage_table = []
            for sr_no, (variable, info) in enumerate(detailed_lineage.items(), 1):
                lineage_table.append({
                    'sr_no': sr_no,
                    'variable_name': variable,
                    'variable_type': info.get('category', 'Unknown').title(),
                    'dataset': info.get('dataset', 'N/A'),
                    'variable_source': ', '.join(info.get('source_variables', [])) if info.get('source_variables') else 'N/A',
                    'calculation_methodology': info.get('calculation', 'N/A'),
                    'explanation': info.get('description', 'N/A'),
                    'line_no': info.get('code_location', 'N/A')
                })
            
            # Step 3: Dataset extraction
            input_datasets, output_datasets = self._ai_extract_datasets(code_files)
            
            return {
                'lineage_table': lineage_table,
                'input_datasets': input_datasets,
                'output_datasets': output_datasets,
                'total_variables': len(lineage_table),
                'discovered_variables': len(all_variables['all_variables']),
                'processing_method': "ai_only_parallel_approach",
                'file_order': self.file_order,
                'functions_macros': self.functions_macros_map
            }
            
        except Exception as e:
            st.error(f"❌ Error in AI lineage analysis: {str(e)}")
            return {
                'error': str(e),
                'lineage_table': [],
                'input_datasets': [],
                'output_datasets': []
            }
    
    def _ai_extract_datasets(self, code_files:list) -> Tuple[List[str], List[str]]:
        """Extract input and output datasets using multiple robust methods."""
        
        input_datasets = []
        output_datasets = []
        
        # Combine all code
        all_code = ""
        file_info = []
        
        for file in code_files:
            content = file.read().decode('utf-8')
            file_type = file.name.split('.')[-1].upper()
            all_code += f"\n\n=== FILE: {file.name} ({file_type}) ===\n{content}"
            file_info.append(file.name)
            file.seek(0)
        
        all_code = ""
        for file in code_files:
            content = file.read().decode('utf-8')
            all_code += f"\n{content}\n"
            file.seek(0)
                
        # Method 1: Try AI first
        try:
            dataset_prompt = f"""Find dataset names in this code. Return simple lists.
    CODE:
    {all_code}

    Find INPUT datasets (files being read) and OUTPUT datasets (files being written).
    Just list the dataset names, one per line:

    INPUT DATASETS:
    [list here]

    OUTPUT DATASETS:  
    [list here]"""
            
            response = self.send_message(dataset_prompt, {
                "type": "simple_dataset_extraction",
                "temperature": 0.0,
                "max_tokens": 60000
            })
            
            if response:
                ai_input, ai_output = self._simple_parse_ai_response(response)
                input_datasets.extend(ai_input)
                output_datasets.extend(ai_output)
                
        except Exception as e:
            st.warning(f"⚠️ AI extraction failed: {str(e)}")
        
        # Method 2: Simple regex patterns (always try this as backup)
        try:
            regex_input, regex_output = self._regex_extract_datasets()
            input_datasets.extend(regex_input)
            output_datasets.extend(regex_output)
            
        except Exception as e:
            st.warning(f"⚠️ Regex extraction failed: {str(e)}")
        
        # Method 3: Keyword search (most basic fallback)
        try:
            keyword_input, keyword_output = self._keyword_extract_datasets()
            input_datasets.extend(keyword_input)
            output_datasets.extend(keyword_output)
            
        except Exception as e:
            st.warning(f"⚠️ Keyword extraction failed: {str(e)}")
        
        # Clean up and deduplicate (but keep names even if not perfectly clean)
        input_datasets = self._gentle_clean_dataset_list(input_datasets)
        output_datasets = self._gentle_clean_dataset_list(output_datasets)
        
        st.info(f"📊 Found {len(input_datasets)} input datasets, {len(output_datasets)} output datasets")
        
        return input_datasets, output_datasets

    def _simple_parse_ai_response(self, response: str) -> Tuple[List[str], List[str]]:
        """Simple text parsing - don't try to be too clever."""
        
        input_datasets = []
        output_datasets = []
        
        lines = response.split('\n')
        current_section = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if 'INPUT' in line.upper():
                current_section = 'input'
            elif 'OUTPUT' in line.upper():
                current_section = 'output'
            elif current_section == 'input' and line not in ['INPUT DATASETS:', 'INPUT:']:
                input_datasets.append(line)
            elif current_section == 'output' and line not in ['OUTPUT DATASETS:', 'OUTPUT:']:
                output_datasets.append(line)
        
        return input_datasets, output_datasets

    def _regex_extract_datasets(self) -> Tuple[List[str], List[str]]:
        """Extract datasets using simple regex patterns."""
        
        input_datasets = []
        output_datasets = []
        
        # Simple patterns for common operations
        input_patterns = [
            r'read_csv\([\'"]([^\'"\)]+)[\'"]',
            r'read_excel\([\'"]([^\'"\)]+)[\'"]',
            r'set\s+([a-zA-Z_][a-zA-Z0-9_.]*)',
            r'from\s+([a-zA-Z_][a-zA-Z0-9_.]*)',
        ]
        
        output_patterns = [
            r'to_csv\([\'"]([^\'"\)]+)[\'"]',
            r'to_excel\([\'"]([^\'"\)]+)[\'"]',
            r'data\s+([a-zA-Z_][a-zA-Z0-9_.]*)',
            r'create\s+table\s+([a-zA-Z_][a-zA-Z0-9_.]*)',
        ]
        
        # Search in code
        for pattern in input_patterns:
            matches = re.findall(pattern, self.compiled_code, re.IGNORECASE)
            input_datasets.extend(matches)
        
        for pattern in output_patterns:
            matches = re.findall(pattern, self.compiled_code, re.IGNORECASE)
            output_datasets.extend(matches)
        
        return input_datasets, output_datasets

    def _keyword_extract_datasets(self) -> Tuple[List[str], List[str]]:
        """Extract datasets using keyword search - most basic method."""
        
        input_datasets = []
        output_datasets = []
        
        lines = self.compiled_code.split('\n')
        
        for line in lines:
            line_lower = line.lower()
            
            # Look for input keywords
            if any(keyword in line_lower for keyword in ['read', 'input', 'load', 'import']):
                # Try to extract anything that looks like a filename
                potential_names = re.findall(r'[a-zA-Z_][a-zA-Z0-9_]*', line)
                input_datasets.extend(potential_names[:2])  # Take first 2 candidates
            
            # Look for output keywords  
            if any(keyword in line_lower for keyword in ['write', 'output', 'save', 'export']):
                potential_names = re.findall(r'[a-zA-Z_][a-zA-Z0-9_]*', line)
                output_datasets.extend(potential_names[:2])  # Take first 2 candidates
        
        return input_datasets, output_datasets

    def _gentle_clean_dataset_list(self, dataset_list: List[str]) -> List[str]:
        """Gently clean dataset names - don't be too strict."""
        
        cleaned = []
        seen = set()
        
        for name in dataset_list:
            if not name:
                continue
                
            # Basic cleaning - remove obvious junk but keep most things
            clean_name = str(name).strip()
            
            # Skip very short names or obvious keywords
            if len(clean_name) < 2 or clean_name.lower() in ['set', 'data', 'from', 'to', 'read', 'write']:
                continue
                
            # Keep the name (even with extensions, prefixes, etc.)
            if clean_name not in seen:
                cleaned.append(clean_name)
                seen.add(clean_name)
        
        return cleaned


            
    def create_full_lineage_excel(self, lineage_data: Dict[str, Any]) -> BytesIO:
        """Create Excel file with AI-generated full data lineage analysis."""
        buffer = BytesIO()
        
        try:
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Sheet 1: Summary with dataset info
                summary_data = []
                
                # Add input datasets if available
                detailed_input = lineage_data.get('detailed_input_datasets', [])
                detailed_output = lineage_data.get('detailed_output_datasets', [])
                
                if detailed_input:
                    for dataset in detailed_input:
                        summary_data.append({
                            'Item Type': 'Input Dataset',
                            'Name': dataset.get('name', 'Unknown'),
                            'Operation': dataset.get('operation', 'N/A'),
                            'Details': 'Source data for analysis'
                        })
                else:
                    # Fallback to simple list
                    for dataset in lineage_data.get('input_datasets', []):
                        summary_data.append({
                            'Item Type': 'Input Dataset',
                            'Name': dataset,
                            'Operation': 'READ/INPUT',
                            'Details': 'Source data identified by AI'
                        })
                
                if detailed_output:
                    for dataset in detailed_output:
                        summary_data.append({
                            'Item Type': 'Output Dataset',
                            'Name': dataset.get('name', 'Unknown'),
                            'Operation': dataset.get('operation', 'N/A'),
                            'Details': 'Result data from analysis'
                        })
                else:
                    # Fallback to simple list
                    for dataset in lineage_data.get('output_datasets', []):
                        summary_data.append({
                            'Item Type': 'Output Dataset',
                            'Name': dataset,
                            'Operation': 'WRITE/OUTPUT',
                            'Details': 'Output data identified by AI'
                        })
                
                # Create summary sheet
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                else:
                    summary_df = pd.DataFrame({
                        'Item Type': ['No datasets found'],
                        'Name': ['AI analysis did not detect input/output operations'],
                        'Operation': ['N/A'],
                        'Details': ['Check if code contains file I/O operations']
                    })
                
                summary_df.to_excel(writer, sheet_name='Datasets_Summary', index=False)
                
                # Format summary sheet
                worksheet = writer.sheets['Datasets_Summary']
                worksheet.column_dimensions['A'].width = 15  # Item Type
                worksheet.column_dimensions['B'].width = 30  # Name
                worksheet.column_dimensions['C'].width = 15  # Operation
                worksheet.column_dimensions['D'].width = 50  # Details
                
                # Sheet 2: Enhanced Data Lineage
                lineage_table = lineage_data.get('lineage_table', [])
                
                if lineage_table and len(lineage_table) > 0:
                    lineage_df = pd.DataFrame(lineage_table)
                    
                    # Ensure all required columns exist
                    required_columns = [
                        'sr_no', 'variable_name', 'variable_type', 'dataset',
                        'variable_source', 'calculation_methodology', 'explanation', 'line_no'
                    ]
                    
                    for col in required_columns:
                        if col not in lineage_df.columns:
                            lineage_df[col] = 'N/A'
                    
                    # Rename columns for Excel
                    lineage_df = lineage_df.rename(columns={
                        'sr_no': 'Sr. No',
                        'variable_name': 'Variable Name',
                        'variable_type': 'Variable Type',
                        'dataset': 'Dataset',
                        'variable_source': 'Variable Source',
                        'calculation_methodology': 'Calculation Methodology',
                        'explanation': 'Explanation',
                        'line_no': 'Line No'
                    })
                    
                    # Select and order columns
                    final_columns = [
                        'Sr. No', 'Variable Name', 'Variable Type', 'Dataset',
                        'Variable Source', 'Calculation Methodology', 'Explanation', 'Line No']
                    
                    lineage_df = lineage_df[final_columns]

                    # Define the custom order for sorting
                    category_order = ['Input', 'Intermediate', 'Created']
                    lineage_df['Variable Type'] = pd.Categorical(
                        lineage_df['Variable Type'],
                        categories=category_order,
                        ordered=True)
    
                    # Sort the DataFrame by the custom order
                    lineage_df = lineage_df.sort_values('Variable Type')
                    
                else:
                    # Create empty lineage sheet
                    lineage_df = pd.DataFrame({
                        'Sr. No': [1],
                        'Variable Name': ['No variables found'],
                        'Variable Type': ['N/A'],
                        'Dataset': ['N/A'],
                        'Variable Source': ['N/A'],
                        'Calculation Methodology': ['AI could not detect variable operations'],
                        'Explanation': ['Please check if code contains variable assignments'],
                        'Line No': ['N/A']
                    })
                
                lineage_df.to_excel(writer, sheet_name='AI_Data_Lineage', index=False)
                
                # Format lineage sheet
                worksheet = writer.sheets['AI_Data_Lineage']
                worksheet.column_dimensions['A'].width = 8   # Sr. No
                worksheet.column_dimensions['B'].width = 25  # Variable Name
                worksheet.column_dimensions['C'].width = 15  # Variable Type
                worksheet.column_dimensions['D'].width = 20  # Dataset
                worksheet.column_dimensions['E'].width = 25  # Variable Source
                worksheet.column_dimensions['F'].width = 40  # Calculation Methodology
                worksheet.column_dimensions['G'].width = 50  # Explanation
                worksheet.column_dimensions['H'].width = 10  # Line No
                
                # Apply formatting and color coding
                try:
                    # Create alignment style once
                    wrap_alignment = openpyxl.styles.Alignment(
                        wrap_text=True,
                        vertical='top'
                    )
                    
                    # Apply bounds checking
                    max_row = worksheet.max_row
                    max_col = worksheet.max_column
                    
                    if max_row > 1 and max_col > 0:
                        for row_idx in range(2, min(max_row + 1, 1500)):  
                            for col_idx in range(1, min(max_col + 1, 20)):  
                                try:
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    if cell.value is not None:  
                                        cell.alignment = wrap_alignment
                                except Exception:
                                    continue  # Skip problematic cells
                                    
                except Exception as e:
                    # Don't fail Excel creation for formatting issues
                    pass
                
                # Color coding by variable type
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                    if len(row) > 2:
                        var_type = row[2].value 
                        if var_type == "Input":
                            row[2].fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif var_type == "Created":
                            row[2].fill = openpyxl.styles.PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                        elif var_type == "Intermediate":
                            row[2].fill = openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif var_type == "Output":
                            row[2].fill = openpyxl.styles.PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")  
            
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            st.error(f"Excel creation error: {str(e)}")
            
            # Create simple fallback Excel
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                fallback_df = pd.DataFrame({
                    'Status': ['Excel creation failed'],
                    'Error': [str(e)],
                    'Solution': ['Please try the analysis again'],
                    'AI_Response_Available': ['Yes' if lineage_data.get('ai_response_length', 0) > 0 else 'No']
                })
                fallback_df.to_excel(writer, sheet_name='Error_Report', index=False)
            
            buffer.seek(0)
            return buffer
        
    def debug_ai_connection(self, test_message: str = "Hello, can you respond?") -> bool:
        """Test if AI connection is working properly."""
        
        st.write("**🔧 Testing AI Connection...**")
        
        response = self.send_message(test_message, {"type": "connection_test"})
        
        if response:
            st.success(f"✅ AI Connection Working")
            return True
        else:
            st.error("❌ AI Connection Failed")
            return False



class DataLineageVisualizer:
    
    def __init__(self):
        self.graph = nx.DiGraph()
        self.view_modes = ['overview', 'detailed', 'focused']
        self.current_mode = 'detailed'
        
        # Hierarchical positioning (Y-levels)
        self.hierarchy_levels = {
            'input_datasets': 1,
            'input_variables': 2,
            'intermediate_variables': 3,
            'created_variables': 4,
            'output_datasets': 5
        }
        
        # Color scheme for different node types
        self.node_colors = {
            'input_dataset': '#E91E63',      # Pink
            'input_variable': '#4CAF50',     # Green 
            'intermediate_variable':'#FF9800',  # Orange 
            'created_variable': '#2196F3',   # Blue 
            'output_dataset': '#9C27B0',     # Purple
            'unknown': '#757575'             # Gray
        }
        
        # Node shapes and sizes
        self.node_config = {
            'input_dataset': {'shape': 'circle', 'size': 60},
            'input_variable': {'shape': 'box', 'size': 25},
            'intermediate_variable': {'shape': 'box', 'size': 28},
            'created_variable': {'shape': 'box', 'size': 30},
            'output_dataset': {'shape': 'circle', 'size': 60},
            'unknown': {'shape': 'dot', 'size': 20}
        }
        
        # Text icons
        self.node_icons = {
            'input_dataset': '[DS]',
            'input_variable': '[IN]',
            'intermediate_variable': '[INT]',
            'created_variable': '[CR]',
            'output_dataset': '[OUT]',
            'unknown': '[?]'
        }
    
    def parse_uploaded_lineage_excel(self, uploaded_file):
        """Parse uploaded Excel file to extract lineage data."""
        
        try:
            # Read the Excel file
            df = pd.read_excel(uploaded_file, sheet_name='AI_Data_Lineage')
            
            # Convert to lineage_table format
            lineage_table = []
            for _, row in df.iterrows():
                lineage_table.append({
                    'sr_no': row.get('Sr. No', ''),
                    'variable_name': row.get('Variable Name', ''),
                    'variable_type': row.get('Variable Type', ''),
                    'dataset': row.get('Dataset', 'N/A'),
                    'variable_source': row.get('Variable Source', 'N/A'),
                    'calculation_methodology': row.get('Calculation Methodology', ''),
                    'explanation': row.get('Explanation', ''),
                    'line_no': row.get('Line No', '')
                })
            
            # Try to read datasets from other sheets
            input_datasets = []
            output_datasets = []
            
            try:
                # Read datasets summary if available
                datasets_df = pd.read_excel(uploaded_file, sheet_name='Datasets_Summary')
                for _, row in datasets_df.iterrows():
                    item_type = str(row.get('Item Type', '')).lower()
                    name = str(row.get('Name', ''))
                    if 'input' in item_type and name:
                        input_datasets.append(name)
                    elif 'output' in item_type and name:
                        output_datasets.append(name)
            except:
                # If no datasets summary, extract from lineage table
                for row in lineage_table:
                    dataset = str(row.get('dataset', '')).strip()
                    if dataset and dataset not in ['N/A', 'nan', '']:
                        if dataset not in input_datasets:
                            input_datasets.append(dataset)
            
            # Create lineage data structure
            uploaded_lineage_data = {
                'lineage_table': lineage_table,
                'input_datasets': input_datasets,
                'output_datasets': output_datasets,
                'total_variables': len(lineage_table),
                'processing_method': 'uploaded_excel',
                'file_name': uploaded_file.name
            }
            
            return uploaded_lineage_data
            
        except Exception as e:
            st.error(f"❌ Error parsing uploaded Excel: {str(e)}")
            return None
    
    def _robust_parse_lineage_data(self, lineage_data):
        """Robustly parse lineage data with extensive error handling."""
        
        parsed_data = {
            'nodes': {},
            'edges': [],
            'datasets': {'input': set(), 'output': set()},
            'variables': {'input': [], 'intermediate': [], 'created': [], 'output': []},
            'parsing_stats': {'success': 0, 'errors': 0, 'warnings': []}
        }
        
        try:
            # Get lineage table
            lineage_table = lineage_data.get('lineage_table', [])
            if not lineage_table:
                parsed_data['parsing_stats']['warnings'].append("No lineage table found")
                return parsed_data
            
            # Parse each row with error handling
            for row_idx, row in enumerate(lineage_table):
                try:
                    self._parse_single_row(row, row_idx, parsed_data)
                    parsed_data['parsing_stats']['success'] += 1
                except Exception as e:
                    parsed_data['parsing_stats']['errors'] += 1
                    parsed_data['parsing_stats']['warnings'].append(f"Row {row_idx}: {str(e)}")
                    continue  # Skip this row, don't fail
            
            # Parse input/output datasets from lineage_data
            self._parse_datasets(lineage_data, parsed_data)
            
            # Create connections with corrected intermediate variable arrows
            self._create_smart_connections(parsed_data)
            
        except Exception as e:
            parsed_data['parsing_stats']['warnings'].append(f"Major parsing error: {str(e)}")
        
        return parsed_data
    
    def _parse_single_row(self, row, row_idx, parsed_data):
        """Parse a single row with robust error handling."""
        
        try:
            # Extract basic info with fallbacks
            var_name = str(row.get('variable_name', f'var_{row_idx}')).strip()
            var_type = str(row.get('variable_type', 'unknown')).lower().strip()
            dataset = str(row.get('dataset', 'N/A')).strip()
            variable_source = str(row.get('variable_source', 'N/A')).strip()
            calculation = str(row.get('calculation_methodology', '')).strip()
            explanation = str(row.get('explanation', '')).strip()
            
            if not var_name or var_name == 'nan':
                return  # Skip invalid rows
            
            # Determine node type with exact case matching
            node_type = self._determine_node_type(var_type)
            
            # Create node
            node_id = f"VAR_{var_name}"
            parsed_data['nodes'][node_id] = {
                'id': node_id,
                'label': var_name,
                'type': node_type,
                'original_type': var_type,
                'dataset': dataset,
                'variable_source': variable_source,
                'calculation': calculation,
                'explanation': explanation,
                'level': self.hierarchy_levels.get(f'{node_type}_variables', 3)
            }
            
            # Categorize variable
            if node_type == 'input_variable':
                parsed_data['variables']['input'].append(node_id)
            elif node_type == 'intermediate_variable':
                parsed_data['variables']['intermediate'].append(node_id)
            elif node_type == 'created_variable':
                parsed_data['variables']['created'].append(node_id)
            
            # Parse dataset source
            self._parse_dataset_source(dataset, node_id, parsed_data)
            
            # Parse variable sources
            self._parse_variable_sources(variable_source, node_id, parsed_data)
            
        except Exception as e:
            raise Exception(f"Error parsing row data: {str(e)}")
    
    def _determine_node_type(self, var_type):
        """Determine node type from variable type string."""
        
        var_type_clean = str(var_type).strip().lower()
        
        # Direct matching for exact lineage types with suffixes for color mapping
        if var_type_clean == 'input':
            return 'input_variable'  
        elif var_type_clean == 'intermediate':
            return 'intermediate_variable'  
        elif var_type_clean == 'created':
            return 'created_variable'  
        else:
            # Debug output to see unrecognized types
            print(f"🔍 DEBUG: Unrecognized variable type: '{var_type}' -> defaulting to 'unknown'")
            return 'unknown'
    
    def _parse_dataset_source(self, dataset_source, node_id, parsed_data):
        """Parse dataset source with correct edge direction based on variable type."""
        
        try:
            if dataset_source and dataset_source not in ['N/A', 'nan', 'null', '']:
                # Clean dataset name
                clean_dataset = self._clean_dataset_name(dataset_source)
                
                if clean_dataset:
                    # Get the variable type from the already created node
                    variable_type = parsed_data['nodes'][node_id]['type']
                    
                    # Determine dataset type and edge direction based on variable type
                    if variable_type == 'input_variable':
                        # Input variables come FROM input datasets
                        # Direction: dataset → variable (dataset "contains" variable)
                        dataset_node_id = f"DS_INPUT_{clean_dataset}"
                        
                        # Create input dataset node if not exists
                        if dataset_node_id not in parsed_data['nodes']:
                            parsed_data['nodes'][dataset_node_id] = {
                                'id': dataset_node_id,
                                'label': clean_dataset,
                                'type': 'input_dataset',
                                'level': self.hierarchy_levels['input_datasets']
                            }
                            parsed_data['datasets']['input'].add(clean_dataset)
                        
                        # Create edge: dataset → variable
                        parsed_data['edges'].append({
                            'from': dataset_node_id,  # FROM dataset
                            'to': node_id,            # TO variable
                            'type': 'contains',
                            'label': 'contains'
                        })
                        
                    elif variable_type in ['intermediate_variable', 'created_variable']:
                        # Intermediate/Created variables go TO output datasets
                        # Direction: variable → dataset (variable "feeds into" dataset)
                        dataset_node_id = f"DS_OUTPUT_{clean_dataset}"
                        
                        # Create output dataset node if not exists
                        if dataset_node_id not in parsed_data['nodes']:
                            parsed_data['nodes'][dataset_node_id] = {
                                'id': dataset_node_id,
                                'label': clean_dataset,
                                'type': 'output_dataset',
                                'level': self.hierarchy_levels['output_datasets']
                            }
                            parsed_data['datasets']['output'].add(clean_dataset)
                        
                        # Create edge: variable → dataset
                        parsed_data['edges'].append({
                            'from': node_id,          # FROM variable
                            'to': dataset_node_id,    # TO dataset
                            'type': 'feeds_into',
                            'label': 'feeds into'
                        })
                        
                    else:
                        # For other variable types (like output_variable), treat as output
                        dataset_node_id = f"DS_OUTPUT_{clean_dataset}"
                        
                        if dataset_node_id not in parsed_data['nodes']:
                            parsed_data['nodes'][dataset_node_id] = {
                                'id': dataset_node_id,
                                'label': clean_dataset,
                                'type': 'output_dataset',
                                'level': self.hierarchy_levels['output_datasets']
                            }
                            parsed_data['datasets']['output'].add(clean_dataset)
                        
                        # Create edge: variable → dataset
                        parsed_data['edges'].append({
                            'from': node_id,          # FROM variable
                            'to': dataset_node_id,    # TO dataset
                            'type': 'outputs_to',
                            'label': 'outputs to'
                        })
        except Exception as e:
            # Don't fail, just log warning
            parsed_data['parsing_stats']['warnings'].append(f"Dataset source parsing error: {str(e)}")

    
    def _clean_dataset_name(self, dataset):
        """Clean dataset name from various AI-generated formats."""
        
        try:
            # Remove common prefixes and suffixes
            clean_name = dataset
            
            # Remove file extensions
            for ext in ['.csv', '.xlsx', '.sas7bdat', '.parquet', '.txt']:
                clean_name = clean_name.replace(ext, '')
            
            # Remove common prefixes
            prefixes_to_remove = ['dataset:', 'file:', 'table:', 'data:', 'source:']
            for prefix in prefixes_to_remove:
                if clean_name.lower().startswith(prefix):
                    clean_name = clean_name[len(prefix):].strip()
            
            # Remove quotes and brackets
            clean_name = clean_name.strip('"\'[](){}')
            
            # Remove paths (keep only filename)
            if '/' in clean_name:
                clean_name = clean_name.split('/')[-1]
            if '\\' in clean_name:
                clean_name = clean_name.split('\\')[-1]
            
            return clean_name.strip() if clean_name.strip() else None
            
        except Exception:
            return None
    
    def _parse_variable_sources(self, variable_source, node_id, parsed_data):
        """Parse variable sources with robust handling of AI-generated formats."""
        
        try:
            if variable_source and variable_source not in ['N/A', 'nan', 'null', '']:
                # Parse multiple source variables
                source_vars = self._extract_source_variables(variable_source)
                
                for source_var in source_vars:
                    if source_var:
                        # Create edge from source variable to current variable
                        source_node_id = f"VAR_{source_var}"
                        
                        parsed_data['edges'].append({
                            'from': source_node_id,
                            'to': node_id,
                            'type': 'transforms',
                            'label': 'transforms'
                        })
        except Exception as e:
            # Don't fail, just log warning
            parsed_data['parsing_stats']['warnings'].append(f"Variable source parsing error: {str(e)}")
    
    def _extract_source_variables(self, variable_source):
        """Extract source variables from various AI-generated formats."""
        
        try:
            # Handle different separator formats
            separators = [',', ';', '|', ' and ', ' & ', '\n']
            
            # Split by various separators
            source_vars = [variable_source]
            for sep in separators:
                temp_vars = []
                for var in source_vars:
                    temp_vars.extend(v.strip() for v in var.split(sep))
                source_vars = temp_vars
            
            # Clean each variable name
            cleaned_vars = []
            for var in source_vars:
                cleaned_var = self._clean_variable_name(var)
                if cleaned_var:
                    cleaned_vars.append(cleaned_var)
            
            return cleaned_vars
            
        except Exception:
            return []
    
    def _clean_variable_name(self, var_name):
        """Clean variable name from AI-generated formats."""
        
        try:
            # Remove common prefixes
            prefixes_to_remove = [
                'calculated:', 'derived from:', 'based on:', 'using:', 'from:',
                'variable:', 'var:', 'field:', 'column:', 'col:'
            ]
            
            clean_name = var_name.strip()
            
            for prefix in prefixes_to_remove:
                if clean_name.lower().startswith(prefix):
                    clean_name = clean_name[len(prefix):].strip()
            
            # Remove quotes, brackets, and parentheses content
            clean_name = clean_name.strip('"\'[]')
            
            # Remove content in parentheses 
            if '(' in clean_name and ')' in clean_name:
                clean_name = clean_name.split('(')[0].strip()
            
            # Remove non-alphanumeric characters except underscore
            import re
            clean_name = re.sub(r'[^a-zA-Z0-9_]', '', clean_name)
            
            return clean_name if clean_name and len(clean_name) > 1 else None
            
        except Exception:
            return None
    
    def _parse_datasets(self, lineage_data, parsed_data):
        """Parse input/output datasets from lineage_data."""
        
        try:
            # Input datasets
            input_datasets = lineage_data.get('input_datasets', [])
            for dataset in input_datasets:
                if dataset:
                    clean_dataset = self._clean_dataset_name(str(dataset))
                    if clean_dataset:
                        parsed_data['datasets']['input'].add(clean_dataset)
            
            # Output datasets
            output_datasets = lineage_data.get('output_datasets', [])
            for dataset in output_datasets:
                if dataset:
                    clean_dataset = self._clean_dataset_name(str(dataset))
                    if clean_dataset:
                        parsed_data['datasets']['output'].add(clean_dataset)
        
        except Exception as e:
            parsed_data['parsing_stats']['warnings'].append(f"Dataset parsing error: {str(e)}")
    
    def _create_smart_connections(self, parsed_data):
        """FIXED: Create smart connections with corrected intermediate variable arrow direction."""
        
        try:
            # Create output dataset connections
            for dataset in parsed_data['datasets']['output']:
                dataset_node_id = f"DS_OUTPUT_{dataset}"
                
                # Create output dataset node if not exists
                if dataset_node_id not in parsed_data['nodes']:
                    parsed_data['nodes'][dataset_node_id] = {
                        'id': dataset_node_id,
                        'label': dataset,
                        'type': 'output_dataset',
                        'level': self.hierarchy_levels['output_datasets']
                    }
                
                # Connect created/output variables to output datasets
                for var_id in parsed_data['variables']['created'] + parsed_data['variables']['output']:
                    if var_id in parsed_data['nodes']:
                        parsed_data['edges'].append({
                            'from': var_id,
                            'to': dataset_node_id,
                            'type': 'outputs_to',
                            'label': 'outputs to'
                        })
            
            # Ensure intermediate variables have OUTWARD arrows
            for intermediate_var_id in parsed_data['variables']['intermediate']:
                if intermediate_var_id in parsed_data['nodes']:
                    # Check if intermediate variable has any outgoing connections
                    has_outgoing = any(edge['from'] == intermediate_var_id for edge in parsed_data['edges'])
                    
                    if not has_outgoing:
                        # Connect intermediate variables to created variables (outward direction)
                        target_found = False
                        
                        # First try to connect to created variables
                        for created_var_id in parsed_data['variables']['created']:
                            if created_var_id in parsed_data['nodes']:
                                parsed_data['edges'].append({
                                    'from': intermediate_var_id,  
                                    'to': created_var_id,        
                                    'type': 'feeds_into',
                                    'label': 'feeds into'
                                })
                                target_found = True
                                break  
                        
                        # If no created variables, try connecting to output datasets
                        for dataset in parsed_data['datasets']['output']:
                            dataset_node_id = f"DS_OUTPUT_{dataset}"
                            if dataset_node_id in parsed_data['nodes']:
                                parsed_data['edges'].append({
                                    'from': intermediate_var_id,  
                                    'to': dataset_node_id,        
                                    'type': 'flows_to',
                                    'label': 'flows to'
                                })
                                break  
        
        except Exception as e:
            parsed_data['parsing_stats']['warnings'].append(f"Smart connections error: {str(e)}")
    
    def create_visualization_with_modes(self, lineage_data, mode='overview'):
        """Create visualization with multiple view modes."""
        
        try:
            # Parse data robustly
            parsed_data = self._robust_parse_lineage_data(lineage_data)
            
            # Create visualization based on mode
            if mode == 'overview':
                return self._create_overview_visualization(parsed_data)
            elif mode == 'detailed':
                return self._create_detailed_visualization(parsed_data)
            elif mode == 'focused':
                return self._create_focused_visualization(parsed_data)
            else:
                return self._create_overview_visualization(parsed_data)
        
        except Exception as e:
            return self._create_error_visualization(f"Visualization error: {str(e)}")
    
    def _create_overview_visualization(self, parsed_data):
        """Create overview visualization with font sizing and colors."""
        
        try:
            from pyvis.network import Network
            
            # Create network
            net = Network(
                height="600px",
                width="100%",
                bgcolor="#f8f9fa",
                font_color="#000000",
                directed=True
            )
            
            # Add nodes with styling
            x_positions = {}  # Track X positions for each level
            
            for node_id, node_data in parsed_data['nodes'].items():
                try:
                    node_type = node_data['type']
                    level = node_data['level']
                    
                    # Calculate position
                    if level not in x_positions:
                        x_positions[level] = 0
                    x_positions[level] += 150
                    
                    # Get node configuration
                    config = self.node_config.get(node_type, self.node_config['unknown'])
                    color = self.node_colors.get(node_type, self.node_colors['unknown'])
                    icon = self.node_icons.get(node_type, self.node_icons['unknown'])
                    
                    # Simple text label (no HTML)
                    label = f"{icon} {node_data['label']}"
                    
                    # Font sizing via pyvis parameters - Dataset names LARGER than variable names
                    if 'dataset' in node_type:
                        font_size = 30  
                        font_bold = True
                    else:
                        font_size = 14 
                        font_bold = True
                    
                    # Create hover text
                    hover_text = f"{label}\nType: {node_type}\nLevel: {level}"
                    
                    # Add node with styling
                    net.add_node(
                        node_id,
                        label=label,
                        color=color,
                        size=config['size'],
                        shape=config['shape'],
                        title=hover_text,
                        font={'size': font_size, 'bold': font_bold},
                        x=x_positions[level],
                        y=level * 150,
                        physics=False  # positions for clear hierarchy
                    )
                
                except Exception as e:
                    continue  # Skip problematic nodes
            
            # Add edges with smaller font
            edge_count = 0
            for edge in parsed_data['edges']:
                try:
                    if edge['from'] in parsed_data['nodes'] and edge['to'] in parsed_data['nodes']:
                        net.add_edge(
                            edge['from'],
                            edge['to'],
                            color="#666666",
                            width=2,
                            arrows="to",
                            title=edge.get('label', ''),
                            label=edge.get('label', ''),
                            font={'size': 8}  # Even smaller font for edge labels 
                        )
                        edge_count += 1
                except Exception:
                    continue  # Skip problematic edges
            
            # Configure network
            net.set_options("""
            var options = {
                "physics": {
                    "enabled": false
                },
                "interaction": {
                    "dragNodes": true,
                    "dragView": true,
                    "zoomView": true
                },
                "layout": {
                    "hierarchical": {
                        "enabled": true,
                        "direction": "UD",
                        "sortMethod": "directed",
                        "levelSeparation": 150,
                        "nodeSpacing": 200
                    }
                }
            }
            """)
            
            # Generate HTML
            html_content = net.generate_html()
            
            # Add custom header
            stats_text = f"{len(parsed_data['nodes'])} Nodes | {edge_count} Connections | Overview Mode"
            enhanced_html = html_content.replace(
                '<body>',
                f'''<body>
                <div style="text-align: center; padding: 10px; background: #f0f0f0; margin-bottom: 10px;">
                    <h3 style="margin: 0; color: #333;">🗺️ Data Lineage Overview</h3>
                    <p style="margin: 5px 0 0 0; color: #666;">{stats_text}</p>
                </div>'''
            )
            return enhanced_html
        except Exception as e:
            return self._create_error_visualization(f"Overview visualization error: {str(e)}")
    
    def _create_detailed_visualization(self, parsed_data):
        """Create detailed visualization with font sizing and colors."""
        try:
            from pyvis.network import Network
            
            # Create network with more space
            net = Network(
                height="800px",
                width="100%",
                bgcolor="#f8f9fa",
                font_color="#000000",
                directed=True
            )
            
            # Add all nodes with styling
            for node_id, node_data in parsed_data['nodes'].items():
                try:
                    node_type = node_data['type']
                    config = self.node_config.get(node_type, self.node_config['unknown'])
                    color = self.node_colors.get(node_type, self.node_colors['unknown'])
                    icon = self.node_icons.get(node_type, self.node_icons['unknown'])
                    
                    # Simple text label (no HTML)
                    label = f"{icon} {node_data['label']}"
                    
                    # Font sizing via pyvis parameters - Dataset names LARGER than variable names
                    if 'dataset' in node_type:
                        font_size = 20  # LARGER for datasets 
                        font_bold = True
                    else:
                        font_size = 14  # Medium for variables 
                        font_bold = True
                    
                    # Detailed hover text
                    hover_text = f"{label}\nType: {node_type}\nDataset: {node_data.get('dataset', 'N/A')}\nSources: {node_data.get('variable_source', 'N/A')}\nCalculation: {node_data.get('calculation', 'N/A')[:50]}..."
                    
                    net.add_node(
                        node_id,
                        label=label,
                        color=color,
                        size=config['size'],
                        shape=config['shape'],
                        title=hover_text,
                        font={'size': font_size, 'bold': font_bold}
                    )
                
                except Exception:
                    continue
            
            # Add all edges with FIXED smaller font
            edge_count = 0
            for edge in parsed_data['edges']:
                try:
                    if edge['from'] in parsed_data['nodes'] and edge['to'] in parsed_data['nodes']:
                        net.add_edge(
                            edge['from'],
                            edge['to'],
                            color="#666666",
                            width=2,
                            arrows="to",
                            title=edge.get('label', ''),
                            label=edge.get('label', ''),
                            font={'size': 8}  
                        )
                        edge_count += 1
                except Exception:
                    continue
            
            node_count = len(parsed_data['nodes'])
            edge_count = len(parsed_data['edges'])
            
            # Configure for detailed view for larger lineage - Disable physics
            if node_count > 100 or edge_count > 500:
                net.set_options("""
                var options = {
                "physics": {
                    "enabled": false},
                    "layout": {
                        "hierarchical": {
                        "enabled": true,
                        "direction": "UD",
                        "sortMethod": "directed"
                    }
                },
                "interaction": {
                    "dragNodes": true,
                    "dragView": true,
                    "zoomView": true
                }
            }
            """)
            else:
                # Configure for detailed view for smaller lineage
                net.set_options("""
                var options = {
                    "physics": {
                        "enabled": true,
                        "barnesHut": {
                            "gravitationalConstant": -8000,
                            "springLength": 200,
                            "springConstant": 0.04,
                            "damping": 0.09
                        }
                    },
                    "interaction": {
                        "dragNodes": true,
                        "dragView": true,
                        "zoomView": true
                    }
                }
                """)
            
            html_content = net.generate_html()
            
            # Add detailed header
            stats_text = f"{len(parsed_data['nodes'])} Nodes | {edge_count} Connections | Detailed Mode"
            enhanced_html = html_content.replace(
                '<body>',
                f'''<body>
                <div style="text-align: center; padding: 10px; background: #f0f0f0; margin-bottom: 10px;">
                    <h3 style="margin: 0; color: #333;">🔍 Detailed Data Lineage</h3>
                    <p style="margin: 5px 0 0 0; color: #666;">{stats_text}</p>
                </div>'''
            )
            
            return enhanced_html
            
        except Exception as e:
            return self._create_error_visualization(f"Detailed visualization error: {str(e)}")
    
    def _create_focused_visualization(self, parsed_data):
        """Create focused visualization with FIXED font sizing and colors."""
        
        try:
            from pyvis.network import Network
            
            # Create network
            net = Network(
                height="700px",
                width="100%",
                bgcolor="#f8f9fa",
                font_color="#000000",
                directed=True
            )
            
            # Focus on created variables and their immediate connections
            focus_nodes = set()
            focus_edges = []
            
            # Add created variables
            for var_id in parsed_data['variables']['created']:
                if var_id in parsed_data['nodes']:
                    focus_nodes.add(var_id)
            
            # Add their source variables
            for edge in parsed_data['edges']:
                if edge['to'] in focus_nodes:
                    focus_nodes.add(edge['from'])
                    focus_edges.append(edge)
            
            # Add nodes with styling
            for node_id in focus_nodes:
                if node_id in parsed_data['nodes']:
                    try:
                        node_data = parsed_data['nodes'][node_id]
                        node_type = node_data['type']
                        config = self.node_config.get(node_type, self.node_config['unknown'])
                        color = self.node_colors.get(node_type, self.node_colors['unknown'])
                        icon = self.node_icons.get(node_type, self.node_icons['unknown'])
                        
                        # Simple text label (no HTML)
                        label = f"{icon} {node_data['label']}"
                        
                        # Font sizing via pyvis parameters
                        if 'dataset' in node_type:
                            font_size = 30  # Larger for datasets
                            font_bold = True
                        else:
                            font_size = 14  # Medium for variables
                            font_bold = True
                        
                        # Focused hover text
                        hover_text = f"{label}\nType: {node_type}\nCalculation: {node_data.get('calculation', 'N/A')[:100]}..."
                        
                        net.add_node(
                            node_id,
                            label=label,
                            color=color,
                            size=config['size'] + 10,  # Slightly larger for focus
                            shape=config['shape'],
                            title=hover_text,
                            font={'size': font_size, 'bold': font_bold}
                        )
                    
                    except Exception:
                        continue
            
            # Add focused edges with FIXED smaller font
            for edge in focus_edges:
                try:
                    if edge['from'] in focus_nodes and edge['to'] in focus_nodes:
                        net.add_edge(
                            edge['from'],
                            edge['to'],
                            color="#2196F3",
                            width=3,
                            arrows="to",
                            title=edge.get('label', ''),
                            label=edge.get('label', ''),
                            font={'size': 10}  # Smaller font for edge labels
                        )
                except Exception:
                    continue
            
            # Configure for focused view
            net.set_options("""
            var options = {
                "physics": {
                    "enabled": true,
                    "barnesHut": {
                        "gravitationalConstant": -4000,
                        "springLength": 150,
                        "springConstant": 0.08,
                        "damping": 0.1
                    }
                },
                "interaction": {
                    "dragNodes": true,
                    "dragView": true,
                    "zoomView": true
                }
            }
            """)
            
            html_content = net.generate_html()
            
            # Add focused header
            stats_text = f"{len(focus_nodes)} Focused Nodes | {len(focus_edges)} Connections | Created Variables Focus"
            enhanced_html = html_content.replace(
                '<body>',
                f'''<body>
                <div style="text-align: center; padding: 10px; background: #f0f0f0; margin-bottom: 10px;">
                    <h3 style="margin: 0; color: #333;">🎯 Focused Data Lineage</h3>
                    <p style="margin: 5px 0 0 0; color: #666;">{stats_text}</p>
                </div>'''
            )
            
            return enhanced_html
            
        except Exception as e:
            return self._create_error_visualization(f"Focused visualization error: {str(e)}")
    
    def _create_error_visualization(self, error_message):
        """Create simple error visualization."""
        
        error_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Data Lineage Visualization Error</title>
        </head>
        <body>
            <div style="text-align: center; padding: 50px; font-family: Arial, sans-serif;">
                <h2 style="color: #f44336;">⚠️ Visualization Error</h2>
                <p style="color: #666; margin: 20px 0;">
                    {error_message}
                </p>
                <p style="color: #999; font-size: 14px;">
                    Please check your data format and try again.
                </p>
            </div>
        </body>
        </html>
        """
        
        return error_html
    
    def display_visualization_stats(self, lineage_data):
        """Display visualization statistics with robust error handling."""
        
        try:
            # Parse data to get stats
            parsed_data = self._robust_parse_lineage_data(lineage_data)
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_nodes = len(parsed_data['nodes'])
                st.metric("📊 Total Nodes", total_nodes)
            
            with col2:
                total_edges = len(parsed_data['edges'])
                st.metric("🔗 Connections", total_edges)
            
            with col3:
                input_count = len(parsed_data['variables']['input'])
                st.metric("📥 Input Variables", input_count)
            
            with col4:
                created_count = len(parsed_data['variables']['created'])
                st.metric("⚙️ Created Variables", created_count)
            
            # Show parsing stats
            stats = parsed_data['parsing_stats']
            if stats['errors'] > 0 or stats['warnings']:
                with st.expander("⚠️ Parsing Statistics", expanded=False):
                    st.write(f"**Successful:** {stats['success']}")
                    st.write(f"**Errors:** {stats['errors']}")
                    if stats['warnings']:
                        st.write("**Warnings:**")
                        for warning in stats['warnings'][:5]:  # Show first 5 warnings
                            st.write(f"• {warning}")
            
        except Exception as e:
            st.error(f"Error displaying stats: {str(e)}")
    
    def create_enhanced_table_display(self, lineage_data):
        """Create enhanced table display with error handling."""
        
        try:
            lineage_table = lineage_data.get('lineage_table', [])
            
            if not lineage_table:
                return pd.DataFrame()
            
            # Convert to DataFrame safely
            df = pd.DataFrame(lineage_table)
            
            # Ensure required columns exist
            required_columns = ['variable_name', 'variable_type', 'dataset', 'variable_source']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = 'N/A'
            
            return df[required_columns].reset_index(drop=True)
            
        except Exception as e:
            st.error(f"Error creating table display: {str(e)}")
            return pd.DataFrame()
    
    # Keep the main interface method
    def create_interactive_lineage_visualization(self, lineage_data):
        """Main interface method - creates overview by default."""
        return self.create_visualization_with_modes(lineage_data, 'overview')




