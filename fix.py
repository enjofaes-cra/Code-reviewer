1) running the same code give same inputs and outupts, but created vs intermediate is changing --> therefore network changing:: context to be refined here. eample notice_of_default_fee
2) we need to look at context to define created and input vs parameter by language. to be refined. example term_plus_12
3) SAS adding variables looped (average_ead_i) vs python not - align instructions there.
4) sometime it is adding data sets to variable - context to be refined there


app.py:
import streamlit as st
import pandas as pd
from typing import List
from io import BytesIO
import openpyxl
import json
from models.Code_Reviewer import *
from models.RDE_Tester import *
from models.Data_Lineage import *
from components.AVA_Assistant import *
from components.functions import *
import warnings
import logging

warnings.filterwarnings("ignore", message=".*missing ScriptRunContext*")

logging.getLogger("streamlit.runtime.scriptrunner.script_runner").setLevel(logging.ERROR)

def perform_data_lineage_analysis(code_files: list, data_files: list, selected_vars: List[str] = None):
    """Enhanced AI-powered data lineage analysis."""
    
    if not code_files:
        st.error("❌ Please upload code files to analyze.")
        return
    
    try:
        # Initialize session state if not exists
        if 'lineage_variables_discovered' not in st.session_state:
            st.session_state['lineage_variables_discovered'] = False
        if 'lineage_all_variables' not in st.session_state:
            st.session_state['lineage_all_variables'] = {}
        
        # STEP 1: AI Variable Discovery
        if not st.session_state['lineage_variables_discovered']:
            with st.spinner("🔍 AI analyzing your scripts to discover all variables..."):
                lineage_analyzer = DataLineageAnalyzer()
                all_variables = lineage_analyzer._simple_ai_variable_extraction(code_files)
                
                if not all_variables.get('all_variables'):
                    st.error("❌ No variables found in the uploaded scripts.")
                    with st.expander("🔧 Debug Information", expanded=True):
                        st.write("**Files uploaded:**")
                        for file in code_files:
                            st.write(f"- {file.name}")
                        
                        st.write("**First 500 characters of first file:**")
                        try:
                            content = code_files[0].read().decode('utf-8')
                            st.code(content[:500])
                            code_files[0].seek(0)
                        except Exception as e:
                            st.write(f"Could not read file: {str(e)}")
                    return
                
                # Store in session state
                store_variables_in_session(all_variables)
        
        # Get variables from session state
        all_variables = st.session_state['lineage_all_variables']
        
        # Display discovered variables summary (always show if variables are discovered)
        if st.session_state['lineage_variables_discovered']:
            st.success(f"✅ AI discovered {len(all_variables['all_variables'])} variables in your scripts!")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📥 Input Variables", len(all_variables.get('input_variables', [])))
            with col2:
                st.metric("🔄 Intermediate Variables", len(all_variables.get('intermediate_variables', [])))
            with col3:
                st.metric("⚙️ Created Variables", len(all_variables.get('created_variables', [])))
                
            # with col4:
            #     st.metric("📤 Output Variables", len(all_variables.get('output_variables', [])))
            
            # Show discovered variables by category
            with st.expander("📋 View All Discovered Variables", expanded=False):
                tab1, tab2, tab3 = st.tabs(["📥 Input", "🔄 Intermediate","⚙️ Created"])
                
                with tab1:
                    if all_variables.get('input_variables'):
                        for var in all_variables['input_variables']:
                            st.write(f"• {var}")
                    else:
                        st.info("No input variables detected")
                
                with tab2:
                    if all_variables.get('intermediate_variables'):
                        for var in all_variables['intermediate_variables']:
                            st.write(f"• {var}")
                    else:
                        st.info("No intermediate variables detected")
                
                with tab3:
                    if all_variables.get('created_variables'):
                        for var in all_variables['created_variables']:
                            st.write(f"• {var}")
                    else:
                        st.info("No created variables detected")                
            
            
            if st.session_state.get('lineage_variables_discovered', False):
                
                # ========== SEPARATE MANUAL UPLOAD SECTION ==========
                st.markdown("---")  # Separator
                st.markdown("### 📁 Upload Existing Lineage Excel for Visualization")
                
                # Separate file uploader
                uploaded_excel_file = st.file_uploader(
                    "Upload your existing data lineage Excel file",
                    type=['xlsx'],
                    key="manual_excel_uploader",  # Unique key
                    help="Upload Excel file with 'AI_Data_Lineage' sheet containing Variable Name, Variable Type, Dataset Source, Variable Source columns"
                )
                
                if uploaded_excel_file:
                    
                    # Parse uploaded file button
                    if st.button("📊 Parse Excel & Create Visualization", type="secondary", key="parse_excel_btn"):
                        
                        with st.spinner("📊 Parsing uploaded Excel file..."):
                            try:
                                # Use the same visualizer class but completely separate
                                manual_visualizer = DataLineageVisualizer()
                                manual_lineage_data = manual_visualizer.parse_uploaded_lineage_excel(uploaded_excel_file)
                                
                                if manual_lineage_data and manual_lineage_data.get('lineage_table'):
                                    # Store in separate session state
                                    st.session_state['manual_lineage_data'] = manual_lineage_data
                                    st.session_state['manual_lineage_ready'] = True
                                    st.success(f"✅ Excel parsed successfully! Found {len(manual_lineage_data['lineage_table'])} variables")
                                else:
                                    st.error("❌ Failed to parse Excel file. Please check the format.")
                                    
                            except Exception as e:
                                st.error(f"❌ Error parsing Excel: {str(e)}")
                                
                    # Show manual visualization section if data is ready
                    if st.session_state.get('manual_lineage_ready', False):
                        
                        manual_data = st.session_state.get('manual_lineage_data', {})
                        
                        st.markdown("#### 🎨 Manual Excel Visualization")
                        st.success(f"📊 **Source**: {uploaded_excel_file.name}")
                        
                        # Show data summary
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("📊 Variables", len(manual_data.get('lineage_table', [])))
                        with col2:
                            st.metric("📥 Input Datasets", len(manual_data.get('input_datasets', [])))
                        with col3:
                            st.metric("📤 Output Datasets", len(manual_data.get('output_datasets', [])))
                        
                        # Visualization controls 
                        col1, col2, col3 = st.columns([2, 1, 1])

                        with col1:
                            # View mode selection with unique key
                            manual_view_mode = st.selectbox(
                                "🎯 Select Visualization Mode:",
                                ["overview", "detailed", "focused"],
                                index=0,
                                key="manual_view_mode_selector",  # Unique key
                                help="Overview: Hierarchical flow | Detailed: All connections | Focused: Created variables only"
                            )

                        with col2:
                            # Create visualization button with unique key
                            create_manual_viz_btn = st.button(
                                "🎨 Create Visualization", 
                                type="primary", 
                                use_container_width=True, 
                                key="create_manual_viz_btn"  # Unique key
                            )
                        dl_placeholder2 = col3.empty()
                        manual_viz_key = f"manual_viz_{manual_view_mode}"
                        if st.session_state.get('manual_viz_dict', {}).get(manual_viz_key):
                            manual_viz_html = st.session_state['manual_viz_dict'][manual_viz_key]
                            timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                            with dl_placeholder2:
                                st.download_button(
                                    label="📥 Download HTML",
                                    data=manual_viz_html,
                                    file_name=f"manual_lineage_{manual_view_mode}_{timestamp}.html",
                                    mime="text/html",
                                    key=f"download_manual_viz_{manual_view_mode}_btn"
                                )

                        # Create visualization when button is clicked
                        if create_manual_viz_btn:
                            with st.spinner(f"🎨 Creating {manual_view_mode} visualization from uploaded Excel..."):
                                try:
                                    # Use the same visualizer class but completely independent
                                    manual_visualizer = DataLineageVisualizer()
                                    
                                    # Show statistics
                                    st.markdown("**📊 Manual Excel Visualization Statistics:**")
                                    manual_visualizer.display_visualization_stats(manual_data)
                                    
                                    # Create visualization
                                    manual_viz_html = manual_visualizer.create_visualization_with_modes(manual_data, manual_view_mode)
                                    
                                    if manual_viz_html:
                                        if 'manual_viz_dict' not in st.session_state:
                                            st.session_state['manual_viz_dict'] = {}

                                        manual_viz_key = f"manual_viz_{manual_view_mode}"
                                        st.session_state['manual_viz_dict'][manual_viz_key] = manual_viz_html

                                        st.rerun()
                                    else:
                                        st.error("❌ Failed to create visualization")
                                        
                                except Exception as e:
                                    st.error(f"❌ Visualization error: {str(e)}")

                        # Display the visualization if it exists
                        manual_viz_key = f"manual_viz_{manual_view_mode}"
                        if st.session_state.get('manual_viz_dict', {}).get(manual_viz_key):
                            
                            manual_viz_html = st.session_state['manual_viz_dict'][manual_viz_key]
                            
                            # Display the interactive visualization
                            st.components.v1.html(manual_viz_html, height=750, scrolling=True)
                            
                            # Show features
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.markdown("**🎨 Colors**")
                                st.write("• Green: Input variables")
                                st.write("• Orange: Intermediate variables")
                                st.write("• Blue: Created variables")
                            
                        else:
                            # Show placeholder
                            st.info(f"Click '🎨 Create Visualization' above to generate the {manual_view_mode} visualization from your Excel data")

                # Help section for manual upload
                with st.expander("📋 Expected Excel Format for Manual Upload", expanded=False):
                    st.markdown("""
                    **Required Sheet Name**: `AI_Data_Lineage`
                    
                    **Required Columns**:
                    - **Variable Name**: Name of the variable/column
                    - **Variable Type**: Must be exactly: `Input`, `Intermediate`, `Created`, or `Output`
                    - **Dataset Source**: Source dataset name (for input variables)
                    - **Variable Source**: Source variables used to create this variable
                    
                    **Example Excel Data**:
                    ```
                    Variable Name | Variable Type | Dataset Source | Variable Source
                    customer_id   | Input        | customers.csv  | N/A
                    revenue_2024  | Created      | N/A           | revenue, year
                    profit_calc   | Intermediate | N/A           | revenue_2024, costs
                    ```
                    
                    **📁 Use the Excel files generated by this tool's "Download Excel" feature for guaranteed compatibility!**
                    """)

            # ========== END OF SEPARATE MANUAL UPLOAD SECTION ==========
            
            # Add Full Data Lineage Analysis
            st.markdown("### 📊 Full Data Lineage")

            # Check if data files were also uploaded
            data_files_available = data_files if data_files else []

            col1, col2 = st.columns([2, 1])
            with col1:
                if st.button("🗺️ Generate Full Data Lineage", type="primary", use_container_width=True, key="full_lineage_btn"):
                    
                    # First test AI connection
                    lineage_analyzer = DataLineageAnalyzer()
                    if lineage_analyzer.debug_ai_connection():
                        
                        with st.spinner("🤖 AI analyzing code for comprehensive data lineage..."):
                            st.write(f"📁 Analyzing {len(code_files)} code files...")
                            
                            # Show what files we're analyzing
                            for file in code_files:
                                content = file.read().decode('utf-8')
                                st.write(f"• {file.name}: {len(content.split())} lines, {len(content)} characters")
                                file.seek(0)
                            
                            full_lineage_data = lineage_analyzer.generate_full_data_lineage(all_variables, code_files, data_files_available)
                        
                        if "error" not in full_lineage_data:
                            # Store in session state
                            st.session_state['full_lineage_data'] = full_lineage_data
                            st.session_state['full_lineage_complete'] = True
                            
                            st.success(f"✅ Full data lineage generated!")
                            st.info(f"📊 Found {full_lineage_data['total_variables']} variables")
                        else:
                            st.error("❌ Failed to generate full data lineage")
                    else:
                        st.error("❌ Cannot proceed - AI connection failed")

            with col2:
                if st.session_state.get('full_lineage_complete', False):
                    full_lineage_data = st.session_state.get('full_lineage_data', {})
                    if full_lineage_data and full_lineage_data.get('lineage_table'):
                        lineage_analyzer = DataLineageAnalyzer()
                        excel_buffer = lineage_analyzer.create_full_lineage_excel(full_lineage_data)
                        
                        excel_buffer.seek(0)
                        
                        st.download_button(
                            label="📥 Download Excel",
                            data=excel_buffer,
                            file_name=f"ai_data_lineage.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_full_lineage"
                        )
                        st.info(f"📊 In case of warnings - Try downloading the Lineage Excel again")

            # Display Full Lineage Results if available
            if st.session_state.get('full_lineage_complete', False):
                full_lineage_data = st.session_state.get('full_lineage_data', {})
                
                if full_lineage_data and full_lineage_data.get('lineage_table'):
                    st.markdown("### 📋 Full Data Lineage Results")
                    
                    # Show datasets summary
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("📥 Input Datasets", len(full_lineage_data.get('input_datasets', [])))
                    with col2:
                        st.metric("📤 Output Datasets", len(full_lineage_data.get('output_datasets', [])))
                    with col3:
                        st.metric("🔢 Total Variables", full_lineage_data.get('total_variables', 0))
                    
                    # Show datasets
                    if full_lineage_data.get('input_datasets') or full_lineage_data.get('output_datasets'):
                        with st.expander("📁 Datasets Summary", expanded=False):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.markdown("**📥 Input Datasets:**")
                                for dataset in full_lineage_data.get('input_datasets', []):
                                    st.write(f"• {dataset}")
                            
                            with col2:
                                st.markdown("**📤 Output Datasets:**")
                                for dataset in full_lineage_data.get('output_datasets', []):
                                    st.write(f"• {dataset}")
                    
                    # Show lineage table
                    lineage_df = pd.DataFrame(full_lineage_data['lineage_table'])
                    
                    if not lineage_df.empty:
                        # Rename columns for display
                        display_df = lineage_df.copy()
                        display_df.columns = [
                            'Sr. No', 'Variable Name', 'Variable Type', 'Dataset', 
                            'Variable Source', 'Calculation Methodology', 'Explanation', 'Line No'
                        ]
                        
                        # Add filtering options
                        st.markdown("#### 🔍 Data Lineage Table")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            var_type_filter = st.selectbox(
                                "Filter by Variable Type:",
                                ["All"] + list(display_df['Variable Type'].unique()),
                                key="var_type_filter"
                            )
                        
                        with col2:
                            search_filter = st.text_input(
                                "Search Variables:",
                                placeholder="Type to search...",
                                key="var_search_filter"
                            )
                        
                        # Apply filters
                        filtered_df = display_df.copy()
                        
                        if var_type_filter != "All":
                            filtered_df = filtered_df[filtered_df['Variable Type'] == var_type_filter]
                        
                        if search_filter:
                            mask = filtered_df['Variable Name'].str.contains(search_filter, case=False, na=False)
                            filtered_df = filtered_df[mask]
                        
                        # Display table with styling
                        st.dataframe(
                            filtered_df,
                            use_container_width=True,
                            column_config={
                                "Sr. No": st.column_config.NumberColumn("Sr. No", width="small"),
                                "Variable Name": st.column_config.TextColumn("Variable Name", width="medium"),
                                "Variable Type": st.column_config.TextColumn("Variable Type", width="small"),
                                "Dataset": st.column_config.TextColumn("Dataset", width="medium"),
                                "Variable Source": st.column_config.TextColumn("Variable Source", width="medium"),
                                "Calculation Methodology": st.column_config.TextColumn("Calculation Methodology", width="large"),
                                "Explanation": st.column_config.TextColumn("Explanation", width="large"),
                                "Line No": st.column_config.NumberColumn("Line No", width="small")
                            }
                        )



                # ========== ENHANCED MULTI-MODE VISUALIZATION CODE ==========
                st.markdown("---")  # Separator

                # Visualization controls
                col1, col2, col3 = st.columns([2, 1, 1])

                with col1:
                    # View mode selection
                    view_mode = st.selectbox(
                        "🎯 Select View Mode:",
                        ["overview", "detailed", "focused"],
                        index=0,
                        key="lineage_view_mode",
                        help="Overview: High-level flow | Detailed: All connections | Focused: Created variables only"
                    )
                    
                    # Display mode descriptions
                    mode_descriptions = {
                        "overview": "📊 **Overview Mode**: Shows hierarchical flow from input datasets → variables → output datasets",
                        "detailed": "🔍 **Detailed Mode**: Shows all variables and connections with full details",
                        "focused": "🎯 **Focused Mode**: Shows only created variables and their immediate sources"
                    }
                    st.markdown(mode_descriptions[view_mode])

                with col2:
                    # Create visualization button
                    create_viz_btn = st.button(
                        "🎨 Create Visualization", 
                        type="secondary", 
                        use_container_width=True, 
                        key="create_multi_mode_viz_btn"
                    )
                dl_placeholder = col3.empty()

                if st.session_state.get('lineage_viz_html_dict', {}).get(view_mode):
                    viz_html = st.session_state['lineage_viz_html_dict'][view_mode]
                    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                    with dl_placeholder:
                        st.download_button(
                            label="📥 Download HTML",
                            data=viz_html,
                            file_name=f"lineage_{view_mode}_{timestamp}.html",
                            mime="text/html",
                            key=f"download_viz_{view_mode}_btn",
                            help=f"Download {view_mode} visualization as HTML file"
                        )

                # Create visualization when button is clicked
                if create_viz_btn:
                    with st.spinner(f"🎨 Creating {view_mode} visualization..."):
                        try:
                            # Initialize the optimized visualizer
                            visualizer = DataLineageVisualizer()
                            
                            # Show statistics
                            st.markdown("**📊 Visualization Statistics:**")
                            visualizer.display_visualization_stats(full_lineage_data)
                            
                            # Create visualization with selected mode
                            viz_html = visualizer.create_visualization_with_modes(full_lineage_data, view_mode)

                            if viz_html:
                                if 'lineage_viz_html_dict' not in st.session_state:
                                    st.session_state['lineage_viz_html_dict'] = {}

                                st.session_state['lineage_viz_html_dict'][view_mode] = viz_html
                                st.session_state['lineage_viz_mode'] = view_mode

                                # 🔁 Force immediate UI rebuild so the download button appears right away
                                st.rerun()
                            else:
                                st.error("❌ Failed to create visualization")
                                
                        except Exception as e:
                            st.error(f"❌ Visualization error: {str(e)}")
                            st.info("💡 The visualization will continue to work even if some data parsing fails")

                # Display the visualization if it exists for current mode
                if st.session_state.get('lineage_viz_html_dict', {}).get(view_mode):
                    st.markdown(f"#### 🌐 {view_mode.title()} Data Lineage Network")
                    
                    viz_html = st.session_state['lineage_viz_html_dict'][view_mode]
                    
                    # Display the interactive visualization
                    st.components.v1.html(viz_html, height=750, scrolling=True)
                    
                    if view_mode == "overview":
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            **📊 Overview Mode Features:**
                            - DS **Input Datasets** (Top Level - Red circles)
                            - IN **Input Variables** (Level 2 - Green boxes)
                            - INT **Intermediate Variables** (Level 3 - Orange boxes)
                            - CR **Created Variables** (Level 4 - Blue boxes)
                            - OUT **Output Datasets** (Level 5 - Purple circles)
                            """)
                        with col2:
                            st.markdown("""
                            **🖱️ Interaction Tips:**
                            - **Hover**: View variable details
                            - **Click & Drag**: Reposition nodes
                            - **Scroll**: Zoom in/out
                            - **Clear Layout**: Shows data flow hierarchy
                            """)
                    
                    elif view_mode == "detailed":
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            **🔍 Detailed Mode Features:**
                            - **All Variables**: Complete dataset shown
                            - **All Connections**: Every relationship visible
                            - **Full Details**: Hover for calculations
                            - **Physics Layout**: Nodes organize naturally
                            """)
                        with col2:
                            st.markdown("""
                            **🖱️ Advanced Interaction:**
                            - **Drag Nodes**: Customize layout
                            - **Zoom**: Focus on specific areas
                            - **Connected Highlighting**: Click nodes
                            - **Dynamic Layout**: Nodes settle optimally
                            """)
                    
                    else:  # focused mode
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            **🎯 Focused Mode Features:**
                            - **Created Variables**: Main focus (Blue triangles)
                            - **Source Variables**: What feeds into created vars
                            - **Calculation Focus**: Shows transformation details
                            """)
                        with col2:
                            st.markdown("""
                            **🖱️ Focused Interaction:**
                            - **Hover**: See calculation details
                            - **Trace Sources**: Follow blue arrows
                            - **Minimal Clutter**: Only important connections
                            - **Clear Flow**: Input → Transformation → Output
                            """)
                    
                else:
                    # Show placeholder when no visualization exists
                    st.markdown("#### 🌐 Interactive Data Lineage Network")
                    st.info(f"Click '🎨 Create Visualization' above to generate the {view_mode} mode visualization")
                    
                    # Show mode preview
                    mode_previews = {
                        "overview": "📊 **Overview Preview**: Will show hierarchical flow from datasets → input variables → created variables → output datasets",
                        "detailed": "🔍 **Detailed Preview**: Will show all variables and connections with complete relationship mapping",
                        "focused": "🎯 **Focused Preview**: Will show only created variables and their immediate source variables"
                    }
                    st.markdown(mode_previews[view_mode])

                # ========== END OF VISUALIZATION CODE ==========

            # STEP 2: User selects variables for lineage analysis
            st.markdown("### 🎯 Individual Variable Lineage")
            
            # Add a reset button
            col_reset, col_spacer = st.columns([1, 4])
            with col_reset:
                if st.button("🔄 Reset Analysis", help="Start over with new variable discovery"):
                    reset_lineage_state()
                    st.rerun()
            
            # Selection options
            selection_option = st.radio(
                "Choose your analysis scope:",
                [
                    "🎯 Select Specific Variables", 
                    "⚙️ All Created Variables", 
                    "🔄 All Intermediate Variables",
                    "📊 All Variables"
                ],
                help="Choose which variables you want detailed lineage analysis for",
                key="lineage_selection_option"
            )
            
            selected_variables = []
            
            if selection_option == "🎯 Select Specific Variables":
                # Search functionality
                search_term = st.text_input(
                    "🔍 Search Variables:",
                    placeholder="Type to filter variables...",
                    help="Search for specific variables by name",
                    key="lineage_search_term"
                )
                
                # Filter variables based on search
                available_vars = all_variables['all_variables']
                if search_term:
                    available_vars = [var for var in all_variables['all_variables'] 
                                   if search_term.lower() in var.lower()]
                    if available_vars:
                        st.info(f"📋 Found {len(available_vars)} variables matching '{search_term}'")
                    else:
                        st.warning(f"❌ No variables found matching '{search_term}'")
                
                # Multi-select for variables
                if available_vars:
                    selected_variables = st.multiselect(
                        "📋 Select Variables for Detailed Lineage:",
                        options=sorted(available_vars),
                        help="Choose specific variables to trace their creation and transformation",
                        key="lineage_selected_variables"
                    )
                    
                    if selected_variables:
                        st.success(f"✅ Selected {len(selected_variables)} variables: {', '.join(selected_variables[:3])}{'...' if len(selected_variables) > 3 else ''}")
            
            elif selection_option == "⚙️ All Created Variables":
                selected_variables = all_variables.get('created_variables', [])
                if selected_variables:
                    st.info(f"🎯 Will analyze {len(selected_variables)} created variables")
                else:
                    st.warning("No created variables found to analyze")
            
            elif selection_option == "🔄 All Intermediate Variables":
                selected_variables = all_variables.get('intermediate_variables', [])
                if selected_variables:
                    st.info(f"🎯 Will analyze {len(selected_variables)} intermediate variables")
                else:
                    st.warning("No intermediate variables found to analyze")
            
            else:  # All Variables
                selected_variables = all_variables['all_variables']
                st.info(f"🎯 Will analyze all {len(selected_variables)} variables")
            
            # STEP 3: Generate detailed lineage for selected variables
            if selected_variables:
                if st.button("🚀 Generate Detailed Lineage Analysis", type="primary", use_container_width=True, key="generate_lineage_btn"):
                    
                    # Add debugging
                    # debug_session_state()
                    
                    # st.write(f"**🎯 Selected variables for analysis:** {selected_variables}")
                    
                    with st.spinner(f"🤖 AI generating detailed lineage for {len(selected_variables)} variables..."):
                        lineage_analyzer = DataLineageAnalyzer()
                        detailed_lineage = lineage_analyzer.generate_detailed_variable_lineage(
                            code_files, selected_variables, all_variables
                        )
                    
                    if detailed_lineage:
                        # Store results in session state
                        st.session_state['lineage_detailed_results'] = detailed_lineage
                        st.session_state['lineage_selected_vars'] = selected_variables
                        st.session_state['lineage_analysis_complete'] = True
            
            # STEP 4: Display detailed lineage results (if analysis is complete)
            if st.session_state.get('lineage_analysis_complete', False):
                detailed_lineage = st.session_state.get('lineage_detailed_results', {})
                analysis_selected_vars = st.session_state.get('lineage_selected_vars', [])
                
                if detailed_lineage:
                    st.success(f"✅ Detailed lineage generated for {len(detailed_lineage)} variables!")
                    
                    # Display detailed lineage results
                    st.markdown("## 🗺️ Detailed Variable Lineage Analysis")
                    
                    # Summary of analysis
                    st.markdown("### 📊 Analysis Summary")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Variables Analyzed", len(detailed_lineage))
                    with col2:
                        variables_with_sources = sum(1 for v in detailed_lineage.values() if v.get('source_variables'))
                        st.metric("Variables with Sources", variables_with_sources)
                    with col3:
                        variables_with_calculations = sum(1 for v in detailed_lineage.values() if v.get('calculation'))
                        st.metric("Variables with Calculations", variables_with_calculations)
                    
                    # Detailed lineage for each variable

                    for variable, lineage_info in detailed_lineage.items():
                        with st.expander(f"📊 **{variable}** ({lineage_info.get('category', 'unknown').title()})", expanded=True):
                            
                            # Two-column layout 
                            col1, col2 = st.columns([2, 1])
                            
                            with col1:
                                # Show exact code line 
                                if lineage_info.get('exact_code_line'):
                                    st.markdown(f"**📝 Exact Code Line:**")
                                    st.code(lineage_info['exact_code_line'], language='sas')
                                
                                # Main lineage information
                                if lineage_info.get('description'):
                                    st.markdown(f"**🔍 How it's created:** {lineage_info['description']}")
                                
                                if lineage_info.get('calculation'):
                                    st.markdown(f"**📊 Calculation Formula:** `{lineage_info['calculation']}`")
                                
                                if lineage_info.get('source_variables'):
                                    sources = ", ".join(lineage_info['source_variables'])
                                    st.markdown(f"**📥 Source Variables:** `{sources}`")
                                
                                if lineage_info.get('transformation_type'):
                                    st.markdown(f"**⚙️ Transformation Type:** {lineage_info['transformation_type'].title()}")
                                
                                if lineage_info.get('business_purpose'):
                                    st.markdown(f"**💼 Business Purpose:** {lineage_info['business_purpose']}")
                            
                            with col2:
                                # Metadata and technical details
                                st.markdown("**📋 Technical Details:**")
                                st.write(f"**Category:** {lineage_info.get('category', 'unknown').title()}")
                                if lineage_info.get('code_location'):
                                    st.write(f"**Location:** {lineage_info['code_location']}")
                            
                            # Detailed steps if available
                            if lineage_info.get('detailed_steps'):
                                st.markdown("**🔄 Transformation Steps:**")
                                for i, step in enumerate(lineage_info['detailed_steps'], 1):
                                    st.write(f"{i}. {step}")
                    
                    # Export options
                    st.markdown("### 📥 Export Analysis")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.button("📄 Download Lineage (JSON)", key="download_json_btn"):
                            json_data = {
                                'analysis_type': 'detailed_lineage',
                                'selected_variables': analysis_selected_vars,
                                'all_variables': all_variables,
                                'detailed_lineage': detailed_lineage,
                                'analysis_timestamp': pd.Timestamp.now().isoformat()
                            }
                            
                            json_buffer = BytesIO()
                            json_buffer.write(json.dumps(json_data, indent=2).encode())
                            json_buffer.seek(0)
                            
                            st.download_button(
                                label="📥 Download JSON",
                                data=json_buffer,
                                file_name=f"lineage_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.json",
                                mime="application/json"
                            )
                    
                    with col2:
                        if st.button("📊 Download Summary (CSV)", key="download_csv_btn"):
                            # Create summary CSV
                            summary_data = []
                            for variable, info in detailed_lineage.items():
                                summary_data.append({
                                    'Variable': variable,
                                    'Category': info.get('category', ''),
                                    'Source_Variables': ', '.join(info.get('source_variables', [])),
                                    'Calculation': info.get('calculation', ''),
                                    'Description': info.get('description', ''),
                                    'Business_Purpose': info.get('business_purpose', '')
                                })
                            
                            csv_buffer = BytesIO()
                            pd.DataFrame(summary_data).to_csv(csv_buffer, index=False)
                            csv_buffer.seek(0)
                            
                            st.download_button(
                                label="📥 Download CSV",
                                data=csv_buffer,
                                file_name=f"lineage_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
                                mime="text/csv"
                            )
            else:
                # Show message if no variables selected
                if not selected_variables:
                    st.info("ℹ️ Please select at least one variable to proceed with detailed lineage analysis.")
        
        return all_variables
        
    except Exception as e:
        st.error(f"❌ Error in lineage analysis: {str(e)}")
        with st.expander("🐛 Technical Error Details", expanded=False):
            import traceback
            st.code(traceback.format_exc())
        return {}


def perform_code_review(uploaded_files: list):
    """Execute code review analysis with enhanced semantic batching and threading for better performance."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    
    analyzer = CodeAnalyzer()
    all_analyses = []
    
    # Thread-safe progress tracking
    progress_lock = threading.Lock()
    
    def process_snippet_batch_enhanced(batch_snippets, file_name, file_type, batch_id):
        """Process a batch of 6 semantic snippets in parallel."""
        batch_results = []
        
        # Create combined message for all snippets in the batch
        snippets_data = []
        for idx, (snippet, line_range) in enumerate(batch_snippets, 1):
            snippets_data.append({
                "snippet_id": idx,
                "code": snippet.strip(),
                "lines": f"{line_range[0]}-{line_range[1]}"
            })
        
        combined_message = f"""Analyze these {len(snippets_data)} {file_type} code snippets from batch {batch_id}:

{chr(10).join([f"SNIPPET {s['snippet_id']} (Lines {s['lines']}):```{file_type.lower()}{chr(10)}{s['code']}{chr(10)}```{chr(10)}" for s in snippets_data])}

For each snippet, provide both overview and technical annotation.

Respond in this exact JSON format:
{{
    "analysis_results": [
        {{
            "snippet_id": 1,
            "overview": "Simple explanation in 60 words of what this code does",
            "annotation": "Technical bullet points explaining functions, operations, and data transformations"
        }},
        {{
            "snippet_id": 2,
            "overview": "Simple explanation in 60 words of what this code does", 
            "annotation": "Technical bullet points explaining functions, operations, and data transformations"
        }}
    ]
}}

Instructions:
- Overview: Business-friendly explanation without technical jargon (max 60 words)
- Annotation: Technical details with specific functions, operations, transformations (max 150 words)
- Analyze each snippet separately but return all in one JSON response"""

        try:
            response = analyzer.send_message(combined_message, {
                "type": "batch_snippet_analysis",
                "max_tokens": 60000,  
                "temperature": 0.1,
                "batch_id": batch_id
            })
            
            if response:
                try:
                    # Parse JSON response
                    import json
                    clean_response = response.strip()
                    if clean_response.startswith('```json'):
                        clean_response = clean_response[7:]
                    if clean_response.endswith('```'):
                        clean_response = clean_response[:-3]
                        
                    parsed_response = json.loads(clean_response)
                    
                    # Extract results for each snippet
                    analysis_results = parsed_response.get('analysis_results', [])
                    
                    for snippet_data, (snippet, line_range) in zip(analysis_results, batch_snippets):
                        snippet_id = snippet_data.get('snippet_id', len(batch_results) + 1)
                        overview = snippet_data.get('overview', 'Analysis completed')
                        annotation = snippet_data.get('annotation', 'Technical analysis completed')
                        
                        batch_results.append({
                            "sr_no": len(batch_results) + 1,  # Will be renumbered later
                            "code_snippet": snippet.strip(),
                            "line_numbers": f"Lines {line_range[0]}-{line_range[1]}",
                            "code_annotation": annotation,
                            "code_overview": overview
                        })
                        
                except json.JSONDecodeError:
                    # Fallback: process individually if JSON parsing fails
                    st.warning(f"⚠️ JSON parsing failed for batch {batch_id}, falling back to individual processing")
                    for snippet, line_range in batch_snippets:
                        batch_results.append({
                            "sr_no": len(batch_results) + 1,
                            "code_snippet": snippet.strip(),
                            "line_numbers": f"Lines {line_range[0]}-{line_range[1]}",
                            "code_annotation": "Analysis completed (JSON parse failed)",
                            "code_overview": "Analysis completed (JSON parse failed)"
                        })
            else:
                # Fallback for no response
                for snippet, line_range in batch_snippets:
                    batch_results.append({
                        "sr_no": len(batch_results) + 1,
                        "code_snippet": snippet.strip(),
                        "line_numbers": f"Lines {line_range[0]}-{line_range[1]}",
                        "code_annotation": "Error analyzing this snippet",
                        "code_overview": "Error analyzing this snippet"
                    })
                    
        except Exception as e:
            st.error(f"Error analyzing batch {batch_id} in {file_name}: {str(e)}")
            # Fallback for errors
            for snippet, line_range in batch_snippets:
                batch_results.append({
                    "sr_no": len(batch_results) + 1,
                    "code_snippet": snippet.strip(),
                    "line_numbers": f"Lines {line_range[0]}-{line_range[1]}",
                    "code_annotation": "Error analyzing this snippet",
                    "code_overview": "Error analyzing this snippet"
                })
        
        return batch_results
    
    # Process each file
    for file_idx, file in enumerate(uploaded_files):
        st.markdown(f"### 📄 Analysis for {file.name}")
        
        try:
            content = file.read().decode('utf-8')
            file_type = file.name.split('.')[-1].upper()
            
            # Check file size
            line_count = analyzer._count_lines(content)
            if line_count > MAX_CODE_SIZE_FOR_FULL_ANALYSIS:
                st.warning(f"⚠️ Large file detected ({line_count} lines). Using optimized analysis...")
            
            # Generate summary first (unchanged)
            with st.spinner(f"Generating summary for {file.name}..."):
                summary = analyzer.generate_summary(content, file_type, file.name)
            
            # Display summary
            with st.expander("📊 Code Summary", expanded=True):
                if "PURPOSE:" in summary:
                    lines = summary.split('\n')
                    for line in lines:
                        if line.startswith("PURPOSE:"):
                            st.markdown(f"**🎯 Purpose:** {line.replace('PURPOSE:', '').strip()}")
                        elif line.startswith("INPUTS:"):
                            st.markdown(f"**📥 Inputs:** {line.replace('INPUTS:', '').strip()}")
                        elif line.startswith("PROCESS:"):
                            st.markdown(f"**⚙️ Process:** {line.replace('PROCESS:', '').strip()}")
                        elif line.startswith("OUTPUT:"):
                            st.markdown(f"**📤 Output:** {line.replace('OUTPUT:', '').strip()}")
                        elif line.startswith("VALUE:"):
                            st.markdown(f"**💎 Business Value:** {line.replace('VALUE:', '').strip()}")
                else:
                    st.markdown(summary)
            
            # Analyze snippets with enhanced semantic batching
            st.markdown("#### 🔍 Detailed Analysis")
            
            if line_count > MAX_CODE_SIZE_FOR_FULL_ANALYSIS:
                analyzed_snippets = analyzer._analyze_large_code(content, file_type, file.name)
            else:
                # Use new semantic batching approach
                semantic_batches = analyzer._create_code_batches(content, file_type)
                
                if len(semantic_batches) <= 2:
                    # Small number of batches - process normally
                    total_snippets = sum(len(batch[0]) for batch in semantic_batches)
                    with st.spinner(f"Analyzing {total_snippets} code sections in {len(semantic_batches)} batches..."):
                        analyzed_snippets = []
                        for batch_snippets, batch_id in semantic_batches:
                            batch_result = process_snippet_batch_enhanced(batch_snippets, file.name, file_type, batch_id)
                            analyzed_snippets.extend(batch_result)
                else:
                    # Use threading for larger number of batches
                    total_snippets = sum(len(batch[0]) for batch in semantic_batches)
                    st.info(f"🚀 Processing {total_snippets} code sections in {len(semantic_batches)} batches with parallel analysis...")
                    
                    # Progress tracking
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    analyzed_snippets = []
                    completed_batches = 0
                    
                    # Process batches in parallel
                    with ThreadPoolExecutor(max_workers=5) as executor:
                        # Submit all batch jobs
                        future_to_batch = {
                            executor.submit(process_snippet_batch_enhanced, batch_snippets, file.name, file_type, batch_id): batch_id 
                            for batch_snippets, batch_id in semantic_batches
                        }
                        
                        # Collect results as they complete
                        for future in as_completed(future_to_batch):
                            try:
                                batch_results = future.result()
                                analyzed_snippets.extend(batch_results)
                                
                                # Update progress
                                completed_batches += 1
                                progress = completed_batches / len(semantic_batches)
                                progress_bar.progress(progress)
                                status_text.text(f"Completed {completed_batches}/{len(semantic_batches)} semantic batches")
                                
                            except Exception as e:
                                st.error(f"Error processing semantic batch: {str(e)}")
                    
                    # Clear progress indicators
                    progress_bar.empty()
                    status_text.empty()
                    
                    st.success(f"✅ Completed semantic analysis of {len(analyzed_snippets)} code sections!")
            
            if analyzed_snippets:
                # SORT RESULTS BY LINE NUMBER (for threading order issue)
                def extract_start_line(snippet):
                    """Extract starting line number from 'Lines X-Y' format."""
                    try:
                        line_numbers = snippet.get("line_numbers", "Lines 0-0")
                        # Extract first number from "Lines X-Y" format
                        start_line = int(line_numbers.split()[1].split('-')[0])
                        return start_line
                    except (ValueError, IndexError):
                        return 0  # Fallback for malformed line numbers
                
                # Sort snippets by starting line number
                analyzed_snippets.sort(key=extract_start_line)
                
                # Show semantic analysis info
                snippet_info = f"📋 Analyzed {len(analyzed_snippets)} semantic code sections"
                if len(semantic_batches) > 1:
                    snippet_info += f" from {len(semantic_batches)} batches"
                st.info(snippet_info)
                
                # Renumber snippets sequentially after sorting
                for i, snippet in enumerate(analyzed_snippets, 1):
                    snippet["sr_no"] = i
                
                df = pd.DataFrame(analyzed_snippets)
                
                # Add filters for large files
                if len(df) > 10:
                    search_term = st.text_input(
                        "🔍 Search explanations:", 
                        key=f"search_{file_idx}",
                        placeholder="Filter by keywords..."
                    )
                    
                    if search_term:
                        mask = (df['code_annotation'].str.contains(search_term, case=False, na=False) | 
                               df['code_overview'].str.contains(search_term, case=False, na=False))
                        df = df[mask]
                        st.info(f"Showing {len(df)} results matching '{search_term}'")
                
                # Display with enhanced formatting
                st.dataframe(
                    df,
                    use_container_width=True,
                    column_config={
                        "sr_no": st.column_config.NumberColumn("No.", width="small"),
                        "code_snippet": st.column_config.TextColumn("Code", width="large"),
                        "line_numbers": st.column_config.TextColumn("Lines", width="small"),
                        "code_annotation": st.column_config.TextColumn("Technical Annotation", width="large"),
                        "code_overview": st.column_config.TextColumn("Code Overview", width="large")
                    }
                )
                
                all_analyses.append({
                    'file_name': file.name,
                    'summary': summary,
                    'analysis': df,
                    'line_count': line_count
                })
                
        except Exception as e:
            st.error(f"Error analyzing {file.name}: {str(e)}")
            st.info("Tip: Make sure the file is encoded in UTF-8 format")
    
    # Create Excel report 
    if all_analyses:
        st.markdown("### 📊 Generate Report")
        
        col1, col2 = st.columns(2)
        with col1:
            include_code = st.checkbox("Include code snippets in report", value=True)
        with col2:
            exec_summary_only = st.checkbox("Executive summary only", value=False)
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet
            summaries_data = []
            for i, analysis in enumerate(all_analyses):
                summaries_data.append({
                    'Script': f"{analysis['file_name']}",
                    'Lines of Code': analysis['line_count'],
                    'Code Summary': analysis['summary']
                })
            
            summary_df = pd.DataFrame(summaries_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            summary_sheet_name = 'Summary'
            worksheet = writer.sheets[summary_sheet_name]
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 70
            
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(
                        wrap_text=True,
                        vertical='top'
                    )
            
            # Analysis sheets for each script
            if not exec_summary_only:
                for i, analysis in enumerate(all_analyses, 1):
                    sheet_name = f'Script{i}_Analysis'[:31]
                    
                    if include_code:
                        df_to_export = analysis['analysis']
                    else:
                        df_to_export = analysis['analysis'][['sr_no', 'line_numbers', 'code_annotation', 'code_overview']]
                    
                    df_to_export.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    worksheet = writer.sheets[sheet_name]
                    worksheet.column_dimensions['A'].width = 10
                    
                    if include_code:
                        worksheet.column_dimensions['B'].width = 60
                        worksheet.column_dimensions['C'].width = 15
                        worksheet.column_dimensions['D'].width = 70
                        worksheet.column_dimensions['E'].width = 50
                    else:
                        worksheet.column_dimensions['B'].width = 15
                        worksheet.column_dimensions['C'].width = 70
                        worksheet.column_dimensions['D'].width = 50
                    
                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = openpyxl.styles.Alignment(
                                wrap_text=True,
                                vertical='top'
                            )
        
        buffer.seek(0)
        
        filename = f"code_analysis_report.xlsx"
        
        st.download_button(
            label="📥 Download Complete Report",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download the complete report in Excel format"
        )
        
        # Enhanced summary metrics
        st.markdown("### 📈 Analysis Summary")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_files = len(all_analyses)
            st.metric("Files Analyzed", total_files)
        
        with col2:
            total_lines = sum(a['line_count'] for a in all_analyses)
            st.metric("Total Lines of Code", f"{total_lines:,}")
        
        with col3:
            total_snippets = sum(len(a['analysis']) for a in all_analyses)
            st.metric("Semantic Sections", total_snippets)
            
        with col4:
            avg_snippet_size = total_lines / total_snippets if total_snippets > 0 else 0
            st.metric("Avg Section Size", f"{avg_snippet_size:.1f} lines")



def perform_rde_testing(code_files: list, data_files: list):
    """Perform AI-powered RDE testing."""
    rde_tester = RDETester()
    
    # Progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # STEP 1: Extract input dataset variables 
    status_text.text("🔍 Analyzing input datasets (headers only)...")
    input_datasets_info, all_input_vars = rde_tester._extract_input_dataset_variables(data_files)
    
    st.markdown("### 📋 Input Dataset Variables")
    input_summary = []
    for file_name, variables in input_datasets_info.items():
        input_summary.append({
            'Dataset': file_name,
            'Variable Count': len(variables),
            'Sample Variables': ', '.join(list(variables)[:5]) + ('...' if len(variables) > 5 else '')
        })
    
    if input_summary:
        st.dataframe(pd.DataFrame(input_summary), use_container_width=True)
        st.info(f"🔒 **Data Privacy**: Only variable names extracted - no actual data values accessed")
    
    progress_bar.progress(0.15)
    
    # STEP 2: Use AI to analyze script-created variables and cross-script flows
    status_text.text("🤖 Using AI to analyze script variables and cross-script flows...")
    script_created_info = rde_tester._extract_script_created_variables(code_files)
    
    st.markdown("### 🤖 AI-Analyzed Script Variables")
    
    # Display script-created variables
    if script_created_info['created_by_script']:
        script_vars_df = []
        for script_name, variables in script_created_info['created_by_script'].items():
            script_vars_df.append({
                'Script': script_name,
                'Created Variables': len(variables),
                'Variables': ', '.join(variables[:3]) + ('...' if len(variables) > 3 else '')
            })
        st.dataframe(pd.DataFrame(script_vars_df), use_container_width=True)
    else:
        st.info("No script-created variables detected by AI")
    
    # Display cross-script flows
    if script_created_info['cross_script_flows']:
        st.markdown("#### 🔄 Cross-Script Variable Flows (AI-Detected)")
        flow_df = []
        for var_name, flow_info in script_created_info['cross_script_flows'].items():
            flow_df.append({
                'Variable': var_name,
                'Flow Pattern': flow_info,
                'Classification': 'Recoded (Cross-Script)'
            })
        st.dataframe(pd.DataFrame(flow_df), use_container_width=True)
        st.success(f"✅ AI detected {len(script_created_info['cross_script_flows'])} cross-script variable flows")
    else:
        st.info("No cross-script variable flows detected")
    
    progress_bar.progress(0.3)
    
    # STEP 3: Read code content for context 
    status_text.text("📖 Reading code files for context...")
    code_content = ""
    code_info = []
    
    for idx, file in enumerate(code_files):
        content = file.read().decode('utf-8')
        code_content += f"\n\n# --- File: {file.name} ---\n{content}"
        code_info.append({
            'filename': file.name,
            'lines': len(content.split('\n'))
        })
        progress_bar.progress(0.3 + (idx + 1) / len(code_files) * 0.1)
    
    # Display code info
    st.markdown("### 📝 Code Files Summary")
    st.dataframe(pd.DataFrame(code_info), use_container_width=True)
    
    # STEP 4: Process datasets for classification
    datasets_info = []
    all_analyses = []
    
    st.markdown("### 🚀 AI-Powered Dataset Analysis")
    
    for idx, file in enumerate(data_files, 1):
        status_text.text(f"🚀 Analysis of {file.name}...")
        progress_bar.progress(0.4 + (idx / len(data_files) * 0.4))
        
        # Read dataset
        df = rde_tester.read_dataset(file)
        
        if df is not None:
            # Create expander for each dataset
            with st.expander(f"📁 {file.name} - {len(df):,} rows, {len(df.columns)} columns", expanded=True):
                
                # Dataset quality analysis
                quality_report = rde_tester.analyze_dataset_quality(df)
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Quality Score", f"{quality_report['quality_score']:.1f}%")
                with col2:
                    st.metric("Missing Values", f"{len(quality_report['missing_values'])} columns")
                with col3:
                    st.metric("Duplicate Rows", f"{quality_report['duplicate_rows']:,}")
                
                # Store dataset info
                datasets_info.append({
                    "ID": idx,
                    "Input Datasets": file.name,
                    "Number of Rows": len(df),
                    "Number of Columns": len(df.columns),
                    "Quality Score": f"{quality_report['quality_score']:.1f}%"
                })
                
                # ⚡ AI-powered variable analysis with batch processing
                st.write("⚡ **AI Analysis** - Processing variables in batches...")
                
                # Use the optimized analysis method
                variables_analysis = rde_tester.analyze_dataset_variables_optimized(
                    df, code_content, input_datasets_info, script_created_info
                )
                
                
                # Show sample data 
                st.write("**📊 Sample Data**")
                samples = rde_tester.get_intelligent_samples(df, n_samples=5)
                st.dataframe(samples.head(5), use_container_width=True)
                
                all_analyses.append({
                    "id": idx,
                    "variables": pd.DataFrame(variables_analysis),
                    "samples": rde_tester.get_intelligent_samples(df, n_samples=65),
                    "quality": quality_report
                })
    
    progress_bar.progress(0.8)
    status_text.text("📋 Creating AI report...")
    
    # Create Excel report with AI insights
    if all_analyses:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet with AI insights
            summary_data = datasets_info.copy()
            
            # Add AI insights to summary
            total_input_vars = len(all_input_vars)
            total_script_vars = len(script_created_info['all_created_vars'])
            total_cross_script = len(script_created_info['cross_script_flows'])
            
            summary_data.append({
                "ID": "AI Analysis",
                "Input Datasets": f"Input Variables: {total_input_vars}",
                "Number of Rows": f"Script Variables: {total_script_vars}",
                "Number of Columns": f"Cross-Script: {total_cross_script}",
                "Quality Score": "OPTIMIZED AI-Powered"
            })
            
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary
            worksheet = writer.sheets['Summary']
            for col in ['A', 'B', 'C', 'D', 'E']:
                worksheet.column_dimensions[col].width = 25
            
            # Add AI insights sheet
            ai_insights_data = []
            
            # Input dataset insights
            for dataset, vars_set in input_datasets_info.items():
                ai_insights_data.append({
                    'Type': 'Input Dataset',
                    'Source': dataset,
                    'Variable_Count': len(vars_set),
                    'Details': 'Original data elements for testing'
                })
            
            # Script created insights
            for script, vars_list in script_created_info['created_by_script'].items():
                ai_insights_data.append({
                    'Type': 'Script Created',
                    'Source': script,
                    'Variable_Count': len(vars_list),
                    'Details': 'Variables created within script'
                })
            
            # Cross-script flow insights
            for var, flow in script_created_info['cross_script_flows'].items():
                ai_insights_data.append({
                    'Type': 'Cross-Script Flow',
                    'Source': var,
                    'Variable_Count': 1,
                    'Details': flow
                })
            
            if ai_insights_data:
                pd.DataFrame(ai_insights_data).to_excel(writer, sheet_name='AI_Insights', index=False)
            
            # Analysis and sample sheets for each dataset
            for analysis in all_analyses:
                # Variables analysis
                sheet_name = str(analysis["id"])
                analysis["variables"].to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatting with AI-based color coding
                worksheet = writer.sheets[sheet_name]
                worksheet.column_dimensions['A'].width = 30  # DataID
                worksheet.column_dimensions['B'].width = 25  # Classification
                worksheet.column_dimensions['C'].width = 60  # Description
                worksheet.column_dimensions['D'].width = 80  # Screenshot
                
                # Conditional formatting
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                    classification = row[1].value
                    if classification == "Data element to test":
                        row[1].fill = openpyxl.styles.PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        row[1].font = openpyxl.styles.Font(bold=True)
                    elif classification == "Recoded":
                        row[1].fill = openpyxl.styles.PatternFill(
                            start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    elif classification == "Unique Identifier":
                        row[1].fill = openpyxl.styles.PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif classification == "Dates":
                        row[1].fill = openpyxl.styles.PatternFill(
                            start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                
                # Samples sheet
                sample_sheet = f"{analysis['id']}-Sample"
                analysis["samples"].to_excel(writer, sheet_name=sample_sheet, index=False)
                
                # Format samples sheet
                worksheet = writer.sheets[sample_sheet]
                for col in worksheet.column_dimensions.keys():
                    worksheet.column_dimensions[col].width = 20
        
        buffer.seek(0)
        
        progress_bar.progress(1.0)
        status_text.text("⚡ AI RDE report ready!")
        
        # Download button 
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="📥 Download AI RDE Identification Report",
            data=buffer,
            file_name=f"ai_rde_report_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download the AI-powered RDE testing report with intelligent variable classification"
        )
        
        # AI summary statistics
        st.markdown("### ⚡ AI Testing Summary")
        
        total_vars = sum(len(a['variables']) for a in all_analyses)
        rde_count = sum((a['variables']['Classification'] == 'Data element to test').sum() for a in all_analyses)
        recoded_count = sum((a['variables']['Classification'] == 'Recoded').sum() for a in all_analyses)
        ai_detected_flows = len(script_created_info.get('cross_script_flows', {}))
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Variables", total_vars, delta="⚡ OPTIMIZED")
        with col2:
            st.metric("RDE to Test", rde_count, delta="From Input Data")
        with col3:
            st.metric("Recoded Variables", recoded_count, delta="Script Created")
        with col4:
            st.metric("Cross-Script Flows", ai_detected_flows, delta="AI Detected")
        
        # Performance and privacy confirmation
        st.success("✅ AI RDE Identification completed successfully!")
        
        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()





def main():
    """Main application function with enhanced features."""
    
    st.set_page_config(
        page_title="CodeIQ - AI Powered Code Analysis Assistant", 
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Custom CSS for better UI
    st.markdown("""
    <style>
    .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
        font-size: 18px;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
    }
    .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] {
        background-color: #4CAF50;
    }
    div[data-testid="metric-container"] {
        background-color: rgba(28, 131, 225, 0.1);
        border: 1px solid rgba(28, 131, 225, 0.2);
        padding: 10px;
        border-radius: 5px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.title("🧠 Code IQ - AI Powered Code Analysis Assistant")
    st.markdown("*Powered by AVA*")
    
    # Sidebar for information
    with st.sidebar:
        st.header("ℹ️ About This Tool")
        st.markdown("""
        This AI-powered tool helps you:
        - 📝 **Review Code**: Get detailed and business-friendly explanations
        - 🔍 **Test RDE**: Identify data elements for testing
        - 🗺️ **Trace Data Lineage**: Understand and Visualize data flow
        
        All powered by advanced AI for accurate, intuitive results!
        """)
        
        st.header("🚀 What's New")
        st.markdown("""
        - **AI-Powered Lineage**: Intelligent data flow analysis
        - **Large File Support**: Optimized for large files
        - **Business Language**: Non-technical explanations
        - **Interactive Visuals**: Explore data relationships
        """)
    
    # Create tabs for different functionalities
    tab1, tab2, tab3 = st.tabs(["📝 Code Review", "🔍 RDE Identification", "🗺️ Data Lineage"])
    
    # Code Review Tab
    with tab1:
        st.header("Code Review")
        st.markdown("Upload your code files to get AI-powered, business-friendly explanations.")
        
        code_files = st.file_uploader(
            "Upload code files for review",
            accept_multiple_files=True,
            type=['py', 'r', 'sas', 'sql', 'cpp', 'cc', 'cxx', 'java', 'js', 'txt'],
            key="code_review_files",
            help="Supports Python, R, SAS, SQL, C++, Java, JavaScript"
        )
        
        if code_files:
            # Show file info
            st.info(f"📁 {len(code_files)} file(s) uploaded")
            
            # Check for large files
            large_files = []
            for file in code_files:
                content = file.read().decode('utf-8')
                line_count = len(content.split('\n'))
                if line_count > MAX_CODE_SIZE_FOR_FULL_ANALYSIS:
                    large_files.append(f"{file.name} ({line_count} lines)")
                file.seek(0)  # Reset file pointer
            
            if large_files:
                st.warning(f"⚠️ Large files detected: {', '.join(large_files)}. Analysis will be optimized for performance.")
            
            if st.button("🚀 Execute Code Review", type="primary"):
                with st.spinner("Analyzing code... This may take a few minutes."):
                    perform_code_review(code_files)
    
    # RDE Identification Tab
    with tab2:
        st.header("RDE Identification")
        st.markdown("Upload code scripts and datasets to identify raw data elements for testing.")
        
        col1, col2 = st.columns(2)
        
        with col1:
            rde_code_files = st.file_uploader(
                "Upload Code Scripts",
                accept_multiple_files=True,
                type=['py', 'r', 'sas', 'sql'],
                key="rde_code_files",
                help="Upload the code that processes your data"
            )
        
        with col2:
            rde_data_files = st.file_uploader(
                "Upload Input Datasets",
                accept_multiple_files=True,
                type=['xlsx', 'csv', 'sas7bdat', 'parquet'],
                key="rde_data_files",
                help="Upload the data files referenced in your code"
            )
        
        if rde_code_files and rde_data_files:
            st.info(f"📁 {len(rde_code_files)} code file(s) and {len(rde_data_files)} data file(s) uploaded")
            
            if st.button("🚀 Execute RDE Identification", type="primary"):
                with st.spinner("Performing RDE Identification..."):
                    perform_rde_testing(rde_code_files, rde_data_files)
    

    # Data Lineage Tab
    with tab3:
        st.header("🗺️ AI-Powered Data Lineage Analysis")
        st.markdown("Upload your scripts, select variables, and get detailed AI-powered lineage analysis.")
        
        # File upload section
        st.markdown("### 📁 Upload Code Scripts")
        lineage_code_files = st.file_uploader(
            "Choose Code Files",
            accept_multiple_files=True,
            type=['py', 'r', 'sas', 'sql', 'txt'],
            key="lineage_code_files",
            help="Upload Python, R, SAS, SQL, or text files containing your code"
        )
        
        if lineage_code_files:
            # Show upload summary
            st.markdown("### 📋 Upload Summary")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("📄 Files Uploaded", len(lineage_code_files))
            with col2:
                file_types = list(set([f.name.split('.')[-1].upper() for f in lineage_code_files]))
                st.metric("🏷️ File Types", len(file_types))
            with col3:
                total_size = sum(len(f.read()) for f in lineage_code_files)
                for f in lineage_code_files:
                    f.seek(0)  # Reset file pointers
                st.metric("💾 Total Size", f"{total_size/1024:.1f} KB")
            
            # Show file details
            with st.expander("📁 File Details", expanded=False):
                file_details = []
                for file in lineage_code_files:
                    content = file.read().decode('utf-8')
                    file_details.append({
                        'File Name': file.name,
                        'Type': file.name.split('.')[-1].upper(),
                        'Lines': len(content.split('\n')),
                        'Size (KB)': f"{len(content)/1024:.1f}"
                    })
                    file.seek(0)  # Reset
                
                st.dataframe(pd.DataFrame(file_details), use_container_width=True)
            
            # Analysis workflow explanation
            st.markdown("### 🎯 Start AI Data Lineage")
            st.info("""
            **Step 1:** 🤖 AI discovers all variables in your scripts  
            **Step 2:** 🎯 You select which variables you want lineage for or for all variables
            **Step 3:** 🗺️ AI generates detailed lineage with calculations and source variables  
            **Step 4:** 📊 View detailed visualizations, transformation explanations and export results
            """)
            
            # Check if analysis has already been started
            if st.session_state.get('lineage_variables_discovered', False):
                # Analysis is in progress or complete
                perform_data_lineage_analysis(lineage_code_files, [], None)
            else:
                # Start analysis button
                if st.button("🚀 Generate", type="primary", use_container_width=True):
                    try:
                        perform_data_lineage_analysis(lineage_code_files, [], None)
                    except Exception as e:
                        st.error(f"❌ Error during analysis: {str(e)}")
                        st.info("💡 Please check your code syntax and try again.")
        
        else:
            # Welcome section when no files uploaded
            with st.expander("How it works", expanded=False):
                st.markdown("""
            
                    **🤖 AI-Powered Variable Discovery:**
                    - Automatically finds ALL variables in your code
                    - Categorizes them as Input, Created, Intermediate.
                    - Understands multiple programming languages
                    
                    **🎯 User-Controlled Selection:**
                    - Choose specific variables for lineage analysis
                    - Filter by variable type or search by name
                    - Focus on what matters most to you
                    
                    **🗺️ Detailed Lineage Analysis:**
                    - Shows exact calculations and formulas
                    - Identifies source variables for each transformation
                    - Explains business purpose in plain English
                    - Traces complete transformation steps
                    
                    ### 🚀 Supported Languages
                    - **SAS**: Data steps, PROC statements, variable assignments
                    - **Python**: Pandas operations, variable assignments, calculations
                    - **R**: Data.frame operations, dplyr transformations, assignments
                    - **SQL**: SELECT statements, calculated fields, table operations
                    """)
    
    # Footer
    st.markdown("---")
    st.markdown(
        """
        <div style='text-align: center; color: #666;'>
            🤖 <strong>AI-Powered Discovery</strong> | 🎯 <strong>Code Analysis</strong> | 
            🗺️ <strong>Detailed Lineage</strong> | 📊 <strong>Export Ready</strong>
        </div>
        """, 
        unsafe_allow_html=True
    )
    
    # Footer
    st.markdown("---")
    st.markdown(
        """
        <div style='text-align: center'>
            <p>Powered by AVA AI | KPMG © 2024 | 
            <a href='#'>Documentation</a> | 
            <a href='#'>Support</a></p>
        </div>
        """, 
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()


models/Data_Lineage.py:
import streamlit as st
import pandas as pd
import re
from typing import List, Dict, Any, Tuple
from io import BytesIO
import openpyxl
import json
import networkx as nx 
from concurrent.futures import ThreadPoolExecutor, as_completed 
from components.AVA_Assistant import *
from components.functions import *
import plotly.graph_objects as go
from pyvis.network import Network
from typing import Dict, List, Any
import streamlit as st
import pandas as pd

class DataLineageAnalyzer(AVAAssistant):
    """Data lineage analyzer with parallel processing."""
    
    def __init__(self):
        super().__init__()
        self.graph = nx.DiGraph()
        self.lineage_cache = {}
        self.node_colors = {
            'source': '#4CAF50',
            'transform': '#2196F3',
            'output': '#FF9800',
            'variable': '#9C27B0',
            'operation': '#00BCD4'
        }
        # Attributes for processing
        self.compiled_code = ""
        self.file_order = []
        self.code_batches = []
        self.functions_macros_map = {}
    
    def _compile_code_in_logical_order(self, code_files: list) -> None:
        """Compile all code in logical execution order."""
        
        def get_sort_key(filename: str) -> tuple:
            """Generate sort key for logical ordering."""
            name = filename.lower()
            base_name = name.rsplit('.', 1)[0]
            
            # Pattern 1: Numbers (1, 2, 3, 10, 11)
            if re.match(r'^\d+$', base_name):
                return (1, int(base_name), '')
            
            # Pattern 2: Number + letter (1a, 1b, 2a)
            match = re.match(r'^(\d+)([a-z]+)$', base_name)
            if match:
                return (2, int(match.group(1)), match.group(2))
            
            # Pattern 3: Single letters (a, b, c)
            if re.match(r'^[a-z]$', base_name):
                return (3, 0, base_name)
            
            # Pattern 4: Words with numbers (script1, file2, step3)
            match = re.search(r'(\d+)', base_name)
            if match:
                prefix = base_name[:match.start()]
                number = int(match.group(1))
                suffix = base_name[match.end():]
                return (4, number, prefix + suffix)
            
            # Pattern 5: Common ordering words
            order_words = ['main', 'init', 'setup', 'process', 'transform', 'analyze', 'final', 'output']
            for i, word in enumerate(order_words):
                if word in base_name:
                    return (5, i, base_name)
            
            # Pattern 6: Everything else (alphabetical)
            return (6, 0, base_name)
        
        # Sort files logically
        sorted_files = sorted(code_files, key=lambda f: get_sort_key(f.name))
        
        self.compiled_code = ""
        self.file_order = []
        
        for i, file in enumerate(sorted_files, 1):
            content = file.read().decode('utf-8')
            
            # Add to compiled code
            self.compiled_code += f"\n\n# ========== FILE {i}: {file.name} ==========\n"
            self.compiled_code += content
            
            self.file_order.append(file.name)
            file.seek(0)  # Reset for later use
        
        # Create batches of 200 lines each
        self._create_code_batches()
    
    def _create_code_batches(self) -> None:
        """Create batches of code for parallel processing."""
        lines = self.compiled_code.split('\n')
        batch_size = 200
        
        self.code_batches = []
        
        for i in range(0, len(lines), batch_size):
            batch_lines = lines[i:i + batch_size]
            batch_content = '\n'.join(batch_lines)
            
            self.code_batches.append({
                'batch_id': len(self.code_batches) + 1,
                'start_line': i + 1,
                'end_line': i + len(batch_lines),
                'content': batch_content,
                'line_count': len(batch_lines)
            })
    
    def _extract_functions_and_macros(self) -> None:
        """Extract functions and macros for context."""
        
        # Patterns for different languages
        patterns = {
            'python_function': r'def\s+(\w+)\s*\([^)]*\):',
            'python_class': r'class\s+(\w+).*:',
            'sas_macro': r'%macro\s+(\w+)',
            'sas_function': r'proc\s+(\w+)',
            'r_function': r'(\w+)\s*<-\s*function\s*\(',
            'sql_function': r'CREATE\s+(?:OR\s+REPLACE\s+)?FUNCTION\s+(\w+)',
            'sql_procedure': r'CREATE\s+(?:OR\s+REPLACE\s+)?PROCEDURE\s+(\w+)'
        }
        
        for pattern_name, pattern in patterns.items():
            matches = re.findall(pattern, self.compiled_code, re.IGNORECASE | re.MULTILINE)
            if matches:
                self.functions_macros_map[pattern_name] = matches
    
    def _simple_ai_variable_extraction(self, code_files: list) -> Dict[str, Any]:
        """AI-powered variable extraction with parallel processing"""
        
        # Step 1: Compile code in logical order
        st.write("📋 **Compiling code in logical order...**")
        self._compile_code_in_logical_order(code_files)
        
        # Step 2: Extract functions/macros for context
        st.write("🔧 **Extracting functions and macros...**")
        self._extract_functions_and_macros()
        
        # Step 3: Parallel AI variable discovery
        st.write("🤖 **AI discovering variables in parallel batches...**")
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Process batches in parallel 
        all_variables = {
            'input_variables': [],
            'created_variables': [],
            'intermediate_variables': [],
            'all_variables': []
        }
        
        def process_single_batch(batch_info):
            """Process a single batch for variable discovery """
            
            batch_id = batch_info['batch_id']
            content = batch_info['content']
            
            # Include function context
            function_context = ""
            if self.functions_macros_map:
                function_context = f"\n\nFUNCTIONS/MACROS AVAILABLE: {json.dumps(self.functions_macros_map)}"
            
            # Ultra-strict prompt for variable discovery
            prompt = f"""ANALYZE CODE BATCH {batch_id} FOR VARIABLE/COLUMN NAMES. RESPOND ONLY WITH JSON.

CODE BATCH:
{content}{function_context}

FIND VARIABLES/COLUMN NAMES FROM THE CODE BATCH BY CATEGORY. 
CRITICAL: ONLY IDENTIFY DATA COLUMNS AS VARIABLES

R: ONLY columns that are part of dataframes/datasets
✓ data$new_column <- data$old_column * 2
✓ mutate(new_var = old_var + 1)
no constant_value <- 0.05

Python: ONLY columns that are part of pandas DataFrames
✓ df['new_column'] = df['old_column'] * 2
✓ df.assign(new_var = df.old_var + 1)
no interest_rate = 0.05

SAS: ONLY variables in DATA steps that are dataset columns
✓ new_var = existing_var * 0.1; (in data steps)
✓ PROC SQL: SELECT new_var = old_var * 2
no %let macro_var = 2023;

IGNORE THESE COMPLETELY:
- Constants/literals (numbers, strings)
- Macro variables (%let, %global)
- File paths and filenames
- Loop counters (i, j, k)
- Configuration variables
- Function parameters
- Temporary calculations not assigned to datasets


STRICTLY RESPOND WITH ONLY THIS JSON:

{{
  "input_variables": ["var1", "var2"],
  "created_variables": ["var3", "var4"], 
  "intermediate_variables": ["var5", "var6"]
}}

CATEGORIES:
- INPUT: Variables/Columns assumed to be present in the data that are already EXISTING
- CREATED: New variables created through calculations/assignments 
- INTERMEDIATE: Variables modified/transformed from EXISTING/INPUT ones
CRITICAL - DO NOT MISS ANY VARIABLES FROM List in batch

JSON RESPONSE:"""
            
            try:
                response = self.send_message(prompt, {
                    "type": "batch_variable_discovery",
                    "temperature": 0.0,
                    "max_tokens": 60000,
                    "batch_id": batch_id
                })
                
                if response:
                    return self._parse_ai_variable_response(response, batch_id)
                else:
                    st.warning(f"⚠️ No AI response for batch {batch_id}")
                    return {'input_variables': [], 'created_variables': [], 'intermediate_variables': []}
                    
            except Exception as e:
                st.warning(f"⚠️ AI error for batch {batch_id}: {str(e)}")
                return {'input_variables': [], 'created_variables': [], 'intermediate_variables': []}
        
        # Execute parallel processing
        status_text.text("🚀 Processing batches in parallel...")
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            # Submit all batch jobs
            future_to_batch = {
                executor.submit(process_single_batch, batch): batch['batch_id'] 
                for batch in self.code_batches
            }
            
            completed_batches = 0
            
            # Collect results as they complete
            for future in as_completed(future_to_batch):
                try:
                    batch_results = future.result()
                    
                    # Merge results
                    for category in ['input_variables', 'created_variables', 'intermediate_variables']:
                        all_variables[category].extend(batch_results.get(category, []))
                    
                    # Update progress
                    completed_batches += 1
                    progress = completed_batches / len(self.code_batches)
                    progress_bar.progress(progress)
                    status_text.text(f"Completed {completed_batches}/{len(self.code_batches)} batches")
                    
                except Exception as e:
                    st.error(f"Error in batch processing: {str(e)}")
        
        # Remove duplicates and create all_variables list
        for category in ['input_variables', 'created_variables', 'intermediate_variables']:
            all_variables[category] = list(set(all_variables[category]))
            all_variables['all_variables'].extend(all_variables[category])
        
        all_variables['all_variables'] = list(set(all_variables['all_variables']))
        
        progress_bar.empty()
        status_text.empty()
        
        return all_variables
    
    def _parse_ai_variable_response(self, response: str, batch_id: int) -> Dict[str, List[str]]:
        """Parse AI response for variables - robust JSON parsing."""
        
        variables = {'input_variables': [], 'created_variables': [], 'intermediate_variables': []}
        
        # Multiple parsing attempts with progressive cleaning
        for attempt in range(3):
            try:
                clean_response = response.strip()
                
                if attempt == 1:
                    # Remove markdown formatting
                    clean_response = re.sub(r'```\w*', '', clean_response)
                    clean_response = clean_response.replace('```', '')
                elif attempt == 2:
                    # Extract JSON block
                    json_match = re.search(r'\{.*\}', clean_response, re.DOTALL)
                    if json_match:
                        clean_response = json_match.group()
                
                # Parse JSON
                data = json.loads(clean_response)
                
                # Extract variables by category
                for category in ['input_variables', 'created_variables', 'intermediate_variables']:
                    if category in data and isinstance(data[category], list):
                        # Clean variable names
                        clean_vars = []
                        for var in data[category]:
                            if isinstance(var, str) and var.strip():
                                clean_vars.append(var.strip())
                        variables[category] = clean_vars
                
                st.write(f"✅ Batch {batch_id}: AI found {sum(len(v) for v in variables.values())} variables")
                return variables
                
            except json.JSONDecodeError:
                if attempt == 2:  # Last attempt
                    st.warning(f"⚠️ AI response not valid JSON for batch {batch_id}")
                continue
        
        return variables
    

    def generate_detailed_variable_lineage(self, code_files: list, selected_variables: List[str], all_variables: Dict[str, Any]) -> Dict[str, Any]:
        """Generate detailed lineage for selected variables using AI analysis"""
        
        if not selected_variables:
            return {}
        
        st.write(f"🗺️ **Generating detailed lineage for {len(selected_variables)} variables...**")
        
        # Group variables into batches 
        batch_size = 1
        variable_batches = []
        
        for i in range(0, len(selected_variables), batch_size):
            batch = selected_variables[i:i + batch_size]
            variable_batches.append({
                'batch_id': len(variable_batches) + 1,
                'variables': batch
            })
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        detailed_lineage = {}
        
        def process_lineage_batch(batch_info):
            """Process a batch of variables for lineage analysis"""
            
            batch_id = batch_info['batch_id']
            variables = batch_info['variables']
            
            # Include function context
            function_context = ""
            if self.functions_macros_map:
                function_context = f"\n\nFUNCTIONS/MACROS AVAILABLE: {json.dumps(self.functions_macros_map)}"
            
            # Create variable analysis prompt
            variables_str = ", ".join(variables)
            
            # Combine all code
            all_code = ""
            file_info = []
            
            for file in code_files:
                content = file.read().decode('utf-8')
                file_type = file.name.split('.')[-1].upper()
                all_code += f"\n\n=== FILE: {file.name} ({file_type}) ===\n{content}"
                file_info.append(file.name)
                file.seek(0)
            
            all_code = ""
            for file in code_files:
                content = file.read().decode('utf-8')
                all_code += f"\n{content}\n"
                file.seek(0)

            detailed_lineage = {}
            
            for variable in selected_variables:
                st.write(f"🔍 Analyzing {variable}...")
                
                # Search for variable assignment lines
                variable_lines = []
                code_lines = all_code.split('\n')
                
                for i, line in enumerate(code_lines, 1):
                    line_clean = line.strip()
                    # Check if variable appears before = sign (meaning it's being assigned)
                    if variable in line_clean and '=' in line_clean:
                        equals_pos = line_clean.find('=')
                        variable_pos = line_clean.find(variable)
                        if variable_pos < equals_pos and variable_pos >= 0:
                            variable_lines.append(f"Line {i}: {line_clean}")
                
                # Fallback search if no assignment lines found
                if not variable_lines:
                    for i, line in enumerate(code_lines, 1):
                        if variable in line.strip():
                            variable_lines.append(f"Line {i}: {line.strip()}")
                
                # Prepare code context for AI
                if variable_lines:
                    specific_code = "\n".join([line.split(": ", 1)[1] for line in variable_lines if ": " in line])
                else:
                    specific_code = all_code
            
            lineage_prompt = f"""ANALYZE LINEAGE FOR ALL OF THESE VARIABLES. RESPOND ONLY WITH JSON.

VARIABLES TO ANALYZE: {variables_str}
    CODE:
    {all_code}

FOR EACH VARIABLE, PROVIDE DETAILED LINEAGE. DO NOT MISS ANY VARIABLE.
RESPOND WITH ONLY THIS JSON:

{{
  "lineage_results": [
    {{
      "variable": "variable_name",
      "category": "input|created|intermediate",
      "source_variables": ["source_var1", "source_var2"], 
      "dataset": ["dataset1"],
      "calculation": "exact_formula_or_logic",
      "description": "business_explanation",
      "business_purpose": "why_this_variable_exists",
      "code_location": "file_and_line_reference",
      "detailed_steps": ["step1", "step2", "step3"]
    }}
  ]
}}

FIND:
- Exact and detailed calculation formulas
- Source variables used
- Dataset: ASSIGNMENT RULES (CRITICAL - FOLLOW EXACTLY):
Rule 1: IF variable "category" is "input" → "dataset" field MUST be INPUT dataset name
- Look for: pd.read_csv, SET dataset, FROM table, IMPORT commands
- Example: "customer_data.csv", "sales_table", "input_file.xlsx"

Rule 2: IF variable "category" is "created" OR "intermediate" → "dataset" field MUST be OUTPUT dataset name  
- Look for: to_csv, DATA output, CREATE TABLE, EXPORT commands
- Example: "results.csv", "output_table", "final_report.xlsx"

Rule 3: IF no clear dataset found → Use "N/A" but NEVER mix input/output categories

Rule 4: There can be a single input/output dataset for a variable

- Business purpose explanation
- Step-by-step transformation logic
- If there are source_variables found for the variable being processed, then try to get all the details for all the 
    source_variables as well along with their exact formulae.
- Do not miss any of the 3 variable from the list

JSON RESPONSE:"""
            
            try:
                response = self.send_message(lineage_prompt, {
                    "type": "batch_lineage_analysis",
                    "temperature": 0.0,
                    "max_tokens": 60000,
                    "batch_id": batch_id
                })
                
                if response:
                    return self._parse_ai_lineage_response(response, batch_id)
                else:
                    st.warning(f"⚠️ No AI response for lineage batch {batch_id}")
                    return {}
                    
            except Exception as e:
                st.warning(f"⚠️ AI error in lineage batch {batch_id}: {str(e)}")
                return {}
        
        # Process lineage batches in parallel
        status_text.text("🚀 Processing lineage batches in parallel...")
        
        with ThreadPoolExecutor(max_workers=1) as executor:
            future_to_batch = {
                executor.submit(process_lineage_batch, batch): batch['batch_id']
                for batch in variable_batches
            }
            
            completed_batches = 0
            
            for future in as_completed(future_to_batch):
                try:
                    batch_results = future.result()
                    detailed_lineage.update(batch_results)
                    
                    completed_batches += 1
                    progress = completed_batches / len(variable_batches)
                    progress_bar.progress(progress)
                    status_text.text(f"Analyzed {completed_batches}/{len(variable_batches)} lineage batches")
                    
                except Exception as e:
                    st.error(f"Error in lineage batch processing: {str(e)}")
        
        progress_bar.empty()
        status_text.empty()
        
        st.success(f"✅ AI generated detailed lineage for {len(detailed_lineage)} variables")
        
        return detailed_lineage
    
    def _parse_ai_lineage_response(self, response: str, batch_id: int) -> Dict[str, Any]:
        """Parse AI lineage response - robust JSON parsing."""
        
        lineage_results = {}
        
        # Multiple parsing attempts
        for attempt in range(3):
            try:
                clean_response = response.strip()
                
                if attempt == 1:
                    # Remove markdown
                    clean_response = re.sub(r'```\w*', '', clean_response)
                    clean_response = clean_response.replace('```', '')
                elif attempt == 2:
                    # Extract JSON
                    json_match = re.search(r'\{.*\}', clean_response, re.DOTALL)
                    if json_match:
                        clean_response = json_match.group()
                
                data = json.loads(clean_response)
                
                if 'lineage_results' in data and isinstance(data['lineage_results'], list):
                    for item in data['lineage_results']:
                        if isinstance(item, dict) and 'variable' in item:
                            variable = item['variable']
                            lineage_results[variable] = {
                                'variable': variable,
                                'category': item.get('category', 'unknown'),
                                'source_variables': item.get('source_variables', []),
                                'dataset': item.get('dataset', []),
                                'calculation': item.get('calculation', ''),
                                'description': item.get('description', ''),
                                'business_purpose': item.get('business_purpose', ''),
                                'code_location': item.get('code_location', ''),
                                'detailed_steps': item.get('detailed_steps', [])
                            }
                    
                    st.write(f"✅ Lineage batch {batch_id}: AI analyzed {len(lineage_results)} variables")
                    return lineage_results
                    
            except json.JSONDecodeError:
                if attempt == 2:
                    st.warning(f"⚠️ AI lineage response not valid JSON for batch {batch_id}")
                continue
        
        return lineage_results
    
    def generate_full_data_lineage(self, all_variables, code_files: list, data_files: list = None) -> Dict[str, Any]:
        """Generate comprehensive data lineage."""
        
        try:
            st.write("🤖 **Starting AI-powered full data lineage analysis...**")
            
            if not all_variables.get('all_variables'):
                return {
                    'error': 'No variables found by AI',
                    'lineage_table': [],
                    'input_datasets': [],
                    'output_datasets': []
                }
            
            # Step 1: Generate lineage for all discovered variables
            all_discovered_vars = all_variables['all_variables']
            
            detailed_lineage = self.generate_detailed_variable_lineage(
                code_files, all_discovered_vars, all_variables
            )
            
            # Step 2: Convert to table format
            lineage_table = []
            for sr_no, (variable, info) in enumerate(detailed_lineage.items(), 1):
                lineage_table.append({
                    'sr_no': sr_no,
                    'variable_name': variable,
                    'variable_type': info.get('category', 'Unknown').title(),
                    'dataset': info.get('dataset', 'N/A'),
                    'variable_source': ', '.join(info.get('source_variables', [])) if info.get('source_variables') else 'N/A',
                    'calculation_methodology': info.get('calculation', 'N/A'),
                    'explanation': info.get('description', 'N/A'),
                    'line_no': info.get('code_location', 'N/A')
                })
            
            # Step 3: Dataset extraction
            input_datasets, output_datasets = self._ai_extract_datasets(code_files)
            
            return {
                'lineage_table': lineage_table,
                'input_datasets': input_datasets,
                'output_datasets': output_datasets,
                'total_variables': len(lineage_table),
                'discovered_variables': len(all_variables['all_variables']),
                'processing_method': "ai_only_parallel_approach",
                'file_order': self.file_order,
                'functions_macros': self.functions_macros_map
            }
            
        except Exception as e:
            st.error(f"❌ Error in AI lineage analysis: {str(e)}")
            return {
                'error': str(e),
                'lineage_table': [],
                'input_datasets': [],
                'output_datasets': []
            }
    
    def _ai_extract_datasets(self, code_files:list) -> Tuple[List[str], List[str]]:
        """Extract input and output datasets using multiple robust methods."""
        
        input_datasets = []
        output_datasets = []
        
        # Combine all code
        all_code = ""
        file_info = []
        
        for file in code_files:
            content = file.read().decode('utf-8')
            file_type = file.name.split('.')[-1].upper()
            all_code += f"\n\n=== FILE: {file.name} ({file_type}) ===\n{content}"
            file_info.append(file.name)
            file.seek(0)
        
        all_code = ""
        for file in code_files:
            content = file.read().decode('utf-8')
            all_code += f"\n{content}\n"
            file.seek(0)
                
        # Method 1: Try AI first
        try:
            dataset_prompt = f"""Find dataset names in this code. Return simple lists.
    CODE:
    {all_code}

    Find INPUT datasets (files being read) and OUTPUT datasets (files being written).
    Just list the dataset names, one per line:

    INPUT DATASETS:
    [list here]

    OUTPUT DATASETS:  
    [list here]"""
            
            response = self.send_message(dataset_prompt, {
                "type": "simple_dataset_extraction",
                "temperature": 0.0,
                "max_tokens": 60000
            })
            
            if response:
                ai_input, ai_output = self._simple_parse_ai_response(response)
                input_datasets.extend(ai_input)
                output_datasets.extend(ai_output)
                
        except Exception as e:
            st.warning(f"⚠️ AI extraction failed: {str(e)}")
        
        # Method 2: Simple regex patterns (always try this as backup)
        try:
            regex_input, regex_output = self._regex_extract_datasets()
            input_datasets.extend(regex_input)
            output_datasets.extend(regex_output)
            
        except Exception as e:
            st.warning(f"⚠️ Regex extraction failed: {str(e)}")
        
        # Method 3: Keyword search (most basic fallback)
        try:
            keyword_input, keyword_output = self._keyword_extract_datasets()
            input_datasets.extend(keyword_input)
            output_datasets.extend(keyword_output)
            
        except Exception as e:
            st.warning(f"⚠️ Keyword extraction failed: {str(e)}")
        
        # Clean up and deduplicate (but keep names even if not perfectly clean)
        input_datasets = self._gentle_clean_dataset_list(input_datasets)
        output_datasets = self._gentle_clean_dataset_list(output_datasets)
        
        st.info(f"📊 Found {len(input_datasets)} input datasets, {len(output_datasets)} output datasets")
        
        return input_datasets, output_datasets

    def _simple_parse_ai_response(self, response: str) -> Tuple[List[str], List[str]]:
        """Simple text parsing - don't try to be too clever."""
        
        input_datasets = []
        output_datasets = []
        
        lines = response.split('\n')
        current_section = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if 'INPUT' in line.upper():
                current_section = 'input'
            elif 'OUTPUT' in line.upper():
                current_section = 'output'
            elif current_section == 'input' and line not in ['INPUT DATASETS:', 'INPUT:']:
                input_datasets.append(line)
            elif current_section == 'output' and line not in ['OUTPUT DATASETS:', 'OUTPUT:']:
                output_datasets.append(line)
        
        return input_datasets, output_datasets

    def _regex_extract_datasets(self) -> Tuple[List[str], List[str]]:
        """Extract datasets using simple regex patterns."""
        
        input_datasets = []
        output_datasets = []
        
        # Simple patterns for common operations
        input_patterns = [
            r'read_csv\([\'"]([^\'"\)]+)[\'"]',
            r'read_excel\([\'"]([^\'"\)]+)[\'"]',
            r'set\s+([a-zA-Z_][a-zA-Z0-9_.]*)',
            r'from\s+([a-zA-Z_][a-zA-Z0-9_.]*)',
        ]
        
        output_patterns = [
            r'to_csv\([\'"]([^\'"\)]+)[\'"]',
            r'to_excel\([\'"]([^\'"\)]+)[\'"]',
            r'data\s+([a-zA-Z_][a-zA-Z0-9_.]*)',
            r'create\s+table\s+([a-zA-Z_][a-zA-Z0-9_.]*)',
        ]
        
        # Search in code
        for pattern in input_patterns:
            matches = re.findall(pattern, self.compiled_code, re.IGNORECASE)
            input_datasets.extend(matches)
        
        for pattern in output_patterns:
            matches = re.findall(pattern, self.compiled_code, re.IGNORECASE)
            output_datasets.extend(matches)
        
        return input_datasets, output_datasets

    def _keyword_extract_datasets(self) -> Tuple[List[str], List[str]]:
        """Extract datasets using keyword search - most basic method."""
        
        input_datasets = []
        output_datasets = []
        
        lines = self.compiled_code.split('\n')
        
        for line in lines:
            line_lower = line.lower()
            
            # Look for input keywords
            if any(keyword in line_lower for keyword in ['read', 'input', 'load', 'import']):
                # Try to extract anything that looks like a filename
                potential_names = re.findall(r'[a-zA-Z_][a-zA-Z0-9_]*', line)
                input_datasets.extend(potential_names[:2])  # Take first 2 candidates
            
            # Look for output keywords  
            if any(keyword in line_lower for keyword in ['write', 'output', 'save', 'export']):
                potential_names = re.findall(r'[a-zA-Z_][a-zA-Z0-9_]*', line)
                output_datasets.extend(potential_names[:2])  # Take first 2 candidates
        
        return input_datasets, output_datasets

    def _gentle_clean_dataset_list(self, dataset_list: List[str]) -> List[str]:
        """Gently clean dataset names - don't be too strict."""
        
        cleaned = []
        seen = set()
        
        for name in dataset_list:
            if not name:
                continue
                
            # Basic cleaning - remove obvious junk but keep most things
            clean_name = str(name).strip()
            
            # Skip very short names or obvious keywords
            if len(clean_name) < 2 or clean_name.lower() in ['set', 'data', 'from', 'to', 'read', 'write']:
                continue
                
            # Keep the name (even with extensions, prefixes, etc.)
            if clean_name not in seen:
                cleaned.append(clean_name)
                seen.add(clean_name)
        
        return cleaned


            
    def create_full_lineage_excel(self, lineage_data: Dict[str, Any]) -> BytesIO:
        """Create Excel file with AI-generated full data lineage analysis."""
        buffer = BytesIO()
        
        try:
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Sheet 1: Summary with dataset info
                summary_data = []
                
                # Add input datasets if available
                detailed_input = lineage_data.get('detailed_input_datasets', [])
                detailed_output = lineage_data.get('detailed_output_datasets', [])
                
                if detailed_input:
                    for dataset in detailed_input:
                        summary_data.append({
                            'Item Type': 'Input Dataset',
                            'Name': dataset.get('name', 'Unknown'),
                            'Operation': dataset.get('operation', 'N/A'),
                            'Details': 'Source data for analysis'
                        })
                else:
                    # Fallback to simple list
                    for dataset in lineage_data.get('input_datasets', []):
                        summary_data.append({
                            'Item Type': 'Input Dataset',
                            'Name': dataset,
                            'Operation': 'READ/INPUT',
                            'Details': 'Source data identified by AI'
                        })
                
                if detailed_output:
                    for dataset in detailed_output:
                        summary_data.append({
                            'Item Type': 'Output Dataset',
                            'Name': dataset.get('name', 'Unknown'),
                            'Operation': dataset.get('operation', 'N/A'),
                            'Details': 'Result data from analysis'
                        })
                else:
                    # Fallback to simple list
                    for dataset in lineage_data.get('output_datasets', []):
                        summary_data.append({
                            'Item Type': 'Output Dataset',
                            'Name': dataset,
                            'Operation': 'WRITE/OUTPUT',
                            'Details': 'Output data identified by AI'
                        })
                
                # Create summary sheet
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                else:
                    summary_df = pd.DataFrame({
                        'Item Type': ['No datasets found'],
                        'Name': ['AI analysis did not detect input/output operations'],
                        'Operation': ['N/A'],
                        'Details': ['Check if code contains file I/O operations']
                    })
                
                summary_df.to_excel(writer, sheet_name='Datasets_Summary', index=False)
                
                # Format summary sheet
                worksheet = writer.sheets['Datasets_Summary']
                worksheet.column_dimensions['A'].width = 15  # Item Type
                worksheet.column_dimensions['B'].width = 30  # Name
                worksheet.column_dimensions['C'].width = 15  # Operation
                worksheet.column_dimensions['D'].width = 50  # Details
                
                # Sheet 2: Enhanced Data Lineage
                lineage_table = lineage_data.get('lineage_table', [])
                
                if lineage_table and len(lineage_table) > 0:
                    lineage_df = pd.DataFrame(lineage_table)
                    
                    # Ensure all required columns exist
                    required_columns = [
                        'sr_no', 'variable_name', 'variable_type', 'dataset',
                        'variable_source', 'calculation_methodology', 'explanation', 'line_no'
                    ]
                    
                    for col in required_columns:
                        if col not in lineage_df.columns:
                            lineage_df[col] = 'N/A'
                    
                    # Rename columns for Excel
                    lineage_df = lineage_df.rename(columns={
                        'sr_no': 'Sr. No',
                        'variable_name': 'Variable Name',
                        'variable_type': 'Variable Type',
                        'dataset': 'Dataset',
                        'variable_source': 'Variable Source',
                        'calculation_methodology': 'Calculation Methodology',
                        'explanation': 'Explanation',
                        'line_no': 'Line No'
                    })
                    
                    # Select and order columns
                    final_columns = [
                        'Sr. No', 'Variable Name', 'Variable Type', 'Dataset',
                        'Variable Source', 'Calculation Methodology', 'Explanation', 'Line No']
                    
                    lineage_df = lineage_df[final_columns]

                    # Define the custom order for sorting
                    category_order = ['Input', 'Intermediate', 'Created']
                    lineage_df['Variable Type'] = pd.Categorical(
                        lineage_df['Variable Type'],
                        categories=category_order,
                        ordered=True)
    
                    # Sort the DataFrame by the custom order
                    lineage_df = lineage_df.sort_values('Variable Type')
                    
                else:
                    # Create empty lineage sheet
                    lineage_df = pd.DataFrame({
                        'Sr. No': [1],
                        'Variable Name': ['No variables found'],
                        'Variable Type': ['N/A'],
                        'Dataset': ['N/A'],
                        'Variable Source': ['N/A'],
                        'Calculation Methodology': ['AI could not detect variable operations'],
                        'Explanation': ['Please check if code contains variable assignments'],
                        'Line No': ['N/A']
                    })
                
                lineage_df.to_excel(writer, sheet_name='AI_Data_Lineage', index=False)
                
                # Format lineage sheet
                worksheet = writer.sheets['AI_Data_Lineage']
                worksheet.column_dimensions['A'].width = 8   # Sr. No
                worksheet.column_dimensions['B'].width = 25  # Variable Name
                worksheet.column_dimensions['C'].width = 15  # Variable Type
                worksheet.column_dimensions['D'].width = 20  # Dataset
                worksheet.column_dimensions['E'].width = 25  # Variable Source
                worksheet.column_dimensions['F'].width = 40  # Calculation Methodology
                worksheet.column_dimensions['G'].width = 50  # Explanation
                worksheet.column_dimensions['H'].width = 10  # Line No
                
                # Apply formatting and color coding
                try:
                    # Create alignment style once
                    wrap_alignment = openpyxl.styles.Alignment(
                        wrap_text=True,
                        vertical='top'
                    )
                    
                    # Apply bounds checking
                    max_row = worksheet.max_row
                    max_col = worksheet.max_column
                    
                    if max_row > 1 and max_col > 0:
                        for row_idx in range(2, min(max_row + 1, 1500)):  
                            for col_idx in range(1, min(max_col + 1, 20)):  
                                try:
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    if cell.value is not None:  
                                        cell.alignment = wrap_alignment
                                except Exception:
                                    continue  # Skip problematic cells
                                    
                except Exception as e:
                    # Don't fail Excel creation for formatting issues
                    pass
                
                # Color coding by variable type
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                    if len(row) > 2:
                        var_type = row[2].value 
                        if var_type == "Input":
                            row[2].fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif var_type == "Created":
                            row[2].fill = openpyxl.styles.PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                        elif var_type == "Intermediate":
                            row[2].fill = openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif var_type == "Output":
                            row[2].fill = openpyxl.styles.PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")  
            
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            st.error(f"Excel creation error: {str(e)}")
            
            # Create simple fallback Excel
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                fallback_df = pd.DataFrame({
                    'Status': ['Excel creation failed'],
                    'Error': [str(e)],
                    'Solution': ['Please try the analysis again'],
                    'AI_Response_Available': ['Yes' if lineage_data.get('ai_response_length', 0) > 0 else 'No']
                })
                fallback_df.to_excel(writer, sheet_name='Error_Report', index=False)
            
            buffer.seek(0)
            return buffer
        
    def debug_ai_connection(self, test_message: str = "Hello, can you respond?") -> bool:
        """Test if AI connection is working properly."""
        
        st.write("**🔧 Testing AI Connection...**")
        
        response = self.send_message(test_message, {"type": "connection_test"})
        
        if response:
            st.success(f"✅ AI Connection Working")
            return True
        else:
            st.error("❌ AI Connection Failed")
            return False



class DataLineageVisualizer:
    
    def __init__(self):
        self.graph = nx.DiGraph()
        self.view_modes = ['overview', 'detailed', 'focused']
        self.current_mode = 'detailed'
        
        # Hierarchical positioning (Y-levels)
        self.hierarchy_levels = {
            'input_datasets': 1,
            'input_variables': 2,
            'intermediate_variables': 3,
            'created_variables': 4,
            'output_datasets': 5
        }
        
        # Color scheme for different node types
        self.node_colors = {
            'input_dataset': '#E91E63',      # Pink
            'input_variable': '#4CAF50',     # Green 
            'intermediate_variable':'#FF9800',  # Orange 
            'created_variable': '#2196F3',   # Blue 
            'output_dataset': '#9C27B0',     # Purple
            'unknown': '#757575'             # Gray
        }
        
        # Node shapes and sizes
        self.node_config = {
            'input_dataset': {'shape': 'circle', 'size': 60},
            'input_variable': {'shape': 'box', 'size': 25},
            'intermediate_variable': {'shape': 'box', 'size': 28},
            'created_variable': {'shape': 'box', 'size': 30},
            'output_dataset': {'shape': 'circle', 'size': 60},
            'unknown': {'shape': 'dot', 'size': 20}
        }
        
        # Text icons
        self.node_icons = {
            'input_dataset': '[DS]',
            'input_variable': '[IN]',
            'intermediate_variable': '[INT]',
            'created_variable': '[CR]',
            'output_dataset': '[OUT]',
            'unknown': '[?]'
        }
    
    def parse_uploaded_lineage_excel(self, uploaded_file):
        """Parse uploaded Excel file to extract lineage data."""
        
        try:
            # Read the Excel file
            df = pd.read_excel(uploaded_file, sheet_name='AI_Data_Lineage')
            
            # Convert to lineage_table format
            lineage_table = []
            for _, row in df.iterrows():
                lineage_table.append({
                    'sr_no': row.get('Sr. No', ''),
                    'variable_name': row.get('Variable Name', ''),
                    'variable_type': row.get('Variable Type', ''),
                    'dataset': row.get('Dataset', 'N/A'),
                    'variable_source': row.get('Variable Source', 'N/A'),
                    'calculation_methodology': row.get('Calculation Methodology', ''),
                    'explanation': row.get('Explanation', ''),
                    'line_no': row.get('Line No', '')
                })
            
            # Try to read datasets from other sheets
            input_datasets = []
            output_datasets = []
            
            try:
                # Read datasets summary if available
                datasets_df = pd.read_excel(uploaded_file, sheet_name='Datasets_Summary')
                for _, row in datasets_df.iterrows():
                    item_type = str(row.get('Item Type', '')).lower()
                    name = str(row.get('Name', ''))
                    if 'input' in item_type and name:
                        input_datasets.append(name)
                    elif 'output' in item_type and name:
                        output_datasets.append(name)
            except:
                # If no datasets summary, extract from lineage table
                for row in lineage_table:
                    dataset = str(row.get('dataset', '')).strip()
                    if dataset and dataset not in ['N/A', 'nan', '']:
                        if dataset not in input_datasets:
                            input_datasets.append(dataset)
            
            # Create lineage data structure
            uploaded_lineage_data = {
                'lineage_table': lineage_table,
                'input_datasets': input_datasets,
                'output_datasets': output_datasets,
                'total_variables': len(lineage_table),
                'processing_method': 'uploaded_excel',
                'file_name': uploaded_file.name
            }
            
            return uploaded_lineage_data
            
        except Exception as e:
            st.error(f"❌ Error parsing uploaded Excel: {str(e)}")
            return None
    
    def _robust_parse_lineage_data(self, lineage_data):
        """Robustly parse lineage data with extensive error handling."""
        
        parsed_data = {
            'nodes': {},
            'edges': [],
            'datasets': {'input': set(), 'output': set()},
            'variables': {'input': [], 'intermediate': [], 'created': [], 'output': []},
            'parsing_stats': {'success': 0, 'errors': 0, 'warnings': []}
        }
        
        try:
            # Get lineage table
            lineage_table = lineage_data.get('lineage_table', [])
            if not lineage_table:
                parsed_data['parsing_stats']['warnings'].append("No lineage table found")
                return parsed_data
            
            # Parse each row with error handling
            for row_idx, row in enumerate(lineage_table):
                try:
                    self._parse_single_row(row, row_idx, parsed_data)
                    parsed_data['parsing_stats']['success'] += 1
                except Exception as e:
                    parsed_data['parsing_stats']['errors'] += 1
                    parsed_data['parsing_stats']['warnings'].append(f"Row {row_idx}: {str(e)}")
                    continue  # Skip this row, don't fail
            
            # Parse input/output datasets from lineage_data
            self._parse_datasets(lineage_data, parsed_data)
            
            # Create connections with corrected intermediate variable arrows
            self._create_smart_connections(parsed_data)
            
        except Exception as e:
            parsed_data['parsing_stats']['warnings'].append(f"Major parsing error: {str(e)}")
        
        return parsed_data
    
    def _parse_single_row(self, row, row_idx, parsed_data):
        """Parse a single row with robust error handling."""
        
        try:
            # Extract basic info with fallbacks
            var_name = str(row.get('variable_name', f'var_{row_idx}')).strip()
            var_type = str(row.get('variable_type', 'unknown')).lower().strip()
            dataset = str(row.get('dataset', 'N/A')).strip()
            variable_source = str(row.get('variable_source', 'N/A')).strip()
            calculation = str(row.get('calculation_methodology', '')).strip()
            explanation = str(row.get('explanation', '')).strip()
            
            if not var_name or var_name == 'nan':
                return  # Skip invalid rows
            
            # Determine node type with exact case matching
            node_type = self._determine_node_type(var_type)
            
            # Create node
            node_id = f"VAR_{var_name}"
            parsed_data['nodes'][node_id] = {
                'id': node_id,
                'label': var_name,
                'type': node_type,
                'original_type': var_type,
                'dataset': dataset,
                'variable_source': variable_source,
                'calculation': calculation,
                'explanation': explanation,
                'level': self.hierarchy_levels.get(f'{node_type}_variables', 3)
            }
            
            # Categorize variable
            if node_type == 'input_variable':
                parsed_data['variables']['input'].append(node_id)
            elif node_type == 'intermediate_variable':
                parsed_data['variables']['intermediate'].append(node_id)
            elif node_type == 'created_variable':
                parsed_data['variables']['created'].append(node_id)
            
            # Parse dataset source
            self._parse_dataset_source(dataset, node_id, parsed_data)
            
            # Parse variable sources
            self._parse_variable_sources(variable_source, node_id, parsed_data)
            
        except Exception as e:
            raise Exception(f"Error parsing row data: {str(e)}")
    
    def _determine_node_type(self, var_type):
        """Determine node type from variable type string."""
        
        var_type_clean = str(var_type).strip().lower()
        
        # Direct matching for exact lineage types with suffixes for color mapping
        if var_type_clean == 'input':
            return 'input_variable'  
        elif var_type_clean == 'intermediate':
            return 'intermediate_variable'  
        elif var_type_clean == 'created':
            return 'created_variable'  
        else:
            # Debug output to see unrecognized types
            print(f"🔍 DEBUG: Unrecognized variable type: '{var_type}' -> defaulting to 'unknown'")
            return 'unknown'
    
    def _parse_dataset_source(self, dataset_source, node_id, parsed_data):
        """Parse dataset source with correct edge direction based on variable type."""
        
        try:
            if dataset_source and dataset_source not in ['N/A', 'nan', 'null', '']:
                # Clean dataset name
                clean_dataset = self._clean_dataset_name(dataset_source)
                
                if clean_dataset:
                    # Get the variable type from the already created node
                    variable_type = parsed_data['nodes'][node_id]['type']
                    
                    # Determine dataset type and edge direction based on variable type
                    if variable_type == 'input_variable':
                        # Input variables come FROM input datasets
                        # Direction: dataset → variable (dataset "contains" variable)
                        dataset_node_id = f"DS_INPUT_{clean_dataset}"
                        
                        # Create input dataset node if not exists
                        if dataset_node_id not in parsed_data['nodes']:
                            parsed_data['nodes'][dataset_node_id] = {
                                'id': dataset_node_id,
                                'label': clean_dataset,
                                'type': 'input_dataset',
                                'level': self.hierarchy_levels['input_datasets']
                            }
                            parsed_data['datasets']['input'].add(clean_dataset)
                        
                        # Create edge: dataset → variable
                        parsed_data['edges'].append({
                            'from': dataset_node_id,  # FROM dataset
                            'to': node_id,            # TO variable
                            'type': 'contains',
                            'label': 'contains'
                        })
                        
                    elif variable_type in ['intermediate_variable', 'created_variable']:
                        # Intermediate/Created variables go TO output datasets
                        # Direction: variable → dataset (variable "feeds into" dataset)
                        dataset_node_id = f"DS_OUTPUT_{clean_dataset}"
                        
                        # Create output dataset node if not exists
                        if dataset_node_id not in parsed_data['nodes']:
                            parsed_data['nodes'][dataset_node_id] = {
                                'id': dataset_node_id,
                                'label': clean_dataset,
                                'type': 'output_dataset',
                                'level': self.hierarchy_levels['output_datasets']
                            }
                            parsed_data['datasets']['output'].add(clean_dataset)
                        
                        # Create edge: variable → dataset
                        parsed_data['edges'].append({
                            'from': node_id,          # FROM variable
                            'to': dataset_node_id,    # TO dataset
                            'type': 'feeds_into',
                            'label': 'feeds into'
                        })
                        
                    else:
                        # For other variable types (like output_variable), treat as output
                        dataset_node_id = f"DS_OUTPUT_{clean_dataset}"
                        
                        if dataset_node_id not in parsed_data['nodes']:
                            parsed_data['nodes'][dataset_node_id] = {
                                'id': dataset_node_id,
                                'label': clean_dataset,
                                'type': 'output_dataset',
                                'level': self.hierarchy_levels['output_datasets']
                            }
                            parsed_data['datasets']['output'].add(clean_dataset)
                        
                        # Create edge: variable → dataset
                        parsed_data['edges'].append({
                            'from': node_id,          # FROM variable
                            'to': dataset_node_id,    # TO dataset
                            'type': 'outputs_to',
                            'label': 'outputs to'
                        })
        except Exception as e:
            # Don't fail, just log warning
            parsed_data['parsing_stats']['warnings'].append(f"Dataset source parsing error: {str(e)}")

    
    def _clean_dataset_name(self, dataset):
        """Clean dataset name from various AI-generated formats."""
        
        try:
            # Remove common prefixes and suffixes
            clean_name = dataset
            
            # Remove file extensions
            for ext in ['.csv', '.xlsx', '.sas7bdat', '.parquet', '.txt']:
                clean_name = clean_name.replace(ext, '')
            
            # Remove common prefixes
            prefixes_to_remove = ['dataset:', 'file:', 'table:', 'data:', 'source:']
            for prefix in prefixes_to_remove:
                if clean_name.lower().startswith(prefix):
                    clean_name = clean_name[len(prefix):].strip()
            
            # Remove quotes and brackets
            clean_name = clean_name.strip('"\'[](){}')
            
            # Remove paths (keep only filename)
            if '/' in clean_name:
                clean_name = clean_name.split('/')[-1]
            if '\\' in clean_name:
                clean_name = clean_name.split('\\')[-1]
            
            return clean_name.strip() if clean_name.strip() else None
            
        except Exception:
            return None
    
    def _parse_variable_sources(self, variable_source, node_id, parsed_data):
        """Parse variable sources with robust handling of AI-generated formats."""
        
        try:
            if variable_source and variable_source not in ['N/A', 'nan', 'null', '']:
                # Parse multiple source variables
                source_vars = self._extract_source_variables(variable_source)
                
                for source_var in source_vars:
                    if source_var:
                        # Create edge from source variable to current variable
                        source_node_id = f"VAR_{source_var}"
                        
                        parsed_data['edges'].append({
                            'from': source_node_id,
                            'to': node_id,
                            'type': 'transforms',
                            'label': 'transforms'
                        })
        except Exception as e:
            # Don't fail, just log warning
            parsed_data['parsing_stats']['warnings'].append(f"Variable source parsing error: {str(e)}")
    
    def _extract_source_variables(self, variable_source):
        """Extract source variables from various AI-generated formats."""
        
        try:
            # Handle different separator formats
            separators = [',', ';', '|', ' and ', ' & ', '\n']
            
            # Split by various separators
            source_vars = [variable_source]
            for sep in separators:
                temp_vars = []
                for var in source_vars:
                    temp_vars.extend(v.strip() for v in var.split(sep))
                source_vars = temp_vars
            
            # Clean each variable name
            cleaned_vars = []
            for var in source_vars:
                cleaned_var = self._clean_variable_name(var)
                if cleaned_var:
                    cleaned_vars.append(cleaned_var)
            
            return cleaned_vars
            
        except Exception:
            return []
    
    def _clean_variable_name(self, var_name):
        """Clean variable name from AI-generated formats."""
        
        try:
            # Remove common prefixes
            prefixes_to_remove = [
                'calculated:', 'derived from:', 'based on:', 'using:', 'from:',
                'variable:', 'var:', 'field:', 'column:', 'col:'
            ]
            
            clean_name = var_name.strip()
            
            for prefix in prefixes_to_remove:
                if clean_name.lower().startswith(prefix):
                    clean_name = clean_name[len(prefix):].strip()
            
            # Remove quotes, brackets, and parentheses content
            clean_name = clean_name.strip('"\'[]')
            
            # Remove content in parentheses 
            if '(' in clean_name and ')' in clean_name:
                clean_name = clean_name.split('(')[0].strip()
            
            # Remove non-alphanumeric characters except underscore
            import re
            clean_name = re.sub(r'[^a-zA-Z0-9_]', '', clean_name)
            
            return clean_name if clean_name and len(clean_name) > 1 else None
            
        except Exception:
            return None
    
    def _parse_datasets(self, lineage_data, parsed_data):
        """Parse input/output datasets from lineage_data."""
        
        try:
            # Input datasets
            input_datasets = lineage_data.get('input_datasets', [])
            for dataset in input_datasets:
                if dataset:
                    clean_dataset = self._clean_dataset_name(str(dataset))
                    if clean_dataset:
                        parsed_data['datasets']['input'].add(clean_dataset)
            
            # Output datasets
            output_datasets = lineage_data.get('output_datasets', [])
            for dataset in output_datasets:
                if dataset:
                    clean_dataset = self._clean_dataset_name(str(dataset))
                    if clean_dataset:
                        parsed_data['datasets']['output'].add(clean_dataset)
        
        except Exception as e:
            parsed_data['parsing_stats']['warnings'].append(f"Dataset parsing error: {str(e)}")
    
    def _create_smart_connections(self, parsed_data):
        """FIXED: Create smart connections with corrected intermediate variable arrow direction."""
        
        try:
            # Create output dataset connections
            for dataset in parsed_data['datasets']['output']:
                dataset_node_id = f"DS_OUTPUT_{dataset}"
                
                # Create output dataset node if not exists
                if dataset_node_id not in parsed_data['nodes']:
                    parsed_data['nodes'][dataset_node_id] = {
                        'id': dataset_node_id,
                        'label': dataset,
                        'type': 'output_dataset',
                        'level': self.hierarchy_levels['output_datasets']
                    }
                
                # Connect created/output variables to output datasets
                for var_id in parsed_data['variables']['created'] + parsed_data['variables']['output']:
                    if var_id in parsed_data['nodes']:
                        parsed_data['edges'].append({
                            'from': var_id,
                            'to': dataset_node_id,
                            'type': 'outputs_to',
                            'label': 'outputs to'
                        })
            
            # Ensure intermediate variables have OUTWARD arrows
            for intermediate_var_id in parsed_data['variables']['intermediate']:
                if intermediate_var_id in parsed_data['nodes']:
                    # Check if intermediate variable has any outgoing connections
                    has_outgoing = any(edge['from'] == intermediate_var_id for edge in parsed_data['edges'])
                    
                    if not has_outgoing:
                        # Connect intermediate variables to created variables (outward direction)
                        target_found = False
                        
                        # First try to connect to created variables
                        for created_var_id in parsed_data['variables']['created']:
                            if created_var_id in parsed_data['nodes']:
                                parsed_data['edges'].append({
                                    'from': intermediate_var_id,  
                                    'to': created_var_id,        
                                    'type': 'feeds_into',
                                    'label': 'feeds into'
                                })
                                target_found = True
                                break  
                        
                        # If no created variables, try connecting to output datasets
                        for dataset in parsed_data['datasets']['output']:
                            dataset_node_id = f"DS_OUTPUT_{dataset}"
                            if dataset_node_id in parsed_data['nodes']:
                                parsed_data['edges'].append({
                                    'from': intermediate_var_id,  
                                    'to': dataset_node_id,        
                                    'type': 'flows_to',
                                    'label': 'flows to'
                                })
                                break  
        
        except Exception as e:
            parsed_data['parsing_stats']['warnings'].append(f"Smart connections error: {str(e)}")
    
    def create_visualization_with_modes(self, lineage_data, mode='overview'):
        """Create visualization with multiple view modes."""
        
        try:
            # Parse data robustly
            parsed_data = self._robust_parse_lineage_data(lineage_data)
            
            # Create visualization based on mode
            if mode == 'overview':
                return self._create_overview_visualization(parsed_data)
            elif mode == 'detailed':
                return self._create_detailed_visualization(parsed_data)
            elif mode == 'focused':
                return self._create_focused_visualization(parsed_data)
            else:
                return self._create_overview_visualization(parsed_data)
        
        except Exception as e:
            return self._create_error_visualization(f"Visualization error: {str(e)}")
    
    def _create_overview_visualization(self, parsed_data):
        """Create overview visualization with font sizing and colors."""
        
        try:
            from pyvis.network import Network
            
            # Create network
            net = Network(
                height="600px",
                width="100%",
                bgcolor="#f8f9fa",
                font_color="#000000",
                directed=True
            )
            
            # Add nodes with styling
            x_positions = {}  # Track X positions for each level
            
            for node_id, node_data in parsed_data['nodes'].items():
                try:
                    node_type = node_data['type']
                    level = node_data['level']
                    
                    # Calculate position
                    if level not in x_positions:
                        x_positions[level] = 0
                    x_positions[level] += 150
                    
                    # Get node configuration
                    config = self.node_config.get(node_type, self.node_config['unknown'])
                    color = self.node_colors.get(node_type, self.node_colors['unknown'])
                    icon = self.node_icons.get(node_type, self.node_icons['unknown'])
                    
                    # Simple text label (no HTML)
                    label = f"{icon} {node_data['label']}"
                    
                    # Font sizing via pyvis parameters - Dataset names LARGER than variable names
                    if 'dataset' in node_type:
                        font_size = 30  
                        font_bold = True
                    else:
                        font_size = 14 
                        font_bold = True
                    
                    # Create hover text
                    hover_text = f"{label}\nType: {node_type}\nLevel: {level}"
                    
                    # Add node with styling
                    net.add_node(
                        node_id,
                        label=label,
                        color=color,
                        size=config['size'],
                        shape=config['shape'],
                        title=hover_text,
                        font={'size': font_size, 'bold': font_bold},
                        x=x_positions[level],
                        y=level * 150,
                        physics=False  # positions for clear hierarchy
                    )
                
                except Exception as e:
                    continue  # Skip problematic nodes
            
            # Add edges with smaller font
            edge_count = 0
            for edge in parsed_data['edges']:
                try:
                    if edge['from'] in parsed_data['nodes'] and edge['to'] in parsed_data['nodes']:
                        net.add_edge(
                            edge['from'],
                            edge['to'],
                            color="#666666",
                            width=2,
                            arrows="to",
                            title=edge.get('label', ''),
                            label=edge.get('label', ''),
                            font={'size': 8}  # Even smaller font for edge labels 
                        )
                        edge_count += 1
                except Exception:
                    continue  # Skip problematic edges
            
            # Configure network
            net.set_options("""
            var options = {
                "physics": {
                    "enabled": false
                },
                "interaction": {
                    "dragNodes": true,
                    "dragView": true,
                    "zoomView": true
                },
                "layout": {
                    "hierarchical": {
                        "enabled": true,
                        "direction": "UD",
                        "sortMethod": "directed",
                        "levelSeparation": 150,
                        "nodeSpacing": 200
                    }
                }
            }
            """)
            
            # Generate HTML
            html_content = net.generate_html()
            
            # Add custom header
            stats_text = f"{len(parsed_data['nodes'])} Nodes | {edge_count} Connections | Overview Mode"
            enhanced_html = html_content.replace(
                '<body>',
                f'''<body>
                <div style="text-align: center; padding: 10px; background: #f0f0f0; margin-bottom: 10px;">
                    <h3 style="margin: 0; color: #333;">🗺️ Data Lineage Overview</h3>
                    <p style="margin: 5px 0 0 0; color: #666;">{stats_text}</p>
                </div>'''
            )
            return enhanced_html
        except Exception as e:
            return self._create_error_visualization(f"Overview visualization error: {str(e)}")
    
    def _create_detailed_visualization(self, parsed_data):
        """Create detailed visualization with font sizing and colors."""
        try:
            from pyvis.network import Network
            
            # Create network with more space
            net = Network(
                height="800px",
                width="100%",
                bgcolor="#f8f9fa",
                font_color="#000000",
                directed=True
            )
            
            # Add all nodes with styling
            for node_id, node_data in parsed_data['nodes'].items():
                try:
                    node_type = node_data['type']
                    config = self.node_config.get(node_type, self.node_config['unknown'])
                    color = self.node_colors.get(node_type, self.node_colors['unknown'])
                    icon = self.node_icons.get(node_type, self.node_icons['unknown'])
                    
                    # Simple text label (no HTML)
                    label = f"{icon} {node_data['label']}"
                    
                    # Font sizing via pyvis parameters - Dataset names LARGER than variable names
                    if 'dataset' in node_type:
                        font_size = 20  # LARGER for datasets 
                        font_bold = True
                    else:
                        font_size = 14  # Medium for variables 
                        font_bold = True
                    
                    # Detailed hover text
                    hover_text = f"{label}\nType: {node_type}\nDataset: {node_data.get('dataset', 'N/A')}\nSources: {node_data.get('variable_source', 'N/A')}\nCalculation: {node_data.get('calculation', 'N/A')[:50]}..."
                    
                    net.add_node(
                        node_id,
                        label=label,
                        color=color,
                        size=config['size'],
                        shape=config['shape'],
                        title=hover_text,
                        font={'size': font_size, 'bold': font_bold}
                    )
                
                except Exception:
                    continue
            
            # Add all edges with FIXED smaller font
            edge_count = 0
            for edge in parsed_data['edges']:
                try:
                    if edge['from'] in parsed_data['nodes'] and edge['to'] in parsed_data['nodes']:
                        net.add_edge(
                            edge['from'],
                            edge['to'],
                            color="#666666",
                            width=2,
                            arrows="to",
                            title=edge.get('label', ''),
                            label=edge.get('label', ''),
                            font={'size': 8}  
                        )
                        edge_count += 1
                except Exception:
                    continue
            
            node_count = len(parsed_data['nodes'])
            edge_count = len(parsed_data['edges'])
            
            # Configure for detailed view for larger lineage - Disable physics
            if node_count > 100 or edge_count > 500:
                net.set_options("""
                var options = {
                "physics": {
                    "enabled": false},
                    "layout": {
                        "hierarchical": {
                        "enabled": true,
                        "direction": "UD",
                        "sortMethod": "directed"
                    }
                },
                "interaction": {
                    "dragNodes": true,
                    "dragView": true,
                    "zoomView": true
                }
            }
            """)
            else:
                # Configure for detailed view for smaller lineage
                net.set_options("""
                var options = {
                    "physics": {
                        "enabled": true,
                        "barnesHut": {
                            "gravitationalConstant": -8000,
                            "springLength": 200,
                            "springConstant": 0.04,
                            "damping": 0.09
                        }
                    },
                    "interaction": {
                        "dragNodes": true,
                        "dragView": true,
                        "zoomView": true
                    }
                }
                """)
            
            html_content = net.generate_html()
            
            # Add detailed header
            stats_text = f"{len(parsed_data['nodes'])} Nodes | {edge_count} Connections | Detailed Mode"
            enhanced_html = html_content.replace(
                '<body>',
                f'''<body>
                <div style="text-align: center; padding: 10px; background: #f0f0f0; margin-bottom: 10px;">
                    <h3 style="margin: 0; color: #333;">🔍 Detailed Data Lineage</h3>
                    <p style="margin: 5px 0 0 0; color: #666;">{stats_text}</p>
                </div>'''
            )
            
            return enhanced_html
            
        except Exception as e:
            return self._create_error_visualization(f"Detailed visualization error: {str(e)}")
    
    def _create_focused_visualization(self, parsed_data):
        """Create focused visualization with FIXED font sizing and colors."""
        
        try:
            from pyvis.network import Network
            
            # Create network
            net = Network(
                height="700px",
                width="100%",
                bgcolor="#f8f9fa",
                font_color="#000000",
                directed=True
            )
            
            # Focus on created variables and their immediate connections
            focus_nodes = set()
            focus_edges = []
            
            # Add created variables
            for var_id in parsed_data['variables']['created']:
                if var_id in parsed_data['nodes']:
                    focus_nodes.add(var_id)
            
            # Add their source variables
            for edge in parsed_data['edges']:
                if edge['to'] in focus_nodes:
                    focus_nodes.add(edge['from'])
                    focus_edges.append(edge)
            
            # Add nodes with styling
            for node_id in focus_nodes:
                if node_id in parsed_data['nodes']:
                    try:
                        node_data = parsed_data['nodes'][node_id]
                        node_type = node_data['type']
                        config = self.node_config.get(node_type, self.node_config['unknown'])
                        color = self.node_colors.get(node_type, self.node_colors['unknown'])
                        icon = self.node_icons.get(node_type, self.node_icons['unknown'])
                        
                        # Simple text label (no HTML)
                        label = f"{icon} {node_data['label']}"
                        
                        # Font sizing via pyvis parameters
                        if 'dataset' in node_type:
                            font_size = 30  # Larger for datasets
                            font_bold = True
                        else:
                            font_size = 14  # Medium for variables
                            font_bold = True
                        
                        # Focused hover text
                        hover_text = f"{label}\nType: {node_type}\nCalculation: {node_data.get('calculation', 'N/A')[:100]}..."
                        
                        net.add_node(
                            node_id,
                            label=label,
                            color=color,
                            size=config['size'] + 10,  # Slightly larger for focus
                            shape=config['shape'],
                            title=hover_text,
                            font={'size': font_size, 'bold': font_bold}
                        )
                    
                    except Exception:
                        continue
            
            # Add focused edges with FIXED smaller font
            for edge in focus_edges:
                try:
                    if edge['from'] in focus_nodes and edge['to'] in focus_nodes:
                        net.add_edge(
                            edge['from'],
                            edge['to'],
                            color="#2196F3",
                            width=3,
                            arrows="to",
                            title=edge.get('label', ''),
                            label=edge.get('label', ''),
                            font={'size': 10}  # Smaller font for edge labels
                        )
                except Exception:
                    continue
            
            # Configure for focused view
            net.set_options("""
            var options = {
                "physics": {
                    "enabled": true,
                    "barnesHut": {
                        "gravitationalConstant": -4000,
                        "springLength": 150,
                        "springConstant": 0.08,
                        "damping": 0.1
                    }
                },
                "interaction": {
                    "dragNodes": true,
                    "dragView": true,
                    "zoomView": true
                }
            }
            """)
            
            html_content = net.generate_html()
            
            # Add focused header
            stats_text = f"{len(focus_nodes)} Focused Nodes | {len(focus_edges)} Connections | Created Variables Focus"
            enhanced_html = html_content.replace(
                '<body>',
                f'''<body>
                <div style="text-align: center; padding: 10px; background: #f0f0f0; margin-bottom: 10px;">
                    <h3 style="margin: 0; color: #333;">🎯 Focused Data Lineage</h3>
                    <p style="margin: 5px 0 0 0; color: #666;">{stats_text}</p>
                </div>'''
            )
            
            return enhanced_html
            
        except Exception as e:
            return self._create_error_visualization(f"Focused visualization error: {str(e)}")
    
    def _create_error_visualization(self, error_message):
        """Create simple error visualization."""
        
        error_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Data Lineage Visualization Error</title>
        </head>
        <body>
            <div style="text-align: center; padding: 50px; font-family: Arial, sans-serif;">
                <h2 style="color: #f44336;">⚠️ Visualization Error</h2>
                <p style="color: #666; margin: 20px 0;">
                    {error_message}
                </p>
                <p style="color: #999; font-size: 14px;">
                    Please check your data format and try again.
                </p>
            </div>
        </body>
        </html>
        """
        
        return error_html
    
    def display_visualization_stats(self, lineage_data):
        """Display visualization statistics with robust error handling."""
        
        try:
            # Parse data to get stats
            parsed_data = self._robust_parse_lineage_data(lineage_data)
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_nodes = len(parsed_data['nodes'])
                st.metric("📊 Total Nodes", total_nodes)
            
            with col2:
                total_edges = len(parsed_data['edges'])
                st.metric("🔗 Connections", total_edges)
            
            with col3:
                input_count = len(parsed_data['variables']['input'])
                st.metric("📥 Input Variables", input_count)
            
            with col4:
                created_count = len(parsed_data['variables']['created'])
                st.metric("⚙️ Created Variables", created_count)
            
            # Show parsing stats
            stats = parsed_data['parsing_stats']
            if stats['errors'] > 0 or stats['warnings']:
                with st.expander("⚠️ Parsing Statistics", expanded=False):
                    st.write(f"**Successful:** {stats['success']}")
                    st.write(f"**Errors:** {stats['errors']}")
                    if stats['warnings']:
                        st.write("**Warnings:**")
                        for warning in stats['warnings'][:5]:  # Show first 5 warnings
                            st.write(f"• {warning}")
            
        except Exception as e:
            st.error(f"Error displaying stats: {str(e)}")
    
    def create_enhanced_table_display(self, lineage_data):
        """Create enhanced table display with error handling."""
        
        try:
            lineage_table = lineage_data.get('lineage_table', [])
            
            if not lineage_table:
                return pd.DataFrame()
            
            # Convert to DataFrame safely
            df = pd.DataFrame(lineage_table)
            
            # Ensure required columns exist
            required_columns = ['variable_name', 'variable_type', 'dataset', 'variable_source']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = 'N/A'
            
            return df[required_columns].reset_index(drop=True)
            
        except Exception as e:
            st.error(f"Error creating table display: {str(e)}")
            return pd.DataFrame()
    
    # Keep the main interface method
    def create_interactive_lineage_visualization(self, lineage_data):
        """Main interface method - creates overview by default."""
        return self.create_visualization_with_modes(lineage_data, 'overview')



components/functions.py:
import streamlit as st

def reset_lineage_state():
    """Reset lineage analysis session state."""
    keys_to_reset = [
        'lineage_variables_discovered', 
        'lineage_all_variables', 
        'lineage_analysis_complete',
        'lineage_detailed_results'
    ]
    for key in keys_to_reset:
        if key in st.session_state:
            del st.session_state[key]

def store_variables_in_session(all_variables):
    """Store discovered variables in session state."""
    st.session_state['lineage_variables_discovered'] = True
    st.session_state['lineage_all_variables'] = all_variables

please check now what can be done to fix it.
