import streamlit as st
import pandas as pd
from typing import List
from io import BytesIO
import openpyxl
import json
from models.Code_Reviewer import *
from models.RDE_Tester import *
from models.Data_Lineage import *
from components.AVA_Assistant import *
from components.functions import *
import warnings
import logging

warnings.filterwarnings("ignore", message=".*missing ScriptRunContext*")

logging.getLogger("streamlit.runtime.scriptrunner.script_runner").setLevel(logging.ERROR)

def perform_data_lineage_analysis(code_files: list, data_files: list, selected_vars: List[str] = None):
    """Enhanced AI-powered data lineage analysis."""
    
    if not code_files:
        st.error("❌ Please upload code files to analyze.")
        return
    
    try:
        # Initialize session state if not exists
        if 'lineage_variables_discovered' not in st.session_state:
            st.session_state['lineage_variables_discovered'] = False
        if 'lineage_all_variables' not in st.session_state:
            st.session_state['lineage_all_variables'] = {}
        
        # STEP 1: AI Variable Discovery
        if not st.session_state['lineage_variables_discovered']:
            with st.spinner("🔍 AI analyzing your scripts to discover all variables..."):
                lineage_analyzer = DataLineageAnalyzer()
                all_variables = lineage_analyzer._simple_ai_variable_extraction(code_files)
                
                if not all_variables.get('all_variables'):
                    st.error("❌ No variables found in the uploaded scripts.")
                    with st.expander("🔧 Debug Information", expanded=True):
                        st.write("**Files uploaded:**")
                        for file in code_files:
                            st.write(f"- {file.name}")
                        
                        st.write("**First 500 characters of first file:**")
                        try:
                            content = code_files[0].read().decode('utf-8')
                            st.code(content[:500])
                            code_files[0].seek(0)
                        except Exception as e:
                            st.write(f"Could not read file: {str(e)}")
                    return
                
                # Store in session state
                store_variables_in_session(all_variables)
        
        # Get variables from session state
        all_variables = st.session_state['lineage_all_variables']
        
        # Display discovered variables summary (always show if variables are discovered)
        if st.session_state['lineage_variables_discovered']:
            st.success(f"✅ AI discovered {len(all_variables['all_variables'])} variables in your scripts!")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📥 Input Variables", len(all_variables.get('input_variables', [])))
            with col2:
                st.metric("🔄 Intermediate Variables", len(all_variables.get('intermediate_variables', [])))
            with col3:
                st.metric("⚙️ Created Variables", len(all_variables.get('created_variables', [])))
                
            # with col4:
            #     st.metric("📤 Output Variables", len(all_variables.get('output_variables', [])))
            
            # Show discovered variables by category
            with st.expander("📋 View All Discovered Variables", expanded=False):
                tab1, tab2, tab3 = st.tabs(["📥 Input", "🔄 Intermediate","⚙️ Created"])
                
                with tab1:
                    if all_variables.get('input_variables'):
                        for var in all_variables['input_variables']:
                            st.write(f"• {var}")
                    else:
                        st.info("No input variables detected")
                
                with tab2:
                    if all_variables.get('intermediate_variables'):
                        for var in all_variables['intermediate_variables']:
                            st.write(f"• {var}")
                    else:
                        st.info("No intermediate variables detected")
                
                with tab3:
                    if all_variables.get('created_variables'):
                        for var in all_variables['created_variables']:
                            st.write(f"• {var}")
                    else:
                        st.info("No created variables detected")                
            
            
            if st.session_state.get('lineage_variables_discovered', False):
                
                # ========== SEPARATE MANUAL UPLOAD SECTION ==========
                st.markdown("---")  # Separator
                st.markdown("### 📁 Upload Existing Lineage Excel for Visualization")
                
                # Separate file uploader
                uploaded_excel_file = st.file_uploader(
                    "Upload your existing data lineage Excel file",
                    type=['xlsx'],
                    key="manual_excel_uploader",  # Unique key
                    help="Upload Excel file with 'AI_Data_Lineage' sheet containing Variable Name, Variable Type, Dataset Source, Variable Source columns"
                )
                
                if uploaded_excel_file:
                    
                    # Parse uploaded file button
                    if st.button("📊 Parse Excel & Create Visualization", type="secondary", key="parse_excel_btn"):
                        
                        with st.spinner("📊 Parsing uploaded Excel file..."):
                            try:
                                # Use the same visualizer class but completely separate
                                manual_visualizer = DataLineageVisualizer()
                                manual_lineage_data = manual_visualizer.parse_uploaded_lineage_excel(uploaded_excel_file)
                                
                                if manual_lineage_data and manual_lineage_data.get('lineage_table'):
                                    # Store in separate session state
                                    st.session_state['manual_lineage_data'] = manual_lineage_data
                                    st.session_state['manual_lineage_ready'] = True
                                    st.success(f"✅ Excel parsed successfully! Found {len(manual_lineage_data['lineage_table'])} variables")
                                else:
                                    st.error("❌ Failed to parse Excel file. Please check the format.")
                                    
                            except Exception as e:
                                st.error(f"❌ Error parsing Excel: {str(e)}")
                                
                    # Show manual visualization section if data is ready
                    if st.session_state.get('manual_lineage_ready', False):
                        
                        manual_data = st.session_state.get('manual_lineage_data', {})
                        
                        st.markdown("#### 🎨 Manual Excel Visualization")
                        st.success(f"📊 **Source**: {uploaded_excel_file.name}")
                        
                        # Show data summary
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("📊 Variables", len(manual_data.get('lineage_table', [])))
                        with col2:
                            st.metric("📥 Input Datasets", len(manual_data.get('input_datasets', [])))
                        with col3:
                            st.metric("📤 Output Datasets", len(manual_data.get('output_datasets', [])))
                        
                        # Visualization controls 
                        col1, col2, col3 = st.columns([2, 1, 1])

                        with col1:
                            # View mode selection with unique key
                            manual_view_mode = st.selectbox(
                                "🎯 Select Visualization Mode:",
                                ["overview", "detailed", "focused"],
                                index=0,
                                key="manual_view_mode_selector",  # Unique key
                                help="Overview: Hierarchical flow | Detailed: All connections | Focused: Created variables only"
                            )

                        with col2:
                            # Create visualization button with unique key
                            create_manual_viz_btn = st.button(
                                "🎨 Create Visualization", 
                                type="primary", 
                                use_container_width=True, 
                                key="create_manual_viz_btn"  # Unique key
                            )

                        with col3:
                            # Download button with unique key
                            manual_viz_key = f"manual_viz_{manual_view_mode}"
                            if st.session_state.get('manual_viz_dict', {}).get(manual_viz_key):
                                manual_viz_html = st.session_state['manual_viz_dict'][manual_viz_key]
                                timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                                st.download_button(
                                    label="📥 Download HTML",
                                    data=manual_viz_html,
                                    file_name=f"manual_lineage_{manual_view_mode}_{timestamp}.html",
                                    mime="text/html",
                                    key=f"download_manual_viz_{manual_view_mode}_btn"  # Unique key
                                )

                        # Create visualization when button is clicked
                        if create_manual_viz_btn:
                            with st.spinner(f"🎨 Creating {manual_view_mode} visualization from uploaded Excel..."):
                                try:
                                    # Use the same visualizer class but completely independent
                                    manual_visualizer = DataLineageVisualizer()
                                    
                                    # Show statistics
                                    st.markdown("**📊 Manual Excel Visualization Statistics:**")
                                    manual_visualizer.display_visualization_stats(manual_data)
                                    
                                    # Create visualization
                                    manual_viz_html = manual_visualizer.create_visualization_with_modes(manual_data, manual_view_mode)
                                    
                                    if manual_viz_html:
                                        # Store in separate session state
                                        if 'manual_viz_dict' not in st.session_state:
                                            st.session_state['manual_viz_dict'] = {}
                                        
                                        manual_viz_key = f"manual_viz_{manual_view_mode}"
                                        st.session_state['manual_viz_dict'][manual_viz_key] = manual_viz_html
                                        
                                        st.success(f"✅ {manual_view_mode.title()} visualization created from your Excel file!")
                                    else:
                                        st.error("❌ Failed to create visualization")
                                        
                                except Exception as e:
                                    st.error(f"❌ Visualization error: {str(e)}")

                        # Display the visualization if it exists
                        manual_viz_key = f"manual_viz_{manual_view_mode}"
                        if st.session_state.get('manual_viz_dict', {}).get(manual_viz_key):
                            
                            manual_viz_html = st.session_state['manual_viz_dict'][manual_viz_key]
                            
                            # Display the interactive visualization
                            st.components.v1.html(manual_viz_html, height=750, scrolling=True)
                            
                            # Show features
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.markdown("**🎨 Colors**")
                                st.write("• Green: Input variables")
                                st.write("• Orange: Intermediate variables")
                                st.write("• Blue: Created variables")
                            
                        else:
                            # Show placeholder
                            st.info(f"Click '🎨 Create Visualization' above to generate the {manual_view_mode} visualization from your Excel data")

                # Help section for manual upload
                with st.expander("📋 Expected Excel Format for Manual Upload", expanded=False):
                    st.markdown("""
                    **Required Sheet Name**: `AI_Data_Lineage`
                    
                    **Required Columns**:
                    - **Variable Name**: Name of the variable/column
                    - **Variable Type**: Must be exactly: `Input`, `Intermediate`, `Created`, or `Output`
                    - **Dataset Source**: Source dataset name (for input variables)
                    - **Variable Source**: Source variables used to create this variable
                    
                    **Example Excel Data**:
                    ```
                    Variable Name | Variable Type | Dataset Source | Variable Source
                    customer_id   | Input        | customers.csv  | N/A
                    revenue_2024  | Created      | N/A           | revenue, year
                    profit_calc   | Intermediate | N/A           | revenue_2024, costs
                    ```
                    
                    **📁 Use the Excel files generated by this tool's "Download Excel" feature for guaranteed compatibility!**
                    """)

            # ========== END OF SEPARATE MANUAL UPLOAD SECTION ==========
            
            # Add Full Data Lineage Analysis
            st.markdown("### 📊 Full Data Lineage")

            # Check if data files were also uploaded
            data_files_available = data_files if data_files else []

            col1, col2 = st.columns([2, 1])
            with col1:
                if st.button("🗺️ Generate Full Data Lineage", type="primary", use_container_width=True, key="full_lineage_btn"):
                    
                    # First test AI connection
                    lineage_analyzer = DataLineageAnalyzer()
                    if lineage_analyzer.debug_ai_connection():
                        
                        with st.spinner("🤖 AI analyzing code for comprehensive data lineage..."):
                            st.write(f"📁 Analyzing {len(code_files)} code files...")
                            
                            # Show what files we're analyzing
                            for file in code_files:
                                content = file.read().decode('utf-8')
                                st.write(f"• {file.name}: {len(content.split())} lines, {len(content)} characters")
                                file.seek(0)
                            
                            full_lineage_data = lineage_analyzer.generate_full_data_lineage(all_variables, code_files, data_files_available)
                        
                        if "error" not in full_lineage_data:
                            # Store in session state
                            st.session_state['full_lineage_data'] = full_lineage_data
                            st.session_state['full_lineage_complete'] = True
                            
                            st.success(f"✅ Full data lineage generated!")
                            st.info(f"📊 Found {full_lineage_data['total_variables']} variables")
                        else:
                            st.error("❌ Failed to generate full data lineage")
                    else:
                        st.error("❌ Cannot proceed - AI connection failed")

            with col2:
                if st.session_state.get('full_lineage_complete', False):
                    full_lineage_data = st.session_state.get('full_lineage_data', {})
                    if full_lineage_data and full_lineage_data.get('lineage_table'):
                        lineage_analyzer = DataLineageAnalyzer()
                        excel_buffer = lineage_analyzer.create_full_lineage_excel(full_lineage_data)
                        
                        excel_buffer.seek(0)
                        
                        st.download_button(
                            label="📥 Download Excel",
                            data=excel_buffer,
                            file_name=f"ai_data_lineage.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_full_lineage"
                        )
                        st.info(f"📊 In case of warnings - Try downloading the Lineage Excel again")

            # Display Full Lineage Results if available
            if st.session_state.get('full_lineage_complete', False):
                full_lineage_data = st.session_state.get('full_lineage_data', {})
                
                if full_lineage_data and full_lineage_data.get('lineage_table'):
                    st.markdown("### 📋 Full Data Lineage Results")
                    
                    # Show datasets summary
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("📥 Input Datasets", len(full_lineage_data.get('input_datasets', [])))
                    with col2:
                        st.metric("📤 Output Datasets", len(full_lineage_data.get('output_datasets', [])))
                    with col3:
                        st.metric("🔢 Total Variables", full_lineage_data.get('total_variables', 0))
                    
                    # Show datasets
                    if full_lineage_data.get('input_datasets') or full_lineage_data.get('output_datasets'):
                        with st.expander("📁 Datasets Summary", expanded=False):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.markdown("**📥 Input Datasets:**")
                                for dataset in full_lineage_data.get('input_datasets', []):
                                    st.write(f"• {dataset}")
                            
                            with col2:
                                st.markdown("**📤 Output Datasets:**")
                                for dataset in full_lineage_data.get('output_datasets', []):
                                    st.write(f"• {dataset}")
                    
                    # Show lineage table
                    lineage_df = pd.DataFrame(full_lineage_data['lineage_table'])
                    
                    if not lineage_df.empty:
                        # Rename columns for display
                        display_df = lineage_df.copy()
                        display_df.columns = [
                            'Sr. No', 'Variable Name', 'Variable Type', 'Dataset', 
                            'Variable Source', 'Calculation Methodology', 'Explanation', 'Line No'
                        ]
                        
                        # Add filtering options
                        st.markdown("#### 🔍 Data Lineage Table")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            var_type_filter = st.selectbox(
                                "Filter by Variable Type:",
                                ["All"] + list(display_df['Variable Type'].unique()),
                                key="var_type_filter"
                            )
                        
                        with col2:
                            search_filter = st.text_input(
                                "Search Variables:",
                                placeholder="Type to search...",
                                key="var_search_filter"
                            )
                        
                        # Apply filters
                        filtered_df = display_df.copy()
                        
                        if var_type_filter != "All":
                            filtered_df = filtered_df[filtered_df['Variable Type'] == var_type_filter]
                        
                        if search_filter:
                            mask = filtered_df['Variable Name'].str.contains(search_filter, case=False, na=False)
                            filtered_df = filtered_df[mask]
                        
                        # Display table with styling
                        st.dataframe(
                            filtered_df,
                            use_container_width=True,
                            column_config={
                                "Sr. No": st.column_config.NumberColumn("Sr. No", width="small"),
                                "Variable Name": st.column_config.TextColumn("Variable Name", width="medium"),
                                "Variable Type": st.column_config.TextColumn("Variable Type", width="small"),
                                "Dataset": st.column_config.TextColumn("Dataset", width="medium"),
                                "Variable Source": st.column_config.TextColumn("Variable Source", width="medium"),
                                "Calculation Methodology": st.column_config.TextColumn("Calculation Methodology", width="large"),
                                "Explanation": st.column_config.TextColumn("Explanation", width="large"),
                                "Line No": st.column_config.NumberColumn("Line No", width="small")
                            }
                        )



                # ========== ENHANCED MULTI-MODE VISUALIZATION CODE ==========
                st.markdown("---")  # Separator

                # Visualization controls
                col1, col2, col3 = st.columns([2, 1, 1])

                with col1:
                    # View mode selection
                    view_mode = st.selectbox(
                        "🎯 Select View Mode:",
                        ["overview", "detailed", "focused"],
                        index=0,
                        key="lineage_view_mode",
                        help="Overview: High-level flow | Detailed: All connections | Focused: Created variables only"
                    )
                    
                    # Display mode descriptions
                    mode_descriptions = {
                        "overview": "📊 **Overview Mode**: Shows hierarchical flow from input datasets → variables → output datasets",
                        "detailed": "🔍 **Detailed Mode**: Shows all variables and connections with full details",
                        "focused": "🎯 **Focused Mode**: Shows only created variables and their immediate sources"
                    }
                    st.markdown(mode_descriptions[view_mode])

                with col2:
                    # Create visualization button
                    create_viz_btn = st.button(
                        "🎨 Create Visualization", 
                        type="secondary", 
                        use_container_width=True, 
                        key="create_multi_mode_viz_btn"
                    )

                with col3:
                    # Download button (only show if visualization exists)
                    if st.session_state.get('lineage_viz_html_dict', {}).get(view_mode):
                        viz_html = st.session_state['lineage_viz_html_dict'][view_mode]
                        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                        st.download_button(
                            label="📥 Download HTML",
                            data=viz_html,
                            file_name=f"lineage_{view_mode}_{timestamp}.html",
                            mime="text/html",
                            key=f"download_viz_{view_mode}_btn",
                            help=f"Download {view_mode} visualization as HTML file"
                        )

                # Create visualization when button is clicked
                if create_viz_btn:
                    with st.spinner(f"🎨 Creating {view_mode} visualization..."):
                        try:
                            # Initialize the optimized visualizer
                            visualizer = DataLineageVisualizer()
                            
                            # Show statistics
                            st.markdown("**📊 Visualization Statistics:**")
                            visualizer.display_visualization_stats(full_lineage_data)
                            
                            # Create visualization with selected mode
                            viz_html = visualizer.create_visualization_with_modes(full_lineage_data, view_mode)
                            
                            if viz_html:
                                # Store in session state (separate for each mode)
                                if 'lineage_viz_html_dict' not in st.session_state:
                                    st.session_state['lineage_viz_html_dict'] = {}
                                
                                st.session_state['lineage_viz_html_dict'][view_mode] = viz_html
                                st.session_state['lineage_viz_mode'] = view_mode
                                
                                st.success(f"✅ {view_mode.title()} visualization created successfully!")
                            else:
                                st.error("❌ Failed to create visualization")
                                
                        except Exception as e:
                            st.error(f"❌ Visualization error: {str(e)}")
                            st.info("💡 The visualization will continue to work even if some data parsing fails")

                # Display the visualization if it exists for current mode
                if st.session_state.get('lineage_viz_html_dict', {}).get(view_mode):
                    st.markdown(f"#### 🌐 {view_mode.title()} Data Lineage Network")
                    
                    viz_html = st.session_state['lineage_viz_html_dict'][view_mode]
                    
                    # Display the interactive visualization
                    st.components.v1.html(viz_html, height=750, scrolling=True)
                    
                    if view_mode == "overview":
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            **📊 Overview Mode Features:**
                            - DS **Input Datasets** (Top Level - Red circles)
                            - IN **Input Variables** (Level 2 - Green boxes)
                            - INT **Intermediate Variables** (Level 3 - Orange boxes)
                            - CR **Created Variables** (Level 4 - Blue boxes)
                            - OUT **Output Datasets** (Level 5 - Purple circles)
                            """)
                        with col2:
                            st.markdown("""
                            **🖱️ Interaction Tips:**
                            - **Hover**: View variable details
                            - **Click & Drag**: Reposition nodes
                            - **Scroll**: Zoom in/out
                            - **Clear Layout**: Shows data flow hierarchy
                            """)
                    
                    elif view_mode == "detailed":
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            **🔍 Detailed Mode Features:**
                            - **All Variables**: Complete dataset shown
                            - **All Connections**: Every relationship visible
                            - **Full Details**: Hover for calculations
                            - **Physics Layout**: Nodes organize naturally
                            """)
                        with col2:
                            st.markdown("""
                            **🖱️ Advanced Interaction:**
                            - **Drag Nodes**: Customize layout
                            - **Zoom**: Focus on specific areas
                            - **Connected Highlighting**: Click nodes
                            - **Dynamic Layout**: Nodes settle optimally
                            """)
                    
                    else:  # focused mode
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            **🎯 Focused Mode Features:**
                            - **Created Variables**: Main focus (Blue triangles)
                            - **Source Variables**: What feeds into created vars
                            - **Calculation Focus**: Shows transformation details
                            """)
                        with col2:
                            st.markdown("""
                            **🖱️ Focused Interaction:**
                            - **Hover**: See calculation details
                            - **Trace Sources**: Follow blue arrows
                            - **Minimal Clutter**: Only important connections
                            - **Clear Flow**: Input → Transformation → Output
                            """)
                    
                else:
                    # Show placeholder when no visualization exists
                    st.markdown("#### 🌐 Interactive Data Lineage Network")
                    st.info(f"Click '🎨 Create Visualization' above to generate the {view_mode} mode visualization")
                    
                    # Show mode preview
                    mode_previews = {
                        "overview": "📊 **Overview Preview**: Will show hierarchical flow from datasets → input variables → created variables → output datasets",
                        "detailed": "🔍 **Detailed Preview**: Will show all variables and connections with complete relationship mapping",
                        "focused": "🎯 **Focused Preview**: Will show only created variables and their immediate source variables"
                    }
                    st.markdown(mode_previews[view_mode])

                # ========== END OF VISUALIZATION CODE ==========

            # STEP 2: User selects variables for lineage analysis
            st.markdown("### 🎯 Individual Variable Lineage")
            
            # Add a reset button
            col_reset, col_spacer = st.columns([1, 4])
            with col_reset:
                if st.button("🔄 Reset Analysis", help="Start over with new variable discovery"):
                    reset_lineage_state()
                    st.rerun()
            
            # Selection options
            selection_option = st.radio(
                "Choose your analysis scope:",
                [
                    "🎯 Select Specific Variables", 
                    "⚙️ All Created Variables", 
                    "🔄 All Intermediate Variables",
                    "📊 All Variables"
                ],
                help="Choose which variables you want detailed lineage analysis for",
                key="lineage_selection_option"
            )
            
            selected_variables = []
            
            if selection_option == "🎯 Select Specific Variables":
                # Search functionality
                search_term = st.text_input(
                    "🔍 Search Variables:",
                    placeholder="Type to filter variables...",
                    help="Search for specific variables by name",
                    key="lineage_search_term"
                )
                
                # Filter variables based on search
                available_vars = all_variables['all_variables']
                if search_term:
                    available_vars = [var for var in all_variables['all_variables'] 
                                   if search_term.lower() in var.lower()]
                    if available_vars:
                        st.info(f"📋 Found {len(available_vars)} variables matching '{search_term}'")
                    else:
                        st.warning(f"❌ No variables found matching '{search_term}'")
                
                # Multi-select for variables
                if available_vars:
                    selected_variables = st.multiselect(
                        "📋 Select Variables for Detailed Lineage:",
                        options=sorted(available_vars),
                        help="Choose specific variables to trace their creation and transformation",
                        key="lineage_selected_variables"
                    )
                    
                    if selected_variables:
                        st.success(f"✅ Selected {len(selected_variables)} variables: {', '.join(selected_variables[:3])}{'...' if len(selected_variables) > 3 else ''}")
            
            elif selection_option == "⚙️ All Created Variables":
                selected_variables = all_variables.get('created_variables', [])
                if selected_variables:
                    st.info(f"🎯 Will analyze {len(selected_variables)} created variables")
                else:
                    st.warning("No created variables found to analyze")
            
            elif selection_option == "🔄 All Intermediate Variables":
                selected_variables = all_variables.get('intermediate_variables', [])
                if selected_variables:
                    st.info(f"🎯 Will analyze {len(selected_variables)} intermediate variables")
                else:
                    st.warning("No intermediate variables found to analyze")
            
            else:  # All Variables
                selected_variables = all_variables['all_variables']
                st.info(f"🎯 Will analyze all {len(selected_variables)} variables")
            
            # STEP 3: Generate detailed lineage for selected variables
            if selected_variables:
                if st.button("🚀 Generate Detailed Lineage Analysis", type="primary", use_container_width=True, key="generate_lineage_btn"):
                    
                    # Add debugging
                    # debug_session_state()
                    
                    # st.write(f"**🎯 Selected variables for analysis:** {selected_variables}")
                    
                    with st.spinner(f"🤖 AI generating detailed lineage for {len(selected_variables)} variables..."):
                        lineage_analyzer = DataLineageAnalyzer()
                        detailed_lineage = lineage_analyzer.generate_detailed_variable_lineage(
                            code_files, selected_variables, all_variables
                        )
                    
                    if detailed_lineage:
                        # Store results in session state
                        st.session_state['lineage_detailed_results'] = detailed_lineage
                        st.session_state['lineage_selected_vars'] = selected_variables
                        st.session_state['lineage_analysis_complete'] = True
            
            # STEP 4: Display detailed lineage results (if analysis is complete)
            if st.session_state.get('lineage_analysis_complete', False):
                detailed_lineage = st.session_state.get('lineage_detailed_results', {})
                analysis_selected_vars = st.session_state.get('lineage_selected_vars', [])
                
                if detailed_lineage:
                    st.success(f"✅ Detailed lineage generated for {len(detailed_lineage)} variables!")
                    
                    # Display detailed lineage results
                    st.markdown("## 🗺️ Detailed Variable Lineage Analysis")
                    
                    # Summary of analysis
                    st.markdown("### 📊 Analysis Summary")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Variables Analyzed", len(detailed_lineage))
                    with col2:
                        variables_with_sources = sum(1 for v in detailed_lineage.values() if v.get('source_variables'))
                        st.metric("Variables with Sources", variables_with_sources)
                    with col3:
                        variables_with_calculations = sum(1 for v in detailed_lineage.values() if v.get('calculation'))
                        st.metric("Variables with Calculations", variables_with_calculations)
                    
                    # Detailed lineage for each variable

                    for variable, lineage_info in detailed_lineage.items():
                        with st.expander(f"📊 **{variable}** ({lineage_info.get('category', 'unknown').title()})", expanded=True):
                            
                            # Two-column layout 
                            col1, col2 = st.columns([2, 1])
                            
                            with col1:
                                # Show exact code line 
                                if lineage_info.get('exact_code_line'):
                                    st.markdown(f"**📝 Exact Code Line:**")
                                    st.code(lineage_info['exact_code_line'], language='sas')
                                
                                # Main lineage information
                                if lineage_info.get('description'):
                                    st.markdown(f"**🔍 How it's created:** {lineage_info['description']}")
                                
                                if lineage_info.get('calculation'):
                                    st.markdown(f"**📊 Calculation Formula:** `{lineage_info['calculation']}`")
                                
                                if lineage_info.get('source_variables'):
                                    sources = ", ".join(lineage_info['source_variables'])
                                    st.markdown(f"**📥 Source Variables:** `{sources}`")
                                
                                if lineage_info.get('transformation_type'):
                                    st.markdown(f"**⚙️ Transformation Type:** {lineage_info['transformation_type'].title()}")
                                
                                if lineage_info.get('business_purpose'):
                                    st.markdown(f"**💼 Business Purpose:** {lineage_info['business_purpose']}")
                            
                            with col2:
                                # Metadata and technical details
                                st.markdown("**📋 Technical Details:**")
                                st.write(f"**Category:** {lineage_info.get('category', 'unknown').title()}")
                                if lineage_info.get('code_location'):
                                    st.write(f"**Location:** {lineage_info['code_location']}")
                            
                            # Detailed steps if available
                            if lineage_info.get('detailed_steps'):
                                st.markdown("**🔄 Transformation Steps:**")
                                for i, step in enumerate(lineage_info['detailed_steps'], 1):
                                    st.write(f"{i}. {step}")
                    
                    # Export options
                    st.markdown("### 📥 Export Analysis")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.button("📄 Download Lineage (JSON)", key="download_json_btn"):
                            json_data = {
                                'analysis_type': 'detailed_lineage',
                                'selected_variables': analysis_selected_vars,
                                'all_variables': all_variables,
                                'detailed_lineage': detailed_lineage,
                                'analysis_timestamp': pd.Timestamp.now().isoformat()
                            }
                            
                            json_buffer = BytesIO()
                            json_buffer.write(json.dumps(json_data, indent=2).encode())
                            json_buffer.seek(0)
                            
                            st.download_button(
                                label="📥 Download JSON",
                                data=json_buffer,
                                file_name=f"lineage_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.json",
                                mime="application/json"
                            )
                    
                    with col2:
                        if st.button("📊 Download Summary (CSV)", key="download_csv_btn"):
                            # Create summary CSV
                            summary_data = []
                            for variable, info in detailed_lineage.items():
                                summary_data.append({
                                    'Variable': variable,
                                    'Category': info.get('category', ''),
                                    'Source_Variables': ', '.join(info.get('source_variables', [])),
                                    'Calculation': info.get('calculation', ''),
                                    'Description': info.get('description', ''),
                                    'Business_Purpose': info.get('business_purpose', '')
                                })
                            
                            csv_buffer = BytesIO()
                            pd.DataFrame(summary_data).to_csv(csv_buffer, index=False)
                            csv_buffer.seek(0)
                            
                            st.download_button(
                                label="📥 Download CSV",
                                data=csv_buffer,
                                file_name=f"lineage_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
                                mime="text/csv"
                            )
            else:
                # Show message if no variables selected
                if not selected_variables:
                    st.info("ℹ️ Please select at least one variable to proceed with detailed lineage analysis.")
        
        return all_variables
        
    except Exception as e:
        st.error(f"❌ Error in lineage analysis: {str(e)}")
        with st.expander("🐛 Technical Error Details", expanded=False):
            import traceback
            st.code(traceback.format_exc())
        return {}


def perform_code_review(uploaded_files: list):
    """Execute code review analysis with enhanced semantic batching and threading for better performance."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    
    analyzer = CodeAnalyzer()
    all_analyses = []
    
    # Thread-safe progress tracking
    progress_lock = threading.Lock()
    
    def process_snippet_batch_enhanced(batch_snippets, file_name, file_type, batch_id):
        """Process a batch of 6 semantic snippets in parallel."""
        batch_results = []
        
        # Create combined message for all snippets in the batch
        snippets_data = []
        for idx, (snippet, line_range) in enumerate(batch_snippets, 1):
            snippets_data.append({
                "snippet_id": idx,
                "code": snippet.strip(),
                "lines": f"{line_range[0]}-{line_range[1]}"
            })
        
        combined_message = f"""Analyze these {len(snippets_data)} {file_type} code snippets from batch {batch_id}:

{chr(10).join([f"SNIPPET {s['snippet_id']} (Lines {s['lines']}):```{file_type.lower()}{chr(10)}{s['code']}{chr(10)}```{chr(10)}" for s in snippets_data])}

For each snippet, provide both overview and technical annotation.

Respond in this exact JSON format:
{{
    "analysis_results": [
        {{
            "snippet_id": 1,
            "overview": "Simple explanation in 60 words of what this code does",
            "annotation": "Technical bullet points explaining functions, operations, and data transformations"
        }},
        {{
            "snippet_id": 2,
            "overview": "Simple explanation in 60 words of what this code does", 
            "annotation": "Technical bullet points explaining functions, operations, and data transformations"
        }}
    ]
}}

Instructions:
- Overview: Business-friendly explanation without technical jargon (max 60 words)
- Annotation: Technical details with specific functions, operations, transformations (max 150 words)
- Analyze each snippet separately but return all in one JSON response"""

        try:
            response = analyzer.send_message(combined_message, {
                "type": "batch_snippet_analysis",
                "max_tokens": 60000,  
                "temperature": 0.1,
                "batch_id": batch_id
            })
            
            if response:
                try:
                    # Parse JSON response
                    import json
                    clean_response = response.strip()
                    if clean_response.startswith('```json'):
                        clean_response = clean_response[7:]
                    if clean_response.endswith('```'):
                        clean_response = clean_response[:-3]
                        
                    parsed_response = json.loads(clean_response)
                    
                    # Extract results for each snippet
                    analysis_results = parsed_response.get('analysis_results', [])
                    
                    for snippet_data, (snippet, line_range) in zip(analysis_results, batch_snippets):
                        snippet_id = snippet_data.get('snippet_id', len(batch_results) + 1)
                        overview = snippet_data.get('overview', 'Analysis completed')
                        annotation = snippet_data.get('annotation', 'Technical analysis completed')
                        
                        batch_results.append({
                            "sr_no": len(batch_results) + 1,  # Will be renumbered later
                            "code_snippet": snippet.strip(),
                            "line_numbers": f"Lines {line_range[0]}-{line_range[1]}",
                            "code_annotation": annotation,
                            "code_overview": overview
                        })
                        
                except json.JSONDecodeError:
                    # Fallback: process individually if JSON parsing fails
                    st.warning(f"⚠️ JSON parsing failed for batch {batch_id}, falling back to individual processing")
                    for snippet, line_range in batch_snippets:
                        batch_results.append({
                            "sr_no": len(batch_results) + 1,
                            "code_snippet": snippet.strip(),
                            "line_numbers": f"Lines {line_range[0]}-{line_range[1]}",
                            "code_annotation": "Analysis completed (JSON parse failed)",
                            "code_overview": "Analysis completed (JSON parse failed)"
                        })
            else:
                # Fallback for no response
                for snippet, line_range in batch_snippets:
                    batch_results.append({
                        "sr_no": len(batch_results) + 1,
                        "code_snippet": snippet.strip(),
                        "line_numbers": f"Lines {line_range[0]}-{line_range[1]}",
                        "code_annotation": "Error analyzing this snippet",
                        "code_overview": "Error analyzing this snippet"
                    })
                    
        except Exception as e:
            st.error(f"Error analyzing batch {batch_id} in {file_name}: {str(e)}")
            # Fallback for errors
            for snippet, line_range in batch_snippets:
                batch_results.append({
                    "sr_no": len(batch_results) + 1,
                    "code_snippet": snippet.strip(),
                    "line_numbers": f"Lines {line_range[0]}-{line_range[1]}",
                    "code_annotation": "Error analyzing this snippet",
                    "code_overview": "Error analyzing this snippet"
                })
        
        return batch_results
    
    # Process each file
    for file_idx, file in enumerate(uploaded_files):
        st.markdown(f"### 📄 Analysis for {file.name}")
        
        try:
            content = file.read().decode('utf-8')
            file_type = file.name.split('.')[-1].upper()
            
            # Check file size
            line_count = analyzer._count_lines(content)
            if line_count > MAX_CODE_SIZE_FOR_FULL_ANALYSIS:
                st.warning(f"⚠️ Large file detected ({line_count} lines). Using optimized analysis...")
            
            # Generate summary first (unchanged)
            with st.spinner(f"Generating summary for {file.name}..."):
                summary = analyzer.generate_summary(content, file_type, file.name)
            
            # Display summary
            with st.expander("📊 Code Summary", expanded=True):
                if "PURPOSE:" in summary:
                    lines = summary.split('\n')
                    for line in lines:
                        if line.startswith("PURPOSE:"):
                            st.markdown(f"**🎯 Purpose:** {line.replace('PURPOSE:', '').strip()}")
                        elif line.startswith("INPUTS:"):
                            st.markdown(f"**📥 Inputs:** {line.replace('INPUTS:', '').strip()}")
                        elif line.startswith("PROCESS:"):
                            st.markdown(f"**⚙️ Process:** {line.replace('PROCESS:', '').strip()}")
                        elif line.startswith("OUTPUT:"):
                            st.markdown(f"**📤 Output:** {line.replace('OUTPUT:', '').strip()}")
                        elif line.startswith("VALUE:"):
                            st.markdown(f"**💎 Business Value:** {line.replace('VALUE:', '').strip()}")
                else:
                    st.markdown(summary)
            
            # Analyze snippets with enhanced semantic batching
            st.markdown("#### 🔍 Detailed Analysis")
            
            if line_count > MAX_CODE_SIZE_FOR_FULL_ANALYSIS:
                analyzed_snippets = analyzer._analyze_large_code(content, file_type, file.name)
            else:
                # Use new semantic batching approach
                semantic_batches = analyzer._create_code_batches(content, file_type)
                
                if len(semantic_batches) <= 2:
                    # Small number of batches - process normally
                    total_snippets = sum(len(batch[0]) for batch in semantic_batches)
                    with st.spinner(f"Analyzing {total_snippets} code sections in {len(semantic_batches)} batches..."):
                        analyzed_snippets = []
                        for batch_snippets, batch_id in semantic_batches:
                            batch_result = process_snippet_batch_enhanced(batch_snippets, file.name, file_type, batch_id)
                            analyzed_snippets.extend(batch_result)
                else:
                    # Use threading for larger number of batches
                    total_snippets = sum(len(batch[0]) for batch in semantic_batches)
                    st.info(f"🚀 Processing {total_snippets} code sections in {len(semantic_batches)} batches with parallel analysis...")
                    
                    # Progress tracking
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    analyzed_snippets = []
                    completed_batches = 0
                    
                    # Process batches in parallel
                    with ThreadPoolExecutor(max_workers=5) as executor:
                        # Submit all batch jobs
                        future_to_batch = {
                            executor.submit(process_snippet_batch_enhanced, batch_snippets, file.name, file_type, batch_id): batch_id 
                            for batch_snippets, batch_id in semantic_batches
                        }
                        
                        # Collect results as they complete
                        for future in as_completed(future_to_batch):
                            try:
                                batch_results = future.result()
                                analyzed_snippets.extend(batch_results)
                                
                                # Update progress
                                completed_batches += 1
                                progress = completed_batches / len(semantic_batches)
                                progress_bar.progress(progress)
                                status_text.text(f"Completed {completed_batches}/{len(semantic_batches)} semantic batches")
                                
                            except Exception as e:
                                st.error(f"Error processing semantic batch: {str(e)}")
                    
                    # Clear progress indicators
                    progress_bar.empty()
                    status_text.empty()
                    
                    st.success(f"✅ Completed semantic analysis of {len(analyzed_snippets)} code sections!")
            
            if analyzed_snippets:
                # SORT RESULTS BY LINE NUMBER (for threading order issue)
                def extract_start_line(snippet):
                    """Extract starting line number from 'Lines X-Y' format."""
                    try:
                        line_numbers = snippet.get("line_numbers", "Lines 0-0")
                        # Extract first number from "Lines X-Y" format
                        start_line = int(line_numbers.split()[1].split('-')[0])
                        return start_line
                    except (ValueError, IndexError):
                        return 0  # Fallback for malformed line numbers
                
                # Sort snippets by starting line number
                analyzed_snippets.sort(key=extract_start_line)
                
                # Show semantic analysis info
                snippet_info = f"📋 Analyzed {len(analyzed_snippets)} semantic code sections"
                if len(semantic_batches) > 1:
                    snippet_info += f" from {len(semantic_batches)} batches"
                st.info(snippet_info)
                
                # Renumber snippets sequentially after sorting
                for i, snippet in enumerate(analyzed_snippets, 1):
                    snippet["sr_no"] = i
                
                df = pd.DataFrame(analyzed_snippets)
                
                # Add filters for large files
                if len(df) > 10:
                    search_term = st.text_input(
                        "🔍 Search explanations:", 
                        key=f"search_{file_idx}",
                        placeholder="Filter by keywords..."
                    )
                    
                    if search_term:
                        mask = (df['code_annotation'].str.contains(search_term, case=False, na=False) | 
                               df['code_overview'].str.contains(search_term, case=False, na=False))
                        df = df[mask]
                        st.info(f"Showing {len(df)} results matching '{search_term}'")
                
                # Display with enhanced formatting
                st.dataframe(
                    df,
                    use_container_width=True,
                    column_config={
                        "sr_no": st.column_config.NumberColumn("No.", width="small"),
                        "code_snippet": st.column_config.TextColumn("Code", width="large"),
                        "line_numbers": st.column_config.TextColumn("Lines", width="small"),
                        "code_annotation": st.column_config.TextColumn("Technical Annotation", width="large"),
                        "code_overview": st.column_config.TextColumn("Code Overview", width="large")
                    }
                )
                
                all_analyses.append({
                    'file_name': file.name,
                    'summary': summary,
                    'analysis': df,
                    'line_count': line_count
                })
                
        except Exception as e:
            st.error(f"Error analyzing {file.name}: {str(e)}")
            st.info("Tip: Make sure the file is encoded in UTF-8 format")
    
    # Create Excel report 
    if all_analyses:
        st.markdown("### 📊 Generate Report")
        
        col1, col2 = st.columns(2)
        with col1:
            include_code = st.checkbox("Include code snippets in report", value=True)
        with col2:
            exec_summary_only = st.checkbox("Executive summary only", value=False)
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet
            summaries_data = []
            for i, analysis in enumerate(all_analyses):
                summaries_data.append({
                    'Script': f"{analysis['file_name']}",
                    'Lines of Code': analysis['line_count'],
                    'Code Summary': analysis['summary']
                })
            
            summary_df = pd.DataFrame(summaries_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            summary_sheet_name = 'Summary'
            worksheet = writer.sheets[summary_sheet_name]
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 70
            
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(
                        wrap_text=True,
                        vertical='top'
                    )
            
            # Analysis sheets for each script
            if not exec_summary_only:
                for i, analysis in enumerate(all_analyses, 1):
                    sheet_name = f'Script{i}_Analysis'[:31]
                    
                    if include_code:
                        df_to_export = analysis['analysis']
                    else:
                        df_to_export = analysis['analysis'][['sr_no', 'line_numbers', 'code_annotation', 'code_overview']]
                    
                    df_to_export.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    worksheet = writer.sheets[sheet_name]
                    worksheet.column_dimensions['A'].width = 10
                    
                    if include_code:
                        worksheet.column_dimensions['B'].width = 60
                        worksheet.column_dimensions['C'].width = 15
                        worksheet.column_dimensions['D'].width = 70
                        worksheet.column_dimensions['E'].width = 50
                    else:
                        worksheet.column_dimensions['B'].width = 15
                        worksheet.column_dimensions['C'].width = 70
                        worksheet.column_dimensions['D'].width = 50
                    
                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = openpyxl.styles.Alignment(
                                wrap_text=True,
                                vertical='top'
                            )
        
        buffer.seek(0)
        
        filename = f"code_analysis_report.xlsx"
        
        st.download_button(
            label="📥 Download Complete Report",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download the complete report in Excel format"
        )
        
        # Enhanced summary metrics
        st.markdown("### 📈 Analysis Summary")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_files = len(all_analyses)
            st.metric("Files Analyzed", total_files)
        
        with col2:
            total_lines = sum(a['line_count'] for a in all_analyses)
            st.metric("Total Lines of Code", f"{total_lines:,}")
        
        with col3:
            total_snippets = sum(len(a['analysis']) for a in all_analyses)
            st.metric("Semantic Sections", total_snippets)
            
        with col4:
            avg_snippet_size = total_lines / total_snippets if total_snippets > 0 else 0
            st.metric("Avg Section Size", f"{avg_snippet_size:.1f} lines")



def perform_rde_testing(code_files: list, data_files: list):
    """Perform AI-powered RDE testing."""
    rde_tester = RDETester()
    
    # Progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # STEP 1: Extract input dataset variables 
    status_text.text("🔍 Analyzing input datasets (headers only)...")
    input_datasets_info, all_input_vars = rde_tester._extract_input_dataset_variables(data_files)
    
    st.markdown("### 📋 Input Dataset Variables")
    input_summary = []
    for file_name, variables in input_datasets_info.items():
        input_summary.append({
            'Dataset': file_name,
            'Variable Count': len(variables),
            'Sample Variables': ', '.join(list(variables)[:5]) + ('...' if len(variables) > 5 else '')
        })
    
    if input_summary:
        st.dataframe(pd.DataFrame(input_summary), use_container_width=True)
        st.info(f"🔒 **Data Privacy**: Only variable names extracted - no actual data values accessed")
    
    progress_bar.progress(0.15)
    
    # STEP 2: Use AI to analyze script-created variables and cross-script flows
    status_text.text("🤖 Using AI to analyze script variables and cross-script flows...")
    script_created_info = rde_tester._extract_script_created_variables(code_files)
    
    st.markdown("### 🤖 AI-Analyzed Script Variables")
    
    # Display script-created variables
    if script_created_info['created_by_script']:
        script_vars_df = []
        for script_name, variables in script_created_info['created_by_script'].items():
            script_vars_df.append({
                'Script': script_name,
                'Created Variables': len(variables),
                'Variables': ', '.join(variables[:3]) + ('...' if len(variables) > 3 else '')
            })
        st.dataframe(pd.DataFrame(script_vars_df), use_container_width=True)
    else:
        st.info("No script-created variables detected by AI")
    
    # Display cross-script flows
    if script_created_info['cross_script_flows']:
        st.markdown("#### 🔄 Cross-Script Variable Flows (AI-Detected)")
        flow_df = []
        for var_name, flow_info in script_created_info['cross_script_flows'].items():
            flow_df.append({
                'Variable': var_name,
                'Flow Pattern': flow_info,
                'Classification': 'Recoded (Cross-Script)'
            })
        st.dataframe(pd.DataFrame(flow_df), use_container_width=True)
        st.success(f"✅ AI detected {len(script_created_info['cross_script_flows'])} cross-script variable flows")
    else:
        st.info("No cross-script variable flows detected")
    
    progress_bar.progress(0.3)
    
    # STEP 3: Read code content for context 
    status_text.text("📖 Reading code files for context...")
    code_content = ""
    code_info = []
    
    for idx, file in enumerate(code_files):
        content = file.read().decode('utf-8')
        code_content += f"\n\n# --- File: {file.name} ---\n{content}"
        code_info.append({
            'filename': file.name,
            'lines': len(content.split('\n'))
        })
        progress_bar.progress(0.3 + (idx + 1) / len(code_files) * 0.1)
    
    # Display code info
    st.markdown("### 📝 Code Files Summary")
    st.dataframe(pd.DataFrame(code_info), use_container_width=True)
    
    # STEP 4: Process datasets for classification
    datasets_info = []
    all_analyses = []
    
    st.markdown("### 🚀 AI-Powered Dataset Analysis")
    
    for idx, file in enumerate(data_files, 1):
        status_text.text(f"🚀 Analysis of {file.name}...")
        progress_bar.progress(0.4 + (idx / len(data_files) * 0.4))
        
        # Read dataset
        df = rde_tester.read_dataset(file)
        
        if df is not None:
            # Create expander for each dataset
            with st.expander(f"📁 {file.name} - {len(df):,} rows, {len(df.columns)} columns", expanded=True):
                
                # Dataset quality analysis
                quality_report = rde_tester.analyze_dataset_quality(df)
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Quality Score", f"{quality_report['quality_score']:.1f}%")
                with col2:
                    st.metric("Missing Values", f"{len(quality_report['missing_values'])} columns")
                with col3:
                    st.metric("Duplicate Rows", f"{quality_report['duplicate_rows']:,}")
                
                # Store dataset info
                datasets_info.append({
                    "ID": idx,
                    "Input Datasets": file.name,
                    "Number of Rows": len(df),
                    "Number of Columns": len(df.columns),
                    "Quality Score": f"{quality_report['quality_score']:.1f}%"
                })
                
                # ⚡ AI-powered variable analysis with batch processing
                st.write("⚡ **AI Analysis** - Processing variables in batches...")
                
                # Use the optimized analysis method
                variables_analysis = rde_tester.analyze_dataset_variables_optimized(
                    df, code_content, input_datasets_info, script_created_info
                )
                
                
                # Show sample data 
                st.write("**📊 Sample Data**")
                samples = rde_tester.get_intelligent_samples(df, n_samples=5)
                st.dataframe(samples.head(5), use_container_width=True)
                
                all_analyses.append({
                    "id": idx,
                    "variables": pd.DataFrame(variables_analysis),
                    "samples": rde_tester.get_intelligent_samples(df, n_samples=65),
                    "quality": quality_report
                })
    
    progress_bar.progress(0.8)
    status_text.text("📋 Creating AI report...")
    
    # Create Excel report with AI insights
    if all_analyses:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet with AI insights
            summary_data = datasets_info.copy()
            
            # Add AI insights to summary
            total_input_vars = len(all_input_vars)
            total_script_vars = len(script_created_info['all_created_vars'])
            total_cross_script = len(script_created_info['cross_script_flows'])
            
            summary_data.append({
                "ID": "AI Analysis",
                "Input Datasets": f"Input Variables: {total_input_vars}",
                "Number of Rows": f"Script Variables: {total_script_vars}",
                "Number of Columns": f"Cross-Script: {total_cross_script}",
                "Quality Score": "OPTIMIZED AI-Powered"
            })
            
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary
            worksheet = writer.sheets['Summary']
            for col in ['A', 'B', 'C', 'D', 'E']:
                worksheet.column_dimensions[col].width = 25
            
            # Add AI insights sheet
            ai_insights_data = []
            
            # Input dataset insights
            for dataset, vars_set in input_datasets_info.items():
                ai_insights_data.append({
                    'Type': 'Input Dataset',
                    'Source': dataset,
                    'Variable_Count': len(vars_set),
                    'Details': 'Original data elements for testing'
                })
            
            # Script created insights
            for script, vars_list in script_created_info['created_by_script'].items():
                ai_insights_data.append({
                    'Type': 'Script Created',
                    'Source': script,
                    'Variable_Count': len(vars_list),
                    'Details': 'Variables created within script'
                })
            
            # Cross-script flow insights
            for var, flow in script_created_info['cross_script_flows'].items():
                ai_insights_data.append({
                    'Type': 'Cross-Script Flow',
                    'Source': var,
                    'Variable_Count': 1,
                    'Details': flow
                })
            
            if ai_insights_data:
                pd.DataFrame(ai_insights_data).to_excel(writer, sheet_name='AI_Insights', index=False)
            
            # Analysis and sample sheets for each dataset
            for analysis in all_analyses:
                # Variables analysis
                sheet_name = str(analysis["id"])
                analysis["variables"].to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatting with AI-based color coding
                worksheet = writer.sheets[sheet_name]
                worksheet.column_dimensions['A'].width = 30  # DataID
                worksheet.column_dimensions['B'].width = 25  # Classification
                worksheet.column_dimensions['C'].width = 60  # Description
                worksheet.column_dimensions['D'].width = 80  # Screenshot
                
                # Conditional formatting
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                    classification = row[1].value
                    if classification == "Data element to test":
                        row[1].fill = openpyxl.styles.PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        row[1].font = openpyxl.styles.Font(bold=True)
                    elif classification == "Recoded":
                        row[1].fill = openpyxl.styles.PatternFill(
                            start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    elif classification == "Unique Identifier":
                        row[1].fill = openpyxl.styles.PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif classification == "Dates":
                        row[1].fill = openpyxl.styles.PatternFill(
                            start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                
                # Samples sheet
                sample_sheet = f"{analysis['id']}-Sample"
                analysis["samples"].to_excel(writer, sheet_name=sample_sheet, index=False)
                
                # Format samples sheet
                worksheet = writer.sheets[sample_sheet]
                for col in worksheet.column_dimensions.keys():
                    worksheet.column_dimensions[col].width = 20
        
        buffer.seek(0)
        
        progress_bar.progress(1.0)
        status_text.text("⚡ AI RDE report ready!")
        
        # Download button 
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="📥 Download AI RDE Identification Report",
            data=buffer,
            file_name=f"ai_rde_report_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download the AI-powered RDE testing report with intelligent variable classification"
        )
        
        # AI summary statistics
        st.markdown("### ⚡ AI Testing Summary")
        
        total_vars = sum(len(a['variables']) for a in all_analyses)
        rde_count = sum((a['variables']['Classification'] == 'Data element to test').sum() for a in all_analyses)
        recoded_count = sum((a['variables']['Classification'] == 'Recoded').sum() for a in all_analyses)
        ai_detected_flows = len(script_created_info.get('cross_script_flows', {}))
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Variables", total_vars, delta="⚡ OPTIMIZED")
        with col2:
            st.metric("RDE to Test", rde_count, delta="From Input Data")
        with col3:
            st.metric("Recoded Variables", recoded_count, delta="Script Created")
        with col4:
            st.metric("Cross-Script Flows", ai_detected_flows, delta="AI Detected")
        
        # Performance and privacy confirmation
        st.success("✅ AI RDE Identification completed successfully!")
        
        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()





def main():
    """Main application function with enhanced features."""
    
    st.set_page_config(
        page_title="CodeIQ - AI Powered Code Analysis Assistant", 
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Custom CSS for better UI
    st.markdown("""
    <style>
    .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
        font-size: 18px;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
    }
    .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] {
        background-color: #4CAF50;
    }
    div[data-testid="metric-container"] {
        background-color: rgba(28, 131, 225, 0.1);
        border: 1px solid rgba(28, 131, 225, 0.2);
        padding: 10px;
        border-radius: 5px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.title("🧠 Code IQ - AI Powered Code Analysis Assistant")
    st.markdown("*Powered by AVA*")
    
    # Sidebar for information
    with st.sidebar:
        st.header("ℹ️ About This Tool")
        st.markdown("""
        This AI-powered tool helps you:
        - 📝 **Review Code**: Get detailed and business-friendly explanations
        - 🔍 **Test RDE**: Identify data elements for testing
        - 🗺️ **Trace Data Lineage**: Understand and Visualize data flow
        
        All powered by advanced AI for accurate, intuitive results!
        """)
        
        st.header("🚀 What's New")
        st.markdown("""
        - **AI-Powered Lineage**: Intelligent data flow analysis
        - **Large File Support**: Optimized for large files
        - **Business Language**: Non-technical explanations
        - **Interactive Visuals**: Explore data relationships
        """)
    
    # Create tabs for different functionalities
    tab1, tab2, tab3 = st.tabs(["📝 Code Review", "🔍 RDE Identification", "🗺️ Data Lineage"])
    
    # Code Review Tab
    with tab1:
        st.header("Code Review")
        st.markdown("Upload your code files to get AI-powered, business-friendly explanations.")
        
        code_files = st.file_uploader(
            "Upload code files for review",
            accept_multiple_files=True,
            type=['py', 'r', 'sas', 'sql', 'cpp', 'cc', 'cxx', 'java', 'js', 'txt'],
            key="code_review_files",
            help="Supports Python, R, SAS, SQL, C++, Java, JavaScript"
        )
        
        if code_files:
            # Show file info
            st.info(f"📁 {len(code_files)} file(s) uploaded")
            
            # Check for large files
            large_files = []
            for file in code_files:
                content = file.read().decode('utf-8')
                line_count = len(content.split('\n'))
                if line_count > MAX_CODE_SIZE_FOR_FULL_ANALYSIS:
                    large_files.append(f"{file.name} ({line_count} lines)")
                file.seek(0)  # Reset file pointer
            
            if large_files:
                st.warning(f"⚠️ Large files detected: {', '.join(large_files)}. Analysis will be optimized for performance.")
            
            if st.button("🚀 Execute Code Review", type="primary"):
                with st.spinner("Analyzing code... This may take a few minutes."):
                    perform_code_review(code_files)
    
    # RDE Identification Tab
    with tab2:
        st.header("RDE Identification")
        st.markdown("Upload code scripts and datasets to identify raw data elements for testing.")
        
        col1, col2 = st.columns(2)
        
        with col1:
            rde_code_files = st.file_uploader(
                "Upload Code Scripts",
                accept_multiple_files=True,
                type=['py', 'r', 'sas', 'sql'],
                key="rde_code_files",
                help="Upload the code that processes your data"
            )
        
        with col2:
            rde_data_files = st.file_uploader(
                "Upload Input Datasets",
                accept_multiple_files=True,
                type=['xlsx', 'csv', 'sas7bdat', 'parquet'],
                key="rde_data_files",
                help="Upload the data files referenced in your code"
            )
        
        if rde_code_files and rde_data_files:
            st.info(f"📁 {len(rde_code_files)} code file(s) and {len(rde_data_files)} data file(s) uploaded")
            
            if st.button("🚀 Execute RDE Identification", type="primary"):
                with st.spinner("Performing RDE Identification..."):
                    perform_rde_testing(rde_code_files, rde_data_files)
    

    # Data Lineage Tab
    with tab3:
        st.header("🗺️ AI-Powered Data Lineage Analysis")
        st.markdown("Upload your scripts, select variables, and get detailed AI-powered lineage analysis.")
        
        # File upload section
        st.markdown("### 📁 Upload Code Scripts")
        lineage_code_files = st.file_uploader(
            "Choose Code Files",
            accept_multiple_files=True,
            type=['py', 'r', 'sas', 'sql', 'txt'],
            key="lineage_code_files",
            help="Upload Python, R, SAS, SQL, or text files containing your code"
        )
        
        if lineage_code_files:
            # Show upload summary
            st.markdown("### 📋 Upload Summary")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("📄 Files Uploaded", len(lineage_code_files))
            with col2:
                file_types = list(set([f.name.split('.')[-1].upper() for f in lineage_code_files]))
                st.metric("🏷️ File Types", len(file_types))
            with col3:
                total_size = sum(len(f.read()) for f in lineage_code_files)
                for f in lineage_code_files:
                    f.seek(0)  # Reset file pointers
                st.metric("💾 Total Size", f"{total_size/1024:.1f} KB")
            
            # Show file details
            with st.expander("📁 File Details", expanded=False):
                file_details = []
                for file in lineage_code_files:
                    content = file.read().decode('utf-8')
                    file_details.append({
                        'File Name': file.name,
                        'Type': file.name.split('.')[-1].upper(),
                        'Lines': len(content.split('\n')),
                        'Size (KB)': f"{len(content)/1024:.1f}"
                    })
                    file.seek(0)  # Reset
                
                st.dataframe(pd.DataFrame(file_details), use_container_width=True)
            
            # Analysis workflow explanation
            st.markdown("### 🎯 Start AI Data Lineage")
            st.info("""
            **Step 1:** 🤖 AI discovers all variables in your scripts  
            **Step 2:** 🎯 You select which variables you want lineage for or for all variables
            **Step 3:** 🗺️ AI generates detailed lineage with calculations and source variables  
            **Step 4:** 📊 View detailed visualizations, transformation explanations and export results
            """)
            
            # Check if analysis has already been started
            if st.session_state.get('lineage_variables_discovered', False):
                # Analysis is in progress or complete
                perform_data_lineage_analysis(lineage_code_files, [], None)
            else:
                # Start analysis button
                if st.button("🚀 Generate", type="primary", use_container_width=True):
                    try:
                        perform_data_lineage_analysis(lineage_code_files, [], None)
                    except Exception as e:
                        st.error(f"❌ Error during analysis: {str(e)}")
                        st.info("💡 Please check your code syntax and try again.")
        
        else:
            # Welcome section when no files uploaded
            with st.expander("How it works", expanded=False):
                st.markdown("""
            
                    **🤖 AI-Powered Variable Discovery:**
                    - Automatically finds ALL variables in your code
                    - Categorizes them as Input, Created, Intermediate.
                    - Understands multiple programming languages
                    
                    **🎯 User-Controlled Selection:**
                    - Choose specific variables for lineage analysis
                    - Filter by variable type or search by name
                    - Focus on what matters most to you
                    
                    **🗺️ Detailed Lineage Analysis:**
                    - Shows exact calculations and formulas
                    - Identifies source variables for each transformation
                    - Explains business purpose in plain English
                    - Traces complete transformation steps
                    
                    ### 🚀 Supported Languages
                    - **SAS**: Data steps, PROC statements, variable assignments
                    - **Python**: Pandas operations, variable assignments, calculations
                    - **R**: Data.frame operations, dplyr transformations, assignments
                    - **SQL**: SELECT statements, calculated fields, table operations
                    """)
    
    # Footer
    st.markdown("---")
    st.markdown(
        """
        <div style='text-align: center; color: #666;'>
            🤖 <strong>AI-Powered Discovery</strong> | 🎯 <strong>Code Analysis</strong> | 
            🗺️ <strong>Detailed Lineage</strong> | 📊 <strong>Export Ready</strong>
        </div>
        """, 
        unsafe_allow_html=True
    )
    
    # Footer
    st.markdown("---")
    st.markdown(
        """
        <div style='text-align: center'>
            <p>Powered by AVA AI | KPMG © 2024 | 
            <a href='#'>Documentation</a> | 
            <a href='#'>Support</a></p>
        </div>
        """, 
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
